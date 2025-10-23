# -*- coding: utf-8 -*-
import ckan.plugins.toolkit as toolkit, dateutil
from datetime import datetime, timedelta
from logging import getLogger
import ckan.lib.uploader as uploader
# from ckan.common import config
# from six import text_type
from sqlalchemy import Table, select, join, func, and_
from sqlalchemy.sql import text
# # import ckan.plugins as p
# import ckan.model as model
from ckan.model import Session, User, user_table, Package,Resource, Group, PackageExtra
import json
import re
import csv
from goodtables import validate
import pandas as pd
from pandas import read_excel
# from tempfile import TemporaryFile
import math
import tempfile
import hashlib
import chardet
import os.path
import requests
import mimetypes
import openpyxl
from openpyxl import load_workbook
import xlrd
import io
from io import BytesIO
from io import StringIO
from six import string_types
import time
from functools import reduce
import numpy as np
from io import StringIO
from collections import Counter
from urllib.parse import urlparse
from ckanext.opendquality.model import (
    DataQualityMetrics as DataQualityMetricsModel
)
from ckan.plugins.toolkit import config
from ckan.model import Session, Package, Group
import ckan.logic as logic
from ckan import model
from typing import List, Dict, Any, Tuple
# from frictionless import Schema

log = getLogger(__name__)
# cache_enabled = p.toolkit.asbool(
#     config.get('ckanext.stats.cache_enabled', False)
# )

# if cache_enabled:
#     log.warn(
#         'ckanext.quality does not support caching in current implementations'
#     )

DATE_FORMAT = '%Y-%m-%d'
MAX_CONTENT_LENGTH = int(1e9)
MAX_EXCERPT_LINES = int(0)
CHUNK_SIZE = 16 * 1024  # 16kb
DOWNLOAD_TIMEOUT = 180
SSL_VERIFY = True

class LazyStreamingList(object):
    '''Implements a buffered stream that emulates an iterable object.

    It is used to fetch large data in chunks (pages/buffers). The fetch is
    done using a buffer of predefined size. While the iterator iterates over
    the buffer, when the end of the buffer is reached, a call is made to fetch
    the next buffer before iterating to the next record.

    :param fetch_page: `function`, a function used to fetch the next buffer
        data. The function prototype is:

            .. code-block: python

                def fetch_page(page, page_size):
                    return {
                        'total': 100,
                        'records': [...],
                    }

        The function takes two parameters:
            * `page`, `int`, the page number starting from 0
            * `page_size`, `int`, the number of items per page. Default is 512.

        The function must return a dict containing the records. The dict must
        have the following entries:
            * `total` - total number of items
            * `records` - a list (iterable) over the records fetched for this
                page.
    :param page_size: `int`, the size of each page (buffer). Default is 512.
    '''

    def __init__(self, fetch_page, page_size=512):
        self.fetch_page = fetch_page
        self.page_size = page_size
        self.page = 0
        self.buffer = None
        self.total = 0
    def _fetch_buffer(self):
        if self.total:
            if self.page*self.page_size >= self.total:
                self.buffer = []
                return
        result = self.fetch_page(self.page, self.page_size)
        self.total = result.get('total', 0)
        self.buffer = result.get('records', [])
        self.page += 1

    # def iterator(self):
    #     '''Returns an iterator over all of the records.
    #     '''
    #     current = 0
    #     self._fetch_buffer()
    #     while True:
    #         if current >= self.total or not self.buffer:
    #             # end of all results
    #             raise StopIteration()
    #         for row in self.buffer:
    #             current += 1
    #             yield row
    #         # fetch next buffer
    #         self._fetch_buffer()
    def iterator(self):
        '''Returns an iterator over all of the records.
        '''
        current = 0
        self._fetch_buffer()
        if current >= self.total:
            log.debug('end of all results')
        if self.buffer:
            for row in self.buffer:
                current += 1
                yield row
            self._fetch_buffer()

    def __iter__(self):
        return self.iterator()

    def rewind(self):
        self.buffer = None
        self.page = 0
        self.total = 0
class ResourceCSVData(object):
    '''Represents a CSV data source that can be read with pagination.

    It fetches the CSV data (for a resource) and generates metadata related to
    the CSV file data.
    The column names and data types for each column are calculated:
        * column names are calculated from the firs row in the CSV file. If
            there are multiple columns with the same name, then the column
            names are suffixed with index `_<i>`, for example:
                - if we have the following columns: `a`, `b`, `a`, `a`, then
                the generated columns will have the following names: `a_1`,
                `b`, `a_2` and `a_3`.
        * the data types are guessed from the first row with data (second row
        in the file).

    The data is rerieved in pages. Each page is a `dict` containing:
        * `total` - total number of data rows in the CSV file (the header row
            is not counted)
        * `records` - the records in the requested page - a `list` of `dict`
        * `fields` - `list` of `dict` describing the columns in the CSV file.

    The constructor takes one argument - the data loader for the CSV. This is a
    function (callable) that is called without arguments to load the actual
    data from the CSV file. The expected output of this function is a `list`
    of rows where each row is itself a `list` of values.
    '''

    def __init__(self, resource_loader):
        csv_data = resource_loader()
        self.data = csv_data
        self.fields = self._get_field_types(self._get_fields(csv_data),
                                            csv_data)
        self.column_names = [f['id'] for f in self.fields]
        self.total = 0
        if len(csv_data):
            self.total = len(csv_data) - 1
        log.debug('-------------ResourceCSVData----------')
        # log.debug('%s, Resource CSV data. Total: %d, columns=%s, fields=%s',
        #           str(self), self.total, self.column_names, self.fields)

    def _get_fields(self, csv_data):
        if not csv_data:
            return {}
        columns = csv_data[0]
        columns_count = {}
        for column in columns:
            columns_count[column] = columns_count.get(column, 0) + 1
        numbered = {}
        fields = []
        for column in columns:
            if columns_count[column] > 1:
                numbered[column] = numbered.get(column, 0) + 1
                column = '{}_{}'.format(column, numbered[column])
            fields.append({'id': column})
        return fields

    def _get_field_types(self, fields, csv_data):
        if not csv_data or len(csv_data) == 1:
            for field in fields:
                field['type'] = 'text'
            return fields
        row = csv_data[1]
        for i, field in enumerate(fields):
            field['type'] = self._guess_type(row[i])
        return fields

    def _guess_type(self, value):
        if value is None:
            return 'text'
        if detect_numeric_format(value):
            return 'numeric'
        if detect_date_format(value):
            return 'timestamp'
        return 'text'

    def fetch_page(self, page, limit):
        '''Retrieves one page from the CSV data.

        :param page: `int`, the number of the page to fetch, 0-based.
        :param limit: `int`, the page size.

        :returns: the data page, which is a `dict` containing:
            * `total` - total number of records.
            * `records` - `list` of `dict`, the row (records) in this page.
            * `fields` - `list` of `dict`, metadata for the columns.
        '''
        start = max(page, 0) * limit + 1
        limit = min(start + limit, self.total + 1)
        items = []
        if start < (self.total + 1):
            for i in range(start, limit):
                data_row = self.data[i]
                row_dict = {self.column_names[j]: value
                            for j, value in enumerate(data_row)}
                items.append(row_dict)
        # log.debug('ResourceCSVData.fetch_page: '
        #           'page=%d, limit=%d, of total %d. Got %d results.',
        #           page, limit, self.total, len(items))
        return {
            'total': self.total,
            'records': items,
            'fields': self.fields,
        }

class DataQualityMetrics(object):
 
    def __init__(self, metrics=None, force_recalculate=False):
        self.metrics = metrics or []
        self.force_recalculate = force_recalculate
        self.logger = getLogger('ckanext.DataQualityMetrics')

    def _fetch_dataset(self, package_id):
        return toolkit.get_action('package_show')(
            {
                'ignore_auth': True,
            },
            {
                'id': package_id,
            }
        )
    
    def _data_quality_settings(self, resource):
        settings = {}
        #,'machine_readable'
        for dimension in ['completeness', 'uniqueness','validity', 'consistency','openness','availablity','downloadable', 'timeliness','relevance','utf8','preview']:
            for key, value in resource.items():
                prefix = 'dq_%s' % dimension
                if key.startswith(prefix):
                    if settings.get(dimension) is None:
                        settings[dimension] = {}
                    setting = key[len(prefix) + 1:]
                    settings[dimension][setting] = value
        return settings
    def _fetch_resource_data(self, resource):
        def _fetch_page(page, page_size):
            try:
                context = {'ignore_auth': True}
                return toolkit.get_action('datastore_search')(None, {
                    'resource_id': resource['id'],
                    'offset': page*page_size,
                    'limit': page_size,
                })
            except Exception as e:
                self.logger.warning('Failed to fetch data for resource %s. '
                                    'Error: %s', resource['id'], str(e))
                self.logger.exception(e)
                return {
                    'total': 0,
                    'records': [],
                    'error': 'Failed to fetch data for resource'
                }
       

        _fetch_page = ResourceFetchData(resource)

        result = _fetch_page(0, 1)  # to calculate the total from start
        return {
            'total': result.get('total', 0),
            'records': LazyStreamingList(_fetch_page),
            'fields': result['fields'],  # metadata
        }
    def _fetch_resource_data2(self, resource):
        def _fetch_page2(page, page_size):
            try:
                context = {'ignore_auth': True}
                return toolkit.get_action('datastore_search')(None, {
                    'resource_id': resource['id'],
                    'offset': page*page_size,
                    'limit': page_size,
                })
            except Exception as e:
                self.logger.warning('Failed to fetch data for resource %s. '
                                    'Error: %s', resource['id'], str(e))
                self.logger.exception(e)
                return {
                    'total': 0,
                    'fields': [],
                    'records': [],
                    'raw_data':[],
                    'mimetype':'',
                    'error': 'Failed to fetch data for resource'
                }
       

        _fetch_page2 = ResourceFetchData2(resource)
        result = _fetch_page2(0, 1)  # to calculate the total from start       
        raw_data = _fetch_page2._download_resource_from_url(resource['url'],resource['format'])
        mimetype = _fetch_page2.detect_mimetype(resource['url'])
        if 'error' in result:
            return {
                'total': 0,
                'records': [],
                'fields': [],
                'raw_data':[],
                'mimetype':'',
                'error': result['error']
            }
        return {
            'total': result.get('total', 0),
            'records': LazyStreamingList(_fetch_page2),
            'raw_data': raw_data,
            'mimetype':mimetype,
            'fields': result['fields'],  # metadata
            'error': ''
        }
    def _fetch_resource_file(self, resource):
        _fetch_page = ResourceFetchData(resource)

        result = _fetch_page(0, 1)  # to calculate the total from start
        return {
            'total': result.get('total', 0),
            'records': LazyStreamingList(_fetch_page),
            'fields': result['fields'],  # metadata
        }
    def calculate_metrics_for_dataset(self, package_id, job_id=None):
        '''Calculates the Data Qualtity for the given dataset identified with
        `package_id`.

        The metrics for the resources and dataset are calculated or reused from
        an earlier calculation if there were no changes in the resources data
        from the last time the calculation has been performed.

        Additionaly, if some of the metrics were set manually, either for a
        resource or the dataset, then the calculation for that dimension will
        not be performed and the manual value will be kept intact.

        The calculations for both resources and dataset are kept in a database,
        one entry for dataset and one for each resource.

        :param package_id: `str`, the ID of the dataset (package) for which to
            calculate Data Quality metrics.
        '''
        log.debug('Calculating data quality for dataset: %s',
                          package_id)
        dataset = self._fetch_dataset(package_id)

        results = []
        for resource in dataset['resources']:
            self.logger.debug ('Calculating data quality for resource: %s',
                              resource['id'])
            # self.logger.debug ('Calculating data quality for resource: %s',
            #                   resource)
            resource['data_quality_settings'] = self._data_quality_settings(
                resource)
       
            result = self.calculate_metrics_for_resource(resource, job_id=job_id)
            if result is None:
                result = {}
            # self.logger.debug('Result: %s', result)
            results.append(result)

        # calculate cumulative
        self.logger.debug('Calculating cumulative data quality metrics for %s',
                          package_id)
        self.calculate_cumulative_metrics(package_id,
                                          dataset['resources'],
                                          results,
                                          job_id=job_id)
        self.logger.info('Data Quality metrcis calculated for dataset: %s.',
                         package_id)
        
    def _get_metrics_record(self, ref_type, ref_id):
        metrics = DataQualityMetricsModel.get(ref_type, ref_id)
        return metrics
    def _new_metrics_record(self, ref_type, ref_id):
        return DataQualityMetricsModel(type=ref_type, ref_id=ref_id)
    def _delete_metrics_record(self, ref_type, ref_id):
        return DataQualityMetricsModel.remove(ref_type, ref_id)
    #-- check file size ---
    def get_file_size(self, resource_id, resource_url):
        """ดึง size ของ resource ถ้าเป็น upload"""

        # เรียก API resource_show
        context = {"model": model, "session": model.Session}
        data_dict = {"id": resource_id}

        try:
            resource = toolkit.get_action("resource_show")(context, data_dict)
        except toolkit.ObjectNotFound:
            log.error(f"Resource {resource_id} not found")
            return None

        url_type = resource.get("url_type", None)
        log.info(f"Resource {resource_id} url_type={url_type}")

        if url_type == "upload":
            # ถ้าเป็น upload → ดึง size
            size = resource.get("size")
            log.info(f"Resource {resource_id} size={size}")
            return size
        else:
            # ถ้าไม่ใช่ upload → ใช้วิธี custom
            return self.handle_non_upload(resource_url)
    # def handle_non_upload(self, url):
    #     timeout = 5
    #     try:
    #         response = requests.head(url, timeout=timeout)
    #         if response.status_code == 200:
    #             size = int(response.headers['Content-Length'])
    #             return size
    #         else:
    #             log.debug("Error: Could not retrieve file size, status code:", response.status_code)
    #             return None
    #     except Exception as e:
    #         log.debug("Missing Content-Length header for URL: %s", url)
    #         return None 
    def handle_non_upload(self, url):
        timeout = 5
        MAX_SIZE = 10_100_000  # 10.1 MB
        CHUNK_SIZE = 1024 * 1024  # 1MB ต่อรอบ

        try:
            response = requests.head(url, timeout=timeout, allow_redirects=True)

            if response.status_code == 200 and 'Content-Length' in response.headers:
                size = int(response.headers['Content-Length'])
                return size

            # ---- fallback ถ้าไม่มี Content-Length ----
            log.debug("No Content-Length header, fallback to partial download")

            length = 0
            with requests.get(url, stream=True, timeout=timeout, allow_redirects=True) as r:
                r.raise_for_status()
                for chunk in r.iter_content(chunk_size=CHUNK_SIZE):
                    if not chunk:
                        break
                    length += len(chunk)
                    if length > MAX_SIZE:
                        log.debug(f"File exceeded {MAX_SIZE} bytes, stopping early")
                        return length  # return ขนาดที่โหลดมาแล้ว (เกิน limit)

            return length  # กรณีโหลดจบแล้วไม่เกิน limit

        except Exception as e:
            log.debug("Error while checking file size: %s", e)
            return None

    def check_connection_url(self, url, timeout=5):
        log.debug('---check_connection_url--')
        try:
            #  ใช้ HEAD ก่อน
            response = requests.head(url, timeout=timeout, allow_redirects=True)

            if response.status_code == 200:
                return True
            else:
                return False

        except requests.exceptions.Timeout:
            return False
        except Exception as e:
            log.debug("Error: %s", e)
            return False
    def is_tabular(self,resource):
        machine_readable_formats = ['CSV', 'XLSX', 'XLS', 'JSON']
        mimetype_to_format = {
            'text/csv': 'CSV',
            'application/vnd.ms-excel': 'XLS',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': 'XLSX',
            'application/json': 'JSON',
            'application/xml': 'XML',
            'text/xml': 'XML'
        }

        data_format = (resource.get('format') or '').strip().upper()
        mimetype = (resource.get('mimetype') or '').strip().lower()

        # หาจาก mimetype ถ้า format ยังว่าง
        if not data_format and mimetype in mimetype_to_format:
            data_format = mimetype_to_format[mimetype]

        # หาจาก URL ถ้า format ยังว่าง
        if not data_format:
            resource_url = resource.get('url', '').strip()
            if '.' in resource_url:
                clean_url = resource_url.split('?')[0]
                extension = clean_url.split('.')[-1]
                data_format = extension.upper()

        # ลบจุดและเช็คว่าคือ tabular หรือไม่
        data_format = data_format.replace('.', '').upper()
        is_file_tabular = data_format in machine_readable_formats
        return is_file_tabular
    def is_openness_5_star_format(self, data_format): #special format 
        OPENNESS_5_STAR_FORMATS = ['RDF', 'TTL', 'N3', 'GEOJSON',  'GML', 'KML', 'SHP','WMS', 'ESRI REST']
        """ตรวจสอบว่า format นี้เป็นประเภท 5-star หรือไม่"""
        if not data_format:
            return False
        normalized_format = data_format.replace('.', '').upper()
        return normalized_format in OPENNESS_5_STAR_FORMATS
    def convert_mimetype_to_format(self, mimetype, resource_format, resource_url):
        file_format = resource_format.upper().strip()
        mimetype_map = {
            # Excel
            'application/vnd.ms-excel': 'XLS',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': 'XLSX',

            # CSV / Text / JSON
            'text/csv': 'CSV',
            'application/json': 'JSON',
            'application/geo+json': 'GEOJSON',
            'application/vnd.geo+json': 'GEOJSON',

            # PDF
            'application/pdf': 'PDF',

            # Word documents
            'application/msword': 'DOC',
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document': 'DOCX',

            # PowerPoint
            'application/vnd.ms-powerpoint': 'PPT',
            'application/vnd.openxmlformats-officedocument.presentationml.presentation': 'PPTX',

            # Images
            'image/jpeg': 'JPEG',
            'image/png': 'PNG',
            'image/gif': 'GIF',
            'image/tiff': 'TIFF',
            'image/svg+xml': 'SVG',

            # Plain text and XML
            'text/plain': 'TXT',
            'application/xml': 'XML',
            'text/xml': 'XML',
            'application/rdf+xml': 'RDF',
            'text/turtle': 'TURTLE',
            'application/n-triples': 'NTRIPLES',
            'application/ld+json': 'JSONLD',
            'application/trig': 'TRIG',
            'application/n-quads': 'NQUADS',

            # Compressed formats
            # 'application/zip': 'ZIP',
            # 'application/x-tar': 'TAR',
            # 'application/gzip': 'GZ',

            # Default fallback
            'application/octet-stream': 'BINARY'
        }
        # 1. map จาก mimetype
        mimetype_format = mimetype_map.get(mimetype, None)

        # 2. fallback ถ้าเจอ octet-stream → เดาจาก extension
        if mimetype == "application/octet-stream" and resource_url:
            path = urlparse(resource_url).path
            ext = os.path.splitext(path)[1].lower()
            if ext in ['.csv', '.xlsx', '.xls', '.json','.ppt','.pptx','.doc','.docx','.rdf']:
                mimetype_format = ext.replace('.', '').upper()

        # 3. ถ้าไม่มี mapping เลย ให้ใช้ resource_format ที่ส่งมา
        result = mimetype_format or file_format

        log.debug('--mimetype_format ---')
        log.debug(result)
        # if not mimetype_format:
        #     result = file_format  # ถ้าอย่างใดอย่างหนึ่งไม่มีค่า
        return result

    #*** ตรวจสอบว่า user กำหนด format มาถูกหรือไม่ ตอนนี้ไม่ได้ใช้  ****
    def inspect_file(self, resource):
        result = None
        file_path = resource.get('url')
        file_format = resource.get('format', '').upper().strip()  # กัน null และ normalize case

        # 1. ตรวจสอบ mimetype
        mimetype = ResourceFetchData2.detect_mimetype(file_path)

        # 2. แมป mimetype ไปยัง format ที่เราสนใจ
        mimetype_map = {
            # Excel
            'application/vnd.ms-excel': 'XLS',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': 'XLSX',

            # CSV / Text / JSON
            'text/csv': 'CSV',
            'application/json': 'JSON',
            'application/geo+json': 'GEOJSON',
            'application/vnd.geo+json': 'GEOJSON',

            # PDF
            'application/pdf': 'PDF',

            # Word documents
            'application/msword': 'DOC',
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document': 'DOCX',

            # PowerPoint
            'application/vnd.ms-powerpoint': 'PPT',
            'application/vnd.openxmlformats-officedocument.presentationml.presentation': 'PPTX',

            # Images
            'image/jpeg': 'JPEG',
            'image/png': 'PNG',
            'image/gif': 'GIF',
            'image/tiff': 'TIFF',
            'image/svg+xml': 'SVG',

            # Plain text and XML
            'text/plain': 'TXT',
            'application/xml': 'XML',
            'text/xml': 'XML',
            'application/rdf+xml': 'RDF',
            'text/turtle': 'TURTLE',
            'application/n-triples': 'NTRIPLES',
            'application/ld+json': 'JSONLD',
            'application/trig': 'TRIG',
            'application/n-quads': 'NQUADS',

            # Compressed formats
            # 'application/zip': 'ZIP',
            # 'application/x-tar': 'TAR',
            # 'application/gzip': 'GZ',

            # Default fallback
            'application/octet-stream': 'BINARY'
        }
        # 3. ตรวจสอบว่า mimetype ตรงกับฟอร์แมตหรือไม่
        mimetype_format = mimetype_map.get(mimetype, None)

        # 4. เปรียบเทียบ (normalize ทั้งสองข้างก่อนเทียบ)
        log.debug('--inspect file ---')
        log.debug(file_format)
        log.debug('--mimetype_format ---')
        log.debug(mimetype_format)
        log.debug(result)
        if file_format in ['XLS','XLSX','CSV','JSON','GEOJSON','PDF','DOC','PPT','PPTX',
                           'JPEG','PNG','GIF','TIFF','SVG','TXT','XML','RDF']:
            if file_format and mimetype_format:
                if file_format.strip().upper() == mimetype_format.strip().upper():
                    result = True
                else:
                    result = False
            else:
                result = None  # ถ้าอย่างใดอย่างหนึ่งไม่มีค่า

        return result
    def _handle_existing_record(self, data_quality, last_modified, resource):
        """Handle logic when a previous metric record exists."""
        self.logger.debug("Found previous metric record.")

        if data_quality.resource_last_modified < last_modified:
            self.logger.debug("Resource has been updated. Recalculating metrics.")
            self._delete_metrics_record('resource', resource['id'])
            updated = self._new_metrics_record('resource', resource['id'])
            updated.resource_last_modified = last_modified
            return False  # recalculate

        if data_quality.resource_last_modified == last_modified:
            all_calculated = all(
                data_quality.metrics.get(m.name) is not None for m in self.metrics
            )
            if all_calculated:
                self.logger.debug("All metrics already calculated. Using cache.")
                return True
            else:
                self.logger.debug("Partial metrics found. Will recalculate missing.")
                return True

        self.logger.debug("No valid condition met. Creating new record.")
        data_quality = self._new_metrics_record('resource', resource['id'])
        data_quality.resource_last_modified = last_modified
        return False

    def calculate_metrics_for_resource(self, resource, job_id=None):
        log.debug('calculate_metrics_for_resource')
        if resource.get('last_modified') is not None:
            last_modified = dateutil.parser.parse(resource.get('last_modified') or resource.get('created'))
        else:
            last_modified = dateutil.parser.parse(resource.get('created'))
        #-----old version----------------------
        # metadata_modified = resource.get('metadata_modified')
        # if metadata_modified:  # Check if the value is not None
        #     last_modified = datetime.strptime(metadata_modified, '%Y-%m-%dT%H:%M:%S.%f')
        # else:
        #     last_modified = None  # Handle the case where metadata_modified is None

        self.logger.debug ('Resource last modified on: %s', last_modified)
        #-------Check Data Dict using Resource Name -----------       
        resource_id = resource['id']
        resource_url = resource['url']
        if 'url' in resource and resource['url']:
            if "209.15.113.87" in resource_url:
                resource_url = resource_url.replace("209.15.113.87", "data.go.th")
                resource['url'] = resource_url   # set ค่าใหม่เข้า resource object เลย
        
        log.debug(resource_url)
        try:
            resource_name = resource['name']
            resource_name = resource_name.lower()            
            # initializing test list
            datadict_list = ['datadict', 'data dict','data_dictionary','data dictionary','คำอธิบายชุดข้อมูล']
            # checking if string contains list element
            res_datadict = [ele for ele in datadict_list if(ele in resource_name)]
            is_datadict = bool(res_datadict)
        except Exception as e:
            # fallback กรณีมี error
            is_datadict = False
            log.warning("Error checking datadict for resource %s: %s", resource.get('id'), e)
        results = {}
        file_size_mb = 0
        execute_time = 0
        timeout = 5  # in seconds
        connection_url = False
        error_file_size = ''
        error_fetching_resource = ''
        # error_file_not_match = ''
        today = datetime.today().strftime("%Y-%m-%d")
        #ตรวจสอบว่า user ใส่ format มาถูกหรือไม่ ตอนนี้ไม่ได้ใช้แล้ว
        # file_info = self.inspect_file(resource)
        # log.debug(file_info)
        # Check if the request was successful
        if self.check_connection_url(resource_url, timeout):     
            start_time = time.time()
            log.debug(start_time)
            connection_url = True
            file_size = self.get_file_size(resource_id,resource_url)
            #----Check mimetype-----------------------
            mimetype = ResourceFetchData2.detect_mimetype(resource_url)
            detected_format = ''
            detected_format = self.convert_mimetype_to_format(mimetype,resource['format'],resource_url)
            if file_size is not None:
                file_size_mb = file_size/1024**2
                log.debug("File size (MB): %s", file_size_mb)
            # if file_info == False:
            #     error_file_not_match = 'Invalid file format'
            if not is_datadict:                                                   
                #----- connect model: check records----------------------
                data_quality = self._get_metrics_record('resource', resource['id']) #get data from DB   
                cached_calculation = False

                if data_quality:
                    cached_calculation = self._handle_existing_record(data_quality, last_modified, resource)
                else:
                    data_quality = self._new_metrics_record('resource', resource['id'])
                    data_quality.resource_last_modified = last_modified
                    self.logger.debug('First time data quality calculation.')
                # if data_quality:
                #     self.logger.debug('Data Quality calculated for '
                #                     'version modified on: %s',
                #                     data_quality.resource_last_modified)
                #     # if data_quality.resource_last_modified >= last_modified:
                #     if data_quality.resource_last_modified < last_modified:
                #         cached_calculation = True
                #         log.debug('--delete and run new updated file--')
                #         #delete record
                #         self._delete_metrics_record('resource', resource['id'])
                #         #calculate new record
                #         data_quality = self._new_metrics_record('resource', resource['id'])
                #         data_quality.resource_last_modified = last_modified
                #         self.logger.debug('Re-calculation.')
                #     elif (data_quality.resource_last_modified == last_modified):
                #         cached_calculation = True
                #         # check if all metrics have been calculated or some needs to be
                #         # calculated again
                #         if all(map(lambda m: m is not None, [
                #                     data_quality.completeness,
                #                     data_quality.uniqueness,
                #                     data_quality.validity,
                #                     data_quality.timeliness,
                #                     data_quality.consistency,
                #                     data_quality.openness,
                #                     data_quality.downloadable,
                #                     data_quality.access_api,
                #                     data_quality.machine_readable
                #                     ])):
                #             self.logger.debug('Data Quality already calculated.')
                #             return data_quality.metrics
                #         else:
                #             self.logger.debug('Data Quality not calculated for '
                #                             'all dimensions.')
                #     else:
                #         #calculate new record
                #         data_quality = self._new_metrics_record('resource',
                #                                                 resource['id'])
                #         data_quality.resource_last_modified = last_modified
                #         self.logger.debug('Resource changed since last calculated. '
                #                         'Calculating data quality again.')
                # else:
                #     data_quality = self._new_metrics_record('resource', resource['id'])
                #     data_quality.resource_last_modified = last_modified
                #     self.logger.debug('First data quality calculation.')                
                #--GET Package_id and Org_name----------------------------
                resource_id = resource['id']
                query = (
                    Session.query(Package.id.label('package_id'), Group.name.label('org_name'))
                    .select_from(Resource)
                    .join(Package, Resource.package_id == Package.id)
                    .join(Group, Package.owner_org == Group.id)  # หรือ Organization ถ้าใช้นั้น
                    .filter(Resource.id == resource_id)
                )
                res_result = query.first()
                if res_result:
                    log.debug(f"Package: {res_result.package_id}, Org: {res_result.org_name}")
                    data_quality.package_id = res_result.package_id
                    data_quality.org_name   = res_result.org_name
                        
                data_quality.ref_id = resource_id                   
                data_quality.resource_last_modified = last_modified
                data_quality.job_id = job_id  # เก็บ job_id ลง DB
                #----------------------------------------------------
                #----------------Calculate Metrics--------------------
                data_stream2 = None
                if self.force_recalculate:
                    log.info('Forcing recalculation of the data metrics '
                            'has been set. All except the manually set metric data '
                            'will be recalculated.')
                self.logger.debug('******Calculate Metrics *****')
                self.logger.debug(self.metrics)
                # self.logger.debug(resource)
                consistency_val = 0
                # encoding = ''
                validity_report = {}
                for metric in self.metrics:
                    self.logger.debug(metric)      
                    try:
                        if cached_calculation and getattr(data_quality,
                                                        metric.name) is not None:
                            cached = data_quality.metrics[metric.name]
                            if not self.force_recalculate and not cached.get('failed'):
                                self.logger.debug('Dimension %s already calculated. '
                                                'Skipping...', metric.name)
                                results[metric.name] = cached
                                continue
                            if cached.get('manual'):
                                self.logger.debug('Calculation has been performed '
                                                'manually. Skipping...', metric.name)
                                results[metric.name] = cached
                                continue
                        
                        self.logger.debug('Calculating dimension: %s...', metric)
                    #    #-----Fetch DATA------------------------------------------------
                    #     if not data_stream2:
                    #         data_stream2 = self._fetch_resource_data2(resource)
                    #     else:
                    #         if data_stream2.get('records') and \
                    #                 hasattr(data_stream2['records'], 'rewind'):
                    #             data_stream2['records'].rewind()
                    #         else:
                    #             data_stream2 = self._fetch_resource_data2(resource)                   
                    #     # log.debug(data_stream2)
                    #     if data_stream2.get('error'):
                    #         error_fetching_resource = data_stream2.get('error')
                    #     log.debug('------ End call Data Stream2-----')
                        #-------------------------------------------------------------
                        #using metadata for calculate metrics                                     
                        if (file_size_mb <= 10 and connection_url):
                            #-----Fetch DATA------------------------------------------------
                            if not data_stream2:
                                data_stream2 = self._fetch_resource_data2(resource)
                            else:
                                if data_stream2.get('records') and \
                                        hasattr(data_stream2['records'], 'rewind'):
                                    data_stream2['records'].rewind()
                                else:
                                    data_stream2 = self._fetch_resource_data2(resource)                   
                            # log.debug(data_stream2)
                            if data_stream2.get('error'):
                                error_fetching_resource = data_stream2.get('error')
                            log.debug('------ End call Data Stream2-----')
                            log.debug('------ mimetype -----')               
                            # #-------------------------------------------------------------
                            log.debug('------ check all metrics-----')
                            if(metric.name == 'openness' or metric.name == 'availability' or  metric.name == 'downloadable' or metric.name == 'access_api' or metric.name == 'preview'):
                                log.debug('------ check metric-----')
                                log.debug(metric.name)
                                results[metric.name] = metric.calculate_metric(resource)
                            
                            elif metric.name == 'timeliness':     
                                log.debug('----timeliness------')                                         
                                results[metric.name] = metric.calculate_metric(resource)
                                timeliness_val = results[metric.name] 

                            elif metric.name in ['acc_latency', 'freshness']:   
                                if timeliness_val:
                                    log.debug('----acc_latency------')                                         
                                    results[metric.name] = metric.calculate_metric(resource, timeliness_val)
                                else:
                                    continue
                            elif( metric.name == 'relevance'):
                                log.debug('------relevance-----')
                                #ถ้าตรวจแบบ organization
                                level_name = 'nectec'
                                execute_type = 'organization'
                                results[metric.name] = metric.calculate_metric(resource,level_name,execute_type)
                            # elif(metric.name == 'machine_readable'):
                            #     log.debug('----machine_readable_val------')       
                            #     results[metric.name] = metric.calculate_metric_machine(resource,consistency_val,validity_report)
                            tabular_format = self.is_tabular(resource)
                            openness_5_star_format = self.is_openness_5_star_format(resource['format'])
                            log.debug(f"[check tabular_format] => {tabular_format}")
                            if(tabular_format):
                                if(metric.name == 'consistency'):
                                    log.debug('------Call consistency -----')                           
                                    results[metric.name] = metric.calculate_metric(resource,data_stream2)
                                    consistency_val = results[metric.name].get('value')
                                    log.debug(consistency_val)
                                elif(metric.name == 'validity'):
                                    log.debug('----check validity------')
                                    results[metric.name] = metric.calculate_metric(resource,data_stream2)                           
                                    validity_report = results[metric.name].get('report')
                                elif(metric.name == 'completeness'):                           
                                    log.debug('----check completeness------')
                                    log.debug(data_stream2['total'])
                                    results[metric.name] = metric.calculate_metric(resource,data_stream2)                                   
                                    completeness_report = results[metric.name].get('report')
                                    log.debug('-----after calculate completeness--')
                                    # log.debug("---dir tempfile----")
                                    # log.debug(os.listdir(tempfile.gettempdir()))
                                elif(metric.name == 'uniqueness'):                           
                                    log.debug('----check uniqueness------')
                                    log.debug(data_stream2['total'])
                                    results[metric.name] = metric.calculate_metric(resource,data_stream2)                                   
                                    uniqueness_report = results[metric.name].get('report')
                                elif(metric.name == 'utf8'):
                                    log.debug('----utf8_val------')       
                                    results[metric.name] = metric.calculate_metric_utf8(resource,validity_report)
                            elif (openness_5_star_format):
                                results['consistency'] = { 'value': 100 }
                                results['validity']    =    { 'value': 100 }
                                results['completeness'] = { 'value': None }
                                results['uniqueness'] = { 'value': None }
                            else:
                                results['consistency'] = { 'value': None }
                                results['validity']    =    { 'value': None }
                                results['completeness'] = { 'value': None }
                                results['uniqueness'] = { 'value': None }
                                
                            # else:
                            #     log.debug('----else----')    
                            #     log.debug(metric.name)                                         
                            #     results[metric.name] = metric.calculate_metric(resource) #,data_stream
                        #file_size > 10 MB
                        else:
                            error_file_size = 'file_size > 10 MB'
                            openness_5_star_format = self.is_openness_5_star_format(resource['format'])
                            
                            if(metric.name == 'openness' or metric.name == 'availability' or metric.name == 'downloadable' or metric.name == 'access_api' or metric.name == 'preview'):
                                results[metric.name] = metric.calculate_metric(resource)
                            # elif(metric.name == 'utf8'):
                            #     log.debug('----utf8_val------')       
                            #     results[metric.name] = metric.calculate_metric_utf8(resource,{})
                            elif (metric.name == 'timeliness'):     
                                log.debug('----timeliness------')                                         
                                results[metric.name] = metric.calculate_metric(resource)
                                timeliness_val = results[metric.name]
                            elif (metric.name == 'acc_latency' or metric.name == 'freshness'):     
                                log.debug('----acc_latency------')                                         
                                results[metric.name] = metric.calculate_metric(resource,timeliness_val)
                            elif( metric.name == 'relevance'):
                                log.debug('------relevance-----')
                                #ถ้าตรวจแบบ organization
                                org_name = 'nectec'
                                execute_type = 'organization'
                                results[metric.name] = metric.calculate_metric(resource,org_name,execute_type)
 
                            #do not execute big file for consistency and validty
                            results['consistency'] = { 'value': None }
                            results['validity']    =    { 'value': None }
                            results['completeness'] = { 'value': None }
                            results['uniqueness'] = { 'value': None }
                            results['utf8'] = { 'value': None }
                            #except openness_5_star_format
                            if (openness_5_star_format):
                                results['consistency'] = { 'value': 100 }
                                results['validity']    =    { 'value': 100 }
                                results['completeness'] = { 'value': None }
                                results['uniqueness'] = { 'value': None }
                            if not connection_url:
                                results['connection_url'] = { 'error': True}
                        
                    except Exception as e:
                        self.logger.error('Failed to calculate: %s. Error: %s',
                                        metric, str(e))
                        self.logger.exception(e)
                        results['error'] = "Failed to calculate:"+metric.name
                        results[metric.name] = {
                            'failed': True,
                            'error': str(e),
                        }
                # set results -- old version
                # for metric, result in results.items():
                #     if result.get('value') is not None:
                #         setattr(data_quality, metric, result['value'])
                # set results -- new version-----
                # ตรวจสอบและจัดรูปแบบผลลัพธ์ ให้เป็น dict ที่มีโครงสร้างเดียวกัน ({'value': ..., 'error': ...})
                # หลังจากได้ results มา
                for metric, result in list(results.items()):
                    if not isinstance(result, dict):
                        # ห่อข้อความผิดพลาดไว้ใน dict
                        results[metric] = {'value': None, 'error': str(result)}

                # จากนั้นจึงตั้ง attribute ได้อย่างปลอดภัย
                for metric, result in results.items():
                    value = result.get('value')
                    setattr(data_quality, metric, value)

                
                data_quality.modified_at = datetime.now()
                #---- add execute time--
                end_time = time.time()   
                # Calculate the time taken
                execute_time = end_time - start_time
                log.debug("----execute_time----")
                log.debug(execute_time)    
                # if 'error' in results and results['error']:
                #     data_quality.error = results['error'].get('error')
                # else:
                #     data_quality.error = ''
                #รวม error จากหลาย metric ไว้ใน list:
                error_list = []
                if error_file_size != '':
                    error_list.append(error_file_size)
                # if error_file_not_match != '':
                #     error_list.append(error_file_not_match)
                if error_fetching_resource != '':
                    error_list.append(error_fetching_resource)  # ถ้า error จาก fetch มีค่า → เพิ่มเข้าไป
                for metric, result in results.items():
                    if isinstance(result, dict) and result.get('error'):
                        error_list.append(f"{metric}: {result['error']}")

                if error_list:
                    results['error'] = {'error': '; '.join(error_list)}
                    data_quality.error = results['error']['error']
                else:
                    data_quality.error = ''
                data_quality.version = today
                data_quality.format  = detected_format #resource['format']
                data_quality.url = resource_url
                data_quality.metrics = results
                data_quality.file_size    = round(file_size_mb,3)
                data_quality.execute_time = round(execute_time,3)
                data_quality.save()
                self.logger.debug('Metrics calculated for resource: %s',
                                resource['id'])

        else:
            self.logger.debug('Connection Timed Out')
            #----- connect model: check records----------------------
            
            data_quality = self._get_metrics_record('resource', resource['id']) #get data from DB
            if data_quality:
                self.logger.debug('Data Quality already calculated.')
                self.logger.debug(resource['id'])
            else:
                data_quality = self._new_metrics_record('resource', resource['id'])
                 #--GET package_id and Org_name---------------------------
                resource_id = resource['id']
                query = (
                    Session.query(Package.id.label('package_id'), Group.name.label('org_name'))
                    .select_from(Resource)
                    .join(Package, Resource.package_id == Package.id)
                    .join(Group, Package.owner_org == Group.id)  # หรือ Organization ถ้าใช้นั้น
                    .filter(Resource.id == resource_id)
                )
                res_result = query.first()
                if res_result:
                    log.debug(f"Package: {res_result.package_id}, Org: {res_result.org_name}")
                    data_quality.package_id = res_result.package_id
                    data_quality.org_name   = res_result.org_name
                        
                data_quality.ref_id = resource_id                   
                data_quality.resource_last_modified = last_modified
            #----------------------------------------------------------
                self.logger.debug('First data quality calculation.')
                #----------------Calculate Metrics--------------------
                # data_quality.ref_id = resource['id']
                # data_quality.resource_last_modified = last_modified
                data_quality.metrics = {'error':'connection timed out'}
                data_quality.modified_at = datetime.now()
                #----------------------
                data_quality.openness = None
                data_quality.availability = None
                data_quality.downloadable = None
                data_quality.access_api = None
                data_quality.timeliness = None #-1, 999
                data_quality.acc_latency= None
                data_quality.freshness = None
                data_quality.relevance = None
                data_quality.completeness = None
                data_quality.consistency = None
                data_quality.validity = None
                data_quality.uniqueness = None
                data_quality.preview = None 
                data_quality.utf8    = None 
                # data_quality.machine_readable = 0      
                #---- add filepath ----
                # data_quality.filepath = ''
                data_quality.error = 'Connection timed out'
                data_quality.version = today
                data_quality.format  = resource['format']
                data_quality.url = resource_url
                data_quality.job_id = job_id
                data_quality.file_size = None
                data_quality.execute_time = None
                data_quality.save()
                self.logger.debug('Metrics calculated for resource: %s',
                                    resource['id'])

        return results

    def calculate_cumulative_metrics(self, package_id, resources, results, job_id=None):
        '''Calculates the cumulative metrics (reduce phase), from the results
        calculated for each resource in the dataset.

        The cumulative values for the dataset are always calculated from the
        results, except in the case when the values for a particular dimension
        have been set manually. In that case, the manual value is used.

        The results of the calculation are stored in database in a separate
        entry containing the valued for Data Quality metrics for the whole
        dataset.

        :param package_id: `str`, the ID of the CKAN dataset.
        :param resources: `list` of CKAN resources for which Data Qualtiy
            metrics have been calculated.
        :param results: `list` of result `dict`, the results of the calculation
            for each of the resources.
        '''
        self.logger.debug('Cumulative data quality metrics for package: %s',
                          package_id)
        data_quality = self._get_metrics_record('package', package_id)
        if not data_quality:
            data_quality = self._new_metrics_record('package', package_id)
        else:
            self.logger.debug('delete package record-->')
            self._delete_metrics_record('package', package_id)
            data_quality = self._new_metrics_record('package', package_id)
        cumulative = {}
        dataset_results = data_quality.metrics or {}
        for metric in self.metrics:
            cached = dataset_results.get(metric.name, {})
            if cached.get('manual'):
                self.logger.debug('Metric %s is calculated manually. '
                                  'Skipping...', metric.name)
                cumulative[metric.name] = cached
                continue
            metric_results = [res.get(metric.name, {}) for res in results]
            if metric.name == 'relevance':
                cumulative[metric.name] = metric.calculate_cumulative_metric(
                    package_id,
                    resources,
                    metric_results
                )
            else:
                cumulative[metric.name] = metric.calculate_cumulative_metric(
                    resources,
                    metric_results
                )

        if cumulative != dataset_results:
            data_quality = self._new_metrics_record('package', package_id)
        for metric, result in cumulative.items():
            if result.get('value') is not None:
                setattr(data_quality, metric, result['value'])

       #-------Add Package ID------------------- 
        query = (
            Session.query(Package.id.label('package_id'), Group.name.label('org_name'))
            .select_from(Package)
            .join(Group, Package.owner_org == Group.id)
            .filter(Package.id == package_id)
        )
        p_result = query.first()
        if p_result:
            data_quality.package_id = p_result.package_id
            data_quality.org_name   = p_result.org_name

        data_quality.metrics = cumulative
        data_quality.modified_at = datetime.now()
        data_quality.job_id = job_id
        data_quality.save()
        self.logger.debug('Cumulative metrics calculated for: %s', package_id)

class OpendQuality(object):
    
    # @classmethod
    # def top_pageckage_view(cls, limit=3, page_no=0):
    #     pass
   
    def test(cls, limit=3, page_no=0):
        return 'OK'
    @classmethod
    def get_last_modified_datasets(limit=5):
        try:
            package = model.Package
            q = model.Session.query(package.name, package.title, package.type, package.metadata_modified.label('date_modified')).filter(package.state == 'active').order_by(package.metadata_modified.desc()).limit(limit)
            packages = q.all()
        except:
            return []
        return packages
class ResourceFetchData(object):
    '''A callable wrapper for fetching the resource data.

    Based on the availability of the data, when called it will fetch a page of
    the resource data.
    The data for a particular resource may be stored in the Data Store, in CKAN
    storage path or on a remote server. Ideally we want to read the data from
    the data store, however if the data is not available there, we should try
    to download the CSV data directly - either from CKAN or if the resource was
    set as a link to an external server, by downloading it from that server.

    :param resource: `dict`, the resource metadata as retrieved from CKAN
        action `resource_show`.
    '''

    def __init__(self, resource):
        self.resource = resource
        self.download_resource = False
        self.resource_csv = None

    def __call__(self, page, limit):
        '''Fetches one page (with size of `limit`) of data from the resource
        CSV data.

        This call will first try to load the data from the datastore. If that
        fails and the data is not available in the data store, then it will try
        to load the data directly (from CKAN uploads or by downloading the CSV
        file directly from the remote server). Once this happens, and the data
        is loaded into `ResourceCSVData`, then every subsequent call will load
        the data from the cached instance of `ResourceCSVData` and will not try
        to load it from the data store.

        :param page: `int`, the page to fetch, 0-based.
        :param limit: `int`, page size.

        :returns: `dict`, the requested page as `dict` containing:
            * `total` - total number of records in the data.
            * `records` - a `list` of records for this page. Each record is a
                `dict` of format `<column_name>:<row_value>`.
            * `fields` - the records metadata describing each column. It is a
                `list` of `dict`, where each element describes one column. The
                order of the columns is the same as it is appears in the CSV
                file. Each element has:
                    * `id` - `str`, the name of the column
                    * `type` - `str`, the column type (ex. `numeric`, `text`)
        '''
        try:
            page = self.fetch_page(page, limit)
            return page
        except Exception as e:
            log.error('Failed to fetch page %d (limit %d) of resource %s. '
                      'Error: %s',
                      page,
                      limit,
                      self.resource.get('id'),
                      str(e))
            log.exception(e)
            return {
                'total': 0,
                'records': [],
                'fields': {},
            }

    def _fetch_data_datastore(self, page, limit):
        log.debug('Fetch page from datastore: page %d, limit %d', page, limit)
        return toolkit.get_action('datastore_search')({
            'ignore_auth': True,
        }, {
            'resource_id': self.resource['id'],
            'offset': page*limit,
            'limit': limit,
        })

    def _download_resource_from_url(self, url, headers=None):
        resp = requests.get(url, headers=headers)
        resp.raise_for_status()  # Raise an error if request to file failed.
        data = []
        with TemporaryFile(mode='w+b') as tmpf:
            for chunk in resp.iter_content():
                tmpf.write(chunk)
            tmpf.flush()
            tmpf.seek(0, 0)  # rewind to start
            reader = csv.reader(tmpf)
            for row in reader:
                data.append(row)
        return data

    def _download_resource_from_ckan(self, resource):       
        upload = uploader.get_resource_uploader(resource)
        filepath = upload.get_path(resource['id'])
        try:
            data = []
            if os.path.isfile(filepath):
                data = []
                with open(filepath) as csvf:
                    reader = csv.reader(csvf)
                    for row in reader:
                        data.append(row)
                return data
            else:
                # The worker is not in the same machine as CKAN, so it cannot
                # read the resource files from the local file system.
                # We need to retrieve the resource data from the resource URL.
                if 'url' not in resource:
                    raise Exception('Cannot access the resource because the '
                                    'resource URL is not set.')
                log.debug('Resource data is not in this file system '
                          '(File %s does not exist).'
                          'Fetching data from CKAN download url directly.',
                          filepath)
                headers = None
                sysadmin_api_key = _get_sysadmin_user_key()
                if sysadmin_api_key:
                    headers = {'Authorization': sysadmin_api_key}
                return self._download_resource_from_url(
                    resource['url'],
                    resource['format'],
                    headers
                )
        except OSError as e:
            log.error('Failed to download resource from CKAN upload. '
                      'Error: %s', str(e))
            log.exception(e)
            raise e

    def _fetch_data_directly(self):
        if self.resource.get('url_type') == 'upload':
            log.debug('Getting data from CKAN...')
            return self._download_resource_from_ckan(self.resource)
        if self.resource.get('url'):
            log.debug('Getting data from remote URL...')
            return self._download_resource_from_url(self.resource['url'], self.resource['format'])
        raise Exception('Resource {} is not available '
                        'for download.'.format(self.resource.get('id')))

    def fetch_page(self, page, limit):
        '''Fetches one page (with size of `limit`) of data from the resource
        CSV data.

        This call will first try to load the data from the datastore. If that
        fails and the data is not available in the data store, then it will try
        to load the data directly (from CKAN uploads or by downloading the CSV
        file directly from the remote server). Once this happens, and the data
        is loaded into `ResourceCSVData`, then every subsequent call will load
        the data from the cached instance of `ResourceCSVData` and will not try
        to load it from the data store.

        :param page: `int`, the page to fetch, 0-based.
        :param limit: `int`, page size.

        :returns: `dict`, the requested page as `dict` containing:
            * `total` - total number of records in the data.
            * `records` - a `list` of records for this page. Each record is a
                `dict` of format `<column_name>:<row_value>`.
            * `fields` - the records metadata describing each column. It is a
                `list` of `dict`, where each element describes one column. The
                order of the columns is the same as it is appears in the CSV
                file. Each element has:
                    * `id` - `str`, the name of the column
                    * `type` - `str`, the column type (ex. `numeric`, `text`)
        '''
        log.debug('----call fetch_page at ResourceFetchData-------')
        if self.download_resource:
            if not self.resource_csv:
                self.resource_csv = ResourceCSVData(self._fetch_data_directly)
                log.debug('Resource data downloaded directly.')
            return self.resource_csv.fetch_page(page, limit)
        try:
            page = self._fetch_data_datastore(page, limit)
            log.debug('Data is available in DataStore. '
                      'Using datastore for data retrieval.')
            return page
        except Exception as e:
            log.warning('Failed to load resource data from DataStore. '
                        'Error: %s', str(e))
            log.exception(e)
            self.download_resource = True
            log.debug('Will try to download the data directly.')
            return self.fetch_page(page, limit)
class ResourceFetchData2(object):
    '''A callable wrapper for fetching the resource data.

    Based on the availability of the data, when called it will fetch a page of
    the resource data.
    The data for a particular resource may be stored in the Data Store, in CKAN
    storage path or on a remote server. Ideally we want to read the data from
    the data store, however if the data is not available there, we should try
    to download the CSV data directly - either from CKAN or if the resource was
    set as a link to an external server, by downloading it from that server.

    :param resource: `dict`, the resource metadata as retrieved from CKAN
        action `resource_show`.
    '''

    def __init__(self, resource):
        self.resource = resource
        self.download_resource = False
        self.resource_csv = None

    def __call__(self, page, limit):
        '''Fetches one page (with size of `limit`) of data from the resource
        CSV data.

        This call will first try to load the data from the datastore. If that
        fails and the data is not available in the data store, then it will try
        to load the data directly (from CKAN uploads or by downloading the CSV
        file directly from the remote server). Once this happens, and the data
        is loaded into `ResourceCSVData`, then every subsequent call will load
        the data from the cached instance of `ResourceCSVData` and will not try
        to load it from the data store.

        :param page: `int`, the page to fetch, 0-based.
        :param limit: `int`, page size.

        :returns: `dict`, the requested page as `dict` containing:
            * `total` - total number of records in the data.
            * `records` - a `list` of records for this page. Each record is a
                `dict` of format `<column_name>:<row_value>`.
            * `fields` - the records metadata describing each column. It is a
                `list` of `dict`, where each element describes one column. The
                order of the columns is the same as it is appears in the CSV
                file. Each element has:
                    * `id` - `str`, the name of the column
                    * `type` - `str`, the column type (ex. `numeric`, `text`)
        '''
        try:
            page = self.fetch_page2(page, limit)
            return page
        except Exception as e:
            # log.error('2Failed to fetch page %d (limit %d) of resource %s. '
            #           'Error: %s',
            #           page,
            #           limit,
            #           self.resource.get('id'),
            #           str(e))
            # log.exception(e)
            return {
                'total': 0,
                'records': [],
                'fields': {},
                'error': 'Failed to fetch data for resource'
            }

    def _fetch_data_datastore(self, page, limit):
        log.debug('Fetch page from datastore: page %d, limit %d', page, limit)
        return toolkit.get_action('datastore_search')({
            'ignore_auth': True,
        }, {
            'resource_id': self.resource['id'],
            'offset': page*limit,
            'limit': limit,
        })
    def _fetch_data_datastore_defined_row(self, resource):
        log.debug('--data store--')
        page = 0
        limit = 5001
        log.debug("resource_id")
        log.debug(resource.get('id'))
        data = []
        try:
            # ตรวจว่ามี datastore จริงหรือไม่ก่อน
            context = {'ignore_auth': True}
            data_dict = {'id': resource.get('id')}
            log.debug("check datastore1")
            log.debug(f"data_dict: {data_dict}")
            ds_metadata = toolkit.get_action('datastore_info')(context, data_dict)
            log.debug("check datastore2")
            # log.debug(ds_metadata)
            
            if resource['datastore_active'] == True:
                
                result = toolkit.get_action('datastore_search')(
                    {'ignore_auth': True},  
                    {'resource_id': resource.get('id'), 'offset': 0, 'limit': limit} 
                )

                # เปลี่ยน format ให้อยู่ในรูปแบบ rows
                headers = list(result['records'][0].keys()) if result['records'] else []
                data.append(headers)
                for row in result['records']:
                    data.append([row.get(h, None) for h in headers])
            else:
                log.warning("Resource ไม่มีข้อมูลใน datastore")
        except Exception as e:
            log.debug(f"cannot read from datastore: {e}")
            return {
                'source': 'none',
                'data': []
            }
        return data

    def get_url(self, url,headers):
        kwargs = {
            'headers': headers,
            'timeout': DOWNLOAD_TIMEOUT,
            'verify': SSL_VERIFY,
            'stream': True
        }
        return requests.get(url, **kwargs)

    def get_response(self, url, headers):
        response = self.get_url(url, headers)
        if response.status_code == 202:
            wait = 1
            while wait < 120 and response.status_code == 202:
                time.sleep(wait)
                response = self.get_url(url,headers)  # แก้ให้ส่ง url
                wait *= 3
        response.raise_for_status()
        return response
    def _download_resource_from_url(self, url, resource_format, headers=None):
        
        #---------------------------------------
        data = []
        n_rows = 5001
        log.debug('----resource format-----')
        filepath = url
        format_url = filepath.split(".")[-1]
        mimetype = ResourceFetchData2.detect_mimetype(filepath)
        if mimetype == "text/html":
            log.debug(f"Skip HTML file by mimetype: {filepath}")
            return []
        log.debug(f"Downloading {filepath}, detected mimetype: {mimetype}")

        try:
            response = self.get_response(url, {})
            # --- เช็ค Content-Type ---
            content_type = response.headers.get("Content-Type", "").lower()
            if "text/html" in content_type:
                log.debug(f"Skip HTML resource: {url}")
                response.close()
                return []   # หรือ return None ตามที่ระบบหลักคุณรองรับ
        except requests.exceptions.RequestException as e:
            log.error(f"Cannot download file: {e}")
            return []

        # # tmp_file = tempfile.NamedTemporaryFile(suffix=filename, delete=False)
        filename = url.split('/')[-1].split('#')[0].split('?')[0]
        #-------------------------------
        # # แยกชื่อกับนามสกุล
        # basename, ext = os.path.splitext(filename)

        # # ตัดชื่อให้สั้นลง (กันชื่อไฟล์ยาวเกินไป)
        # safe_basename = basename[:30]  # เอาแค่ 30 ตัวอักษรแรก
        #-------------------------------
        # ถ้า filename ไม่มี basename (เช่น .xlsx เฉย ๆ) ให้ใช้ชื่อ default
        if not filename or filename.startswith('.'):
            ext = os.path.splitext(filename)[1] if '.' in filename else ''
            basename = "downloaded_file"
        else:
            basename, ext = os.path.splitext(filename)

        # กันชื่อยาวเกินไป
        safe_basename = basename[:30]

        # ถ้าไม่มีนามสกุลไฟล์ให้เดาจาก mimetype
        if not ext:
            if mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
                ext = ".xlsx"
            elif mimetype == "application/vnd.ms-excel":
                ext = ".xls"
            elif mimetype == "text/csv":
                ext = ".csv"
            elif mimetype == "application/json":
                ext = ".json"

        # สร้าง temp file โดยใช้ prefix + suffix
        tmp_file = tempfile.NamedTemporaryFile(
            prefix=f"{safe_basename}_",  # เช่น report_from_ckan_...
            suffix=ext,                  # เช่น .csv, .xlsx
            delete=False
        )
        log.debug("-----temp file---")
        log.debug(tmp_file.name)
        try:
            length = 0
            m = hashlib.md5()
            raw_data = b''     

            for chunk in response.iter_content(CHUNK_SIZE):
                length += len(chunk)
                if length > MAX_CONTENT_LENGTH:
                    response.close()
                    raise DataTooBigError("File too large")
                tmp_file.write(chunk)
                raw_data += chunk
                m.update(chunk)

            response.close()
            tmp_file.flush()
            tmp_file.seek(0)
            # --- Detect encoding ---
            result = chardet.detect(raw_data[:100000])  # ใช้แค่ 100KB แรกพอ
            encoding = result['encoding'] or 'utf-8'
            log.debug(f"Detected encoding: {encoding}")

            # ใช้ priority: mimetype ก่อน → format ทีหลัง
            if not mimetype and resource_format:
                if resource_format == "CSV":
                    mimetype = "text/csv"
                elif resource_format == "JSON":
                    mimetype = "application/json"
                elif resource_format == "XLSX":
                    mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                elif resource_format == "XLS":
                    mime_type = "application/vnd.ms-excel"


            if(mimetype == 'text/csv'): #(resource_format =='CSV'):
                log.debug('----csv----')     
                # log.debug(filepath)
                #-----------------------------------      
                try:
                    log.debug('--Reading CSV from temp--')
                    reader = csv.reader(io.TextIOWrapper(tmp_file, encoding=encoding, newline='')) #'utf-8'
                    records_read = 0
                    for row in reader:
                        # log.debug(row)
                        data.append(row)
                        records_read += 1
                        if records_read >= n_rows:
                            break
                    
                    # log.debug(data)
                except UnicodeDecodeError:
                    log.debug("An UnicodeDecodeError occurred")
                    data = []
                    # data = self._fetch_data_datastore_defined_row(self.resource) 
                except Exception as e:
                    log.debug("An error occurred, use CKAN datastore to readfile: %s", e)
                    data = []
                    # data = self._fetch_data_datastore_defined_row(self.resource) 

                
            elif(mimetype == 'application/json'): #elif(resource_format == 'JSON'):
                data = []
                try:
                    log.debug('--Reading JSON data--')
                    
                    # Check if the filepath is a URL or a file path
                    if filepath.startswith('http'):  # If it's a URL
                        response = requests.get(filepath)
                        response.encoding = 'utf-8'  # Ensure correct encoding
                        content_type = response.headers.get('Content-Type', '')
                        log.debug(content_type)
                        # Ensure content type is JSON
                        if 'application/json' in content_type or 'application/octet-stream' in content_type:
                            json_data = StringIO(response.text)
                            data_df = pd.read_json(json_data)
                        else:
                            try:
                                json_obj = json.loads(response.text)
                                data_df = pd.json_normalize(json_obj)
                                log.debug("Parsed as JSON despite invalid Content-Type")
                            except json.JSONDecodeError:
                                log.debug("Expected JSON but received:")
                                log.debug("Response content preview:")
                                raise ValueError("Unsupported Content-Type and invalid JSON")                                
                    else:  # If it's a file path
                        data_df = pd.read_json(filepath)

                    # Convert DataFrame to list of lists
                    data = data_df.values.tolist()                  
                    log.debug('--Data successfully read--')
                    # log.debug(data)
                except ValueError as e:
                    log.debug("Error parsing JSON")
                    data = []
                except json.JSONDecodeError as e:
                    log.debug("Error decoding JSON")
                    log.debug("Response content")
                    log.debug(response.content)
                    data = []
                except Exception as e:
                    log.debug("Unexpected error occurred")
                    data = []
            elif(mimetype == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'):
                log.debug('--Reading XLSX data--')
                try:
                    wb = load_workbook(tmp_file.name) #filename=tmp_file
                    # Get the last worksheet instead of the active one
                    last_sheet_name = wb.sheetnames[-1]
                    ws = wb[last_sheet_name]  # Get the last worksheet
                    # Iterate over rows and cells to read the data
                    records_read = 0
                    for row in ws.iter_rows(values_only=True):
                        data.append(row)
                        records_read += 1
                        if records_read >= n_rows:
                            break 
                except Exception as e:
                    log.debug("An error occurred, use CKAN datastore to readfile: %s", e)
                    data = []
                    # data = self._fetch_data_datastore_defined_row(self.resource)
            elif(mimetype == 'application/vnd.ms-excel'):
                log.debug('--Reading XLS data--')
                try:
                    file_bytes = tmp_file.read()
                    book = xlrd.open_workbook(file_contents=file_bytes)
                    # เลือก sheet สุดท้าย
                    last_sheet_index = book.nsheets - 1
                    sheet = book.sheet_by_index(last_sheet_index)
                    # sheet = book.sheet_by_index(0)
                    data = [sheet.row_values(i) for i in range(sheet.nrows)]
                except Exception as e:
                    log.debug("An error occurred, use CKAN datastore to readfile: %s", e)
                    # data = self._fetch_data_datastore_defined_row(self.resource)
            else:
                data = []
                log.debug('mimetype ไม่ตรง ไม่มีเครื่องมืออ่าน')
            # log.debug(data)
        finally:
            # ลบไฟล์ด้วยตัวเอง
            tmp_file.close()       # ปิดไฟล์ก่อน
            os.remove(tmp_file.name)  # ลบไฟล์
            # log.debug("----Temp file deleted:----")
            # log.debug(tmp_file.name)
            # log.debug("---dir tempfile----")
            # log.debug(os.listdir(tempfile.gettempdir()))
        return data
    # def _download_resource_from_url(self, url, headers=None):
        
    #     #---------------------------------------
    #     data = []
    #     log.debug('----resource format-----')
    #     filepath = url #self.resource['url']
    #     format_url = filepath.split(".")[-1]
    #     log.debug('----mimetype:_download_resource_from_url-----')
    #     mimetype = ResourceFetchData2.detect_mimetype(filepath)
    #     log.debug(filepath)
    #     log.debug(mimetype)  
    #     timeout = 5
    #     n_rows = 5001
    #     headers = {"User-Agent": "Mozilla/5.0"}     
    #     try:
    #         response = self.get_response(url, {})
    #         # response = requests.get(filepath, headers=headers, timeout=30)
    #         # response.raise_for_status()  # ถ้าโหลดไม่สำเร็จจะ raise error
    #     except requests.exceptions.RequestException as e:
    #         log.debug("Cannot download file ----:", e)
 
    #     if self.is_url_file(filepath,timeout) and response.status_code == 200 :
    #         log.debug('----read temp file-----')
    #         log.debug(mimetype)
    #         filename = url.split('/')[-1].split('#')[0].split('?')[0]
    #         tmp_file = tempfile.NamedTemporaryFile(suffix=filename)

    #         length = 0
    #         m = hashlib.md5()
    #         raw_data = b''

    #         for chunk in response.iter_content(CHUNK_SIZE):
    #             length += len(chunk)
    #             if length > MAX_CONTENT_LENGTH:
    #                 response.close()
    #                 raise DataTooBigError("File too large")
    #             tmp_file.write(chunk)
    #             raw_data += chunk
    #             m.update(chunk)

    #         response.close()
    #         tmp_file.seek(0)

    #         # --- Detect encoding ---
    #         result = chardet.detect(raw_data[:100000])  # ใช้แค่ 100KB แรกพอ
    #         encoding = result['encoding'] or 'utf-8'
    #         log.debug(f"Detected encoding: {encoding}")
    #         # response = self.get_response(url, {})
    #         # length = 0
    #         # m = hashlib.md5()
    #         # cl = None
    #         # for chunk in response.iter_content(CHUNK_SIZE):
    #         #     length += len(chunk)
    #         #     if length > MAX_CONTENT_LENGTH:
    #         #         raise DataTooBigError
    #         #     tmp_file.write(chunk)
    #         #     m.update(chunk)
    #         # response.close()
    #         # tmp_file.seek(0)


    #         if(mimetype == 'text/csv'): #(resource_format =='CSV'):
    #             log.debug('----csv----')     
    #             # log.debug(filepath)
    #             #-----------------------------------      
    #             try:
    #                 log.debug('--Reading CSV from temp--')
    #                 reader = csv.reader(io.TextIOWrapper(tmp_file, encoding=encoding, newline='')) #'utf-8'
    #                 records_read = 0
    #                 for row in reader:
    #                     data.append(row)
    #                     records_read += 1
    #                     if records_read >= n_rows:
    #                         break
    #                 # log.debug(data)
    #             except UnicodeDecodeError:
    #                 log.debug("An UnicodeDecodeError occurred")
    #                 data = self._fetch_data_datastore_defined_row(self.resource) 
    #             except Exception as e:
    #                 log.debug("An error occurred, use CKAN datastore to readfile: %s", e)
    #                 data = self._fetch_data_datastore_defined_row(self.resource) 
       
                
    #         elif(mimetype == 'application/json'): #elif(resource_format == 'JSON'):
    #             # log.debug('--Read JSON data--')
    #             # try:
    #             #     data_chunks = []
    #             #     log.debug('--chank--')
    #             #     for chunk in pd.read_json(filepath, lines=True, chunksize=n_rows):
    #             #         data_chunks.append(chunk)
    #             #         # Break the loop if the specified number of rows have been read
    #             #         if len(data_chunks) * n_rows >= n_rows:
    #             #             break
    #             #     # Concatenate the data chunks into a single DataFrame
    #             #     data_df = pd.concat(data_chunks, ignore_index=True)
    #             #     data = data_df.values.tolist()
    #             #     log.debug('--chank2--')  
    #             #     log.debug(data)            
    #             # except ValueError as e:
    #             #     log.debug("Error parsing JSON")
    #             #     log.debug(e)
    #             #     log.debug("Response content")
    #             #     log.debug(response.content)
    #             #     data = []
    #             #-------------\
    #             # data = []
    #             # try:
    #             #     # Read the entire JSON file into a DataFrame
    #             #     log.debug('--Reading JSON file--')
    #             #     data_df = pd.read_json(filepath)  # Read the entire JSON content into a DataFrame
                    
    #             #     # Convert DataFrame to list of lists
    #             #     data = data_df.values.tolist()
                    
    #             #     log.debug('--Data successfully read--')
    #             #     log.debug(data)
    #             data = []
    #             try:
    #                 log.debug('--Reading JSON data--')
                    
    #                 # Check if the filepath is a URL or a file path
    #                 if filepath.startswith('http'):  # If it's a URL
    #                     response = requests.get(filepath)
    #                     response.encoding = 'utf-8'  # Ensure correct encoding
    #                     content_type = response.headers.get('Content-Type', '')
    #                     log.debug(content_type)
    #                     # Ensure content type is JSON
    #                     if 'application/json' in content_type or 'application/octet-stream' in content_type:
    #                     # if 'application/json' in response.headers.get('Content-Type', ''):
    #                         # Use `pd.read_json` with a StringIO object for URL content
    #                         json_data = StringIO(response.text)
    #                         data_df = pd.read_json(json_data)
    #                     else:
    #                         try:
    #                             json_obj = json.loads(response.text)
    #                             data_df = pd.json_normalize(json_obj)
    #                             log.debug("Parsed as JSON despite invalid Content-Type")
    #                         except json.JSONDecodeError:
    #                             log.debug("Expected JSON but received:")
    #                             log.debug("Response content preview:")
    #                             raise ValueError("Unsupported Content-Type and invalid JSON")                                
    #                 else:  # If it's a file path
    #                     data_df = pd.read_json(filepath)

    #                 # Convert DataFrame to list of lists
    #                 data = data_df.values.tolist()                  
    #                 log.debug('--Data successfully read--')
    #                 # log.debug(data)
    #             except ValueError as e:
    #                 log.debug("Error parsing JSON")
    #                 data = []
    #             except json.JSONDecodeError as e:
    #                 log.debug("Error decoding JSON")
    #                 log.debug(e)
    #                 log.debug("Response content")
    #                 log.debug(response.content)
    #                 data = []
    #             except Exception as e:
    #                 log.debug("Unexpected error occurred")
    #                 log.debug(e)
    #                 data = []
    #         elif(mimetype == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'):
    #             log.debug('--Reading XLSX data--')
    #             try:
    #                 wb = load_workbook(filename=tmp_file)
    #                 # Get the last worksheet instead of the active one
    #                 last_sheet_name = wb.sheetnames[-1]
    #                 ws = wb[last_sheet_name]  # Get the last worksheet
    #                 # Iterate over rows and cells to read the data
    #                 records_read = 0
    #                 for row in ws.iter_rows(values_only=True):
    #                     data.append(row)
    #                     records_read += 1
    #                     if records_read >= n_rows:
    #                         break
    #                 # if  ResourceFetchData2.has_valid_filename(filepath,'.xlsx'):
    #                 #     log.debug('--load_workbook--')
    #                 #     # Load the workbook from the temporary file
    #                     # wb = load_workbook(filename=BytesIO(response.content), read_only=True)

    #                     # # Get the last worksheet instead of the active one
    #                     # last_sheet_name = wb.sheetnames[-1]
    #                     # ws = wb[last_sheet_name]  # Get the last worksheet
    #                     # # Iterate over rows and cells to read the data
    #                     # records_read = 0
    #                     # for row in ws.iter_rows(values_only=True):
    #                     #     data.append(row)
    #                     #     records_read += 1
    #                     #     if records_read >= n_rows:
    #                     #         break
    #                     # # log.debug(data)
    #                     # # log.debug("end readfile--")
    #                     # wb.close()
    #                 # else:
    #                 #     log.debug('--data store--')
    #                 #     data = self._fetch_data_datastore_defined_row(self.resource)           
                       
    #             except Exception as e:
    #                 log.debug("An error occurred, use CKAN datastore to readfile", e)
    #                 data = self._fetch_data_datastore_defined_row(self.resource)
    #                 # try:
    #                 #     data_df = pd.read_excel(filepath) 
    #                 #     if data_df is not None:
    #                 #         data = data_df.values.tolist()
    #                 #         data[:0] = [list(data_df.keys())]
    #                 # except Exception as e:
    #                 #     log.debug("An error occurred:", e)
    #                 #     data = []
    #         elif(mimetype == 'application/vnd.ms-excel'):
    #             log.debug('--Reading XLS data--')
    #             try:
    #                 file_bytes = tmp_file.read()
    #                 book = xlrd.open_workbook(file_contents=file_bytes)
    #                 # book = xlrd.open_workbook(filename=BytesIO(tmp_file))
    #                 # เลือก sheet สุดท้าย
    #                 last_sheet_index = book.nsheets - 1
    #                 sheet = book.sheet_by_index(last_sheet_index)
    #                 # sheet = book.sheet_by_index(0)
    #                 data = [sheet.row_values(i) for i in range(sheet.nrows)]

    #                 # if  ResourceFetchData2.has_valid_filename(filepath,'.xls'):
    #                 #     book = xlrd.open_workbook(file_contents=response.content)
    #                 #     # เลือก sheet สุดท้าย
    #                 #     last_sheet_index = book.nsheets - 1
    #                 #     sheet = book.sheet_by_index(last_sheet_index)
    #                 #     # sheet = book.sheet_by_index(0)
    #                 #     data = [sheet.row_values(i) for i in range(sheet.nrows)]
    #                 # else: 
    #                 #     log.debug('--data store--')
    #                 #     data = self._fetch_data_datastore_defined_row(self.resource)
    #             except Exception as e:
    #                 log.debug("An error occurred, use CKAN datastore to readfile", e)
    #                 data = self._fetch_data_datastore_defined_row(self.resource)
    #         else:
    #             data = []
    #     else:
    #         log.debug('----is_not_url_file-----')
    #         data = []
    #     return data

    def is_url_file(self,url,timeout):
        try:
            # Send a HEAD request to get headers
            response = requests.head(url, timeout=timeout)
            content_type = response.headers.get('Content-Type')
            
            if content_type:
                # Check if it's 'application/octet-stream'
                if content_type == 'application/octet-stream':
                    log.debug("Content-Type, application/octet-stream...")
                    # Use the URL extension to guess the MIME type
                    mime_type, _ = mimetypes.guess_type(url)
                    if mime_type:
                        log.debug("Guessed MIME type from URL:", mime_type)
                        return True  # It's a file based on the guessed MIME type
                    else:
                        log.debug("Cannot detect mimetype")
                        return True
                else:
                    log.debug("Content-Type: %s", content_type)
                    return True  # It's a valid file based on Content-Type
            else:
                # If there's no Content-Type, guess based on URL or file content
                log.debug("No Content-Type, guessing from URL...")
                mime_type, _ = mimetypes.guess_type(url)
                if mime_type:
                    log.debug("Guessed MIME type:" , mime_type)
                    return True
                else:
                    # As a fallback, attempt a GET request to analyze the content
                    response = requests.get(url, timeout=timeout, stream=True)
                    response.raise_for_status()
                    content_disposition = response.headers.get('Content-Disposition')
                    if content_disposition:
                        log.debug("Guessed from Content-Disposition: ",content_disposition)
                        return True
                    log.debug("Content-Type could not be determined.")
                    return False
        except Exception as e:
            log.debug("An error occurred:", e)
            return False  # Error occurred, not a file
    @staticmethod
    def has_valid_filename(url, format):
        parsed = urlparse(url)
        filename = os.path.basename(parsed.path)
        return filename and filename.lower() != format
    @staticmethod
    def detect_mimetype(url):
        try:
            log.debug("identify mimetype based on content-type")
            response = requests.head(url, timeout=5)
            content_type = response.headers.get('Content-Type', None)
            content_disposition = response.headers.get('Content-Disposition', None)

            mimetype = None
            charset = 'utf-8'
            filename = None

            # --- Parse Content-Type header ---
            if content_type:
                content_parts = content_type.split(";")
                mimetype = content_parts[0].strip()
                for part in content_parts[1:]:
                    if "charset=" in part:
                        charset = part.split("=")[-1].strip()
                        break

            # --- Parse Content-Disposition header for filename ---
            if content_disposition:
                log.debug(f"Content-Disposition: {content_disposition}")
                if 'filename=' in content_disposition:
                    filename = content_disposition.split('filename=')[-1].strip().strip('"')
                elif 'filename*=' in content_disposition:
                    filename = content_disposition.split("filename*=")[-1].split("''")[-1]
                log.debug(f"Downloaded filename: {filename}")

            # --- Try to get filename from URL if not available ---
            parsed_url = urlparse(url)
            filename_from_url = os.path.basename(parsed_url.path)
            log.debug(f"Filename from URL path: {filename_from_url}")

            # --- Select most reliable filename ---
            if filename and os.path.splitext(filename)[1]:
                filename_to_check = filename
            elif filename_from_url and os.path.splitext(filename_from_url)[1]:
                filename_to_check = filename_from_url
            else:
                # กรณีไม่มีชื่อไฟล์หรือไม่มีนามสกุลจริง
                filename_to_check = None

            # --- If filename_to_check ไม่มีค่า หรือ ไม่มีนามสกุล ให้ fallback ตามเงื่อนไข ---
            if not filename_to_check or not os.path.splitext(filename_to_check)[1]:
                # ตั้ง fallback ตามชื่อไฟล์ (ในกรณี source มีคำใน url หรือชื่อไฟล์แอบแฝง)
                url_lower = url.lower()
                fallback_extensions = [
                    'xlsx', 'xls', 'csv', 'json', 'geojson', 'pdf',
                    'doc', 'docx', 'ppt', 'pptx',
                    'jpg', 'jpeg', 'png', 'gif', 'tif', 'tiff', 'svg',
                    'xml', 'txt', 'zip', 'tar', 'gz'
                ]

                for ext in fallback_extensions:
                    if ext in url_lower:
                        filename_to_check = f"fallback.{ext}"
                        break
                
            # --- Check file extension to determine mimetype ---
            mimetype_map = {
                # Excel
                '.xls': 'application/vnd.ms-excel',
                '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',

                # CSV / Text / JSON
                '.csv': 'text/csv',
                '.json': 'application/json',
                '.geojson': 'application/geo+json',

                # PDF
                '.pdf': 'application/pdf',

                # Word documents
                '.doc': 'application/msword',
                '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',

                # PowerPoint
                '.ppt': 'application/vnd.ms-powerpoint',
                '.pptx': 'application/vnd.openxmlformats-officedocument.presentationml.presentation',

                # Images
                '.jpg': 'image/jpeg',
                '.jpeg': 'image/jpeg',
                '.png': 'image/png',
                '.gif': 'image/gif',
                '.tif': 'image/tiff',
                '.tiff': 'image/tiff',
                '.svg': 'image/svg+xml',

                # Plain text and XML
                '.txt': 'text/plain',
                '.xml': 'application/xml',

                # Compressed formats
                '.zip': 'application/zip',
                '.tar': 'application/x-tar',
                '.gz': 'application/gzip',
            }
            ext = os.path.splitext(filename_to_check)[1].lower()
            mimetype = mimetype_map.get(ext, 'application/octet-stream') #'application/octet-stream'
            # ext = os.path.splitext(filename_to_check)[1].lower()
            # if ext == '.xls':
            #     mimetype = 'application/vnd.ms-excel'
            # elif ext == '.xlsx':
            #     mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            # elif ext == '.csv':
            #     mimetype = 'text/csv'
            # elif ext == '.json':
            #     mimetype = 'application/json'
            # elif ext == '.pdf':
            #     mimetype = 'application/pdf'
            # else:
            #     mimetype = 'application/octet-stream'  # fallback

            log.debug(f"Determined mimetype from extension '{ext}': {mimetype}")

        except Exception as e:
            log.debug("Error fetching header:")
            log.debug(e)
            mimetype = None

        return mimetype
    def detect_encoding(self,url):
        timeout = 5
        sample_size=1024
        try:
            response = requests.get(url, timeout=timeout)
            if response.status_code == 200:
                # Use chardet to detect the encoding of the sample
                sample = response.content[:sample_size]
                result = chardet.detect(sample)
                return result['encoding']
            else:
                log.debug(f"Failed to download the file. Status code: {response.status_code}")
                return None
        except Exception as e:
            log.debug(f"Error: {e}")
            return None   
    def _download_resource_from_ckan(self, resource):       
        upload = uploader.get_resource_uploader(resource)
        filepath = upload.get_path(resource['id'])
        # filepath = resource['url']
        resource_format = resource['format']
        try:
            data = []
            if os.path.isfile(filepath):
                if '\0' in open(filepath).read():
                    log.debug("you have null bytes in your input file")
                else:
                    log.debug("you don't")
                if(resource_format == 'XLSX' or resource_format == 'XLS'):                  
                    data_df = pd.read_excel(filepath)
                    data = data_df.values.tolist()
                elif(resource_format == 'JSON'):
                    try:
                        data_json = pd.read_json(filepath)
                        data_df = pd.DataFrame(data_json)
                        data = data_df.values.tolist()
                        data[:0] = [list(data_df.keys())]
                        log.debug(list(data_df.keys()))
                    except ValueError as e:
                        log.debug('ValueError = ', e)
                        data = []
                elif(resource_format == 'CSV'):
                    with open(filepath) as csvf:
                        reader = csv.reader(csvf)
                        for row in reader:
                            data.append(row)
                else:
                    data = []
                return data
            # if self.is_url_file(filepath):
            #     response = requests.get(filepath)
            #     response.raise_for_status()  # Raise an exception for bad status codes
            #     if(resource_format == 'CSV'):
            #         data = response.content.decode('utf-8')  # Decode content to string
            #         data_df = pd.read_csv(io.StringIO(data))  # Read CSV data into a DataFrame
            #         if data_df is not None:
            #             data = data_df.values.tolist()
            #             data[:0] = [list(data_df.keys())]
                        
            #     elif(resource_format == 'JSON'):
            #         try:
            #             data_json = response.json()
            #             data_json = pd.read_json(filepath)
            #             if data_json is not None:
            #                 data_df = pd.DataFrame(data_json)
            #                 data = data_df.values.tolist()
            #                 data[:0] = [list(data_df.keys())]
                                  
            #         except ValueError as e:
            #                 print('ValueError = ', e)
            #                 data = []

            #     elif(resource_format == 'XLSX' or resource_format == 'XLS'): 
            #         try:
            #             excel_data = io.BytesIO(response.content)         
            #             dataframe = openpyxl.load_workbook(excel_data)  # Load Excel file using openpyxl
            #             # Define variable to read sheet
            #             dataframe1 = dataframe.active
            #             # Iterate the loop to read the cell values
            #             for row in range(0, dataframe1.max_row):
            #                 data_row = []
            #                 for col in dataframe1.iter_cols(1, dataframe1.max_column):
            #                     data_row.append(col[row].value)
            #                 data.append(data_row)

            #         except Exception as e:
            #             print("An error occurred:", e)
            #             data = []
            #     else:
            #         data = []
            #     return data
            else:
                # The worker is not in the same machine as CKAN, so it cannot
                # read the resource files from the local file system.
                # We need to retrieve the resource data from the resource URL.
                if 'url' not in resource:
                    raise Exception('Cannot access the resource because the '
                                    'resource URL is not set.')
                log.debug('Resource data is not in this file system '
                          '(File %s does not exist).'
                          'Fetching data from CKAN download url directly.',
                          filepath)
                headers = None
                # sysadmin_api_key = _get_sysadmin_user_key()
                # if sysadmin_api_key:
                #     headers = {'Authorization': sysadmin_api_key}
                return self._download_resource_from_url(
                    resource['url'],
                    resource['format'],
                    headers
                )
        except OSError as e:
            log.error('Failed to download resource from CKAN upload. '
                      'Error: %s', str(e))
            log.exception(e)
            raise e

    def _fetch_data_directly(self):
        if self.resource.get('url_type') == 'upload':
            log.debug('2 Getting data from CKAN...')
            # return self._download_resource_from_ckan(self.resource)
            return self._download_resource_from_url(self.resource['url'], self.resource['format'])
        if self.resource.get('url'):
            log.debug('2 Getting data from remote URL...')
            return self._download_resource_from_url(self.resource['url'], self.resource['format'])
        raise Exception('Resource {} is not available '
                        'for download.'.format(self.resource.get('id')))

    def fetch_page2(self, page, limit):
        '''Fetches one page (with size of `limit`) of data from the resource
        CSV data.

        This call will first try to load the data from the datastore. If that
        fails and the data is not available in the data store, then it will try
        to load the data directly (from CKAN uploads or by downloading the CSV
        file directly from the remote server). Once this happens, and the data
        is loaded into `ResourceCSVData`, then every subsequent call will load
        the data from the cached instance of `ResourceCSVData` and will not try
        to load it from the data store.

        :param page: `int`, the page to fetch, 0-based.
        :param limit: `int`, page size.

        :returns: `dict`, the requested page as `dict` containing:
            * `total` - total number of records in the data.
            * `records` - a `list` of records for this page. Each record is a
                `dict` of format `<column_name>:<row_value>`.
            * `fields` - the records metadata describing each column. It is a
                `list` of `dict`, where each element describes one column. The
                order of the columns is the same as it is appears in the CSV
                file. Each element has:
                    * `id` - `str`, the name of the column
                    * `type` - `str`, the column type (ex. `numeric`, `text`)
        '''
        log.debug('----call fetch_page at ResourceFetchData2-------')
        # log.debug(self.download_resource)
        if self.download_resource:
            if not self.resource_csv:
                self.resource_csv = ResourceCSVData(self._fetch_data_directly)
                log.debug('2Resource data downloaded directly.')
            return self.resource_csv.fetch_page(page, limit)
        try:
            #------ Pang Edit ------------------       
            self.download_resource = True
            log.debug('Will try to download the data directly.')
            return self.fetch_page2(page, limit)
        except Exception as e:
            log.warning('2 Failed to load resource data from URL. '
                        'Error: %s', str(e))
            log.exception(e)
            self.download_resource = True
            log.debug('Will try to download the data directly.')
            return self.fetch_page2(page, limit)
class Openness():#DimensionMetric
    '''Calculates the openness Data Qualtiy dimension.

    The calculation is performed over all values in the resource data.
    In each row, ever cell is inspected if there is a value present in it.

    The calculation is: `cells_with_value/total_numbr_of_cells * 100`, where:
        * `cells_with_value` is the number of cells containing a value. A cell
            contains a value if the value in the cell is not `None` or an empty
            string or a string containing only whitespace.
        * `total_numbr_of_cells` is the total number of cells expected to be
            populted. This is calculated from the number of rows multiplied by
            the number of columns in the tabular data.

    The return value is a percentage of cells that are populated from the total
    number of cells.
    '''
    
    
    def __init__(self):
        self.name = 'openness'

    def get_openness_score(self,data_format,mimetype):
        openness_score = { 
        "TTL": 5, "RDF": 5, "JSON-LD": 5,"N3": 5, "SPARQL": 5, 
        "KML": 3, "GML": 3, "WCS": 3, "NetCDF": 3,#Geo
        "TSV": 3, "WFS": 3, "KMZ": 3, "QGIS": 3,  #Geo
        "WMS": 3, "WMTS": 3,"XYZ": 3,#Geo
        "RSS": 3, "Atom Feed": 3,
        "GeoJSON": 3, 
        "XML": 3,"ODG": 3,
        "CSV": 3, "JSON": 3, "ODS": 3,"ODB": 3, "ODF": 3,#open
        "ArcGIS Map Service": 2,  #Geo
        "ArcGIS Map Preview": 2,  #Geo
        "dBase": 2, "SHP": 2, "Esri REST": 2, #Geo
        "XLS": 2,"XLSX": 2,"MDB": 2, #microsoft access
        "GIF": 1, "TIFF": 1, "JPEG": 1, "BMP": 1,"SVG": 1,"PNG": 1,
        "ODT": 1, "ODP": 1,"ODC": 1,
        "BIN": 1,  "TAR": 1, "ZIP": 1, "GZ": 1, "RAR": 1, 
        "PDF": 1,"DOCX": 1,"DOC": 1 ,"PPT": 1,"PPTX": 1,"TXT": 1,
        "HTML": 1
        }
        mimetype_to_format = {
        'text/csv': 'CSV',
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': 'XLSX',
        'application/vnd.ms-excel': 'XLS',
        'application/pdf': 'PDF',
        'application/rdf+xml': 'RDF',
        'application/ld+json': 'JSON-LD',
        'application/json': 'JSON',
        'application/xml': 'XML',
        'text/xml': 'XML',
        'application/vnd.google-earth.kml+xml': 'KML',
        'application/vnd.google-earth.kmz': 'KMZ',
        'application/gml+xml': 'GML',
        'image/png': 'PNG',
        'image/jpeg': 'JPEG',
        'image/bmp': 'BMP',
        'image/gif': 'GIF',
        'image/tiff': 'TIFF',
        'image/svg+xml': 'SVG',
        'application/zip': 'ZIP',
        'application/gzip': 'GZ',
        'application/x-tar': 'TAR',
        'application/msword': 'DOC',
        'application/vnd.openxmlformats-officedocument.wordprocessingml.document': 'DOCX',
        'application/vnd.ms-powerpoint': 'PPT',
        'application/vnd.openxmlformats-officedocument.presentationml.presentation': 'PPTX',
        'application/vnd.oasis.opendocument.text': 'ODT',
        'application/vnd.oasis.opendocument.spreadsheet': 'ODS',
        'application/vnd.oasis.opendocument.presentation': 'ODP',
        'application/vnd.oasis.opendocument.chart': 'ODC',
        'application/vnd.oasis.opendocument.formula': 'ODF',
        'application/vnd.oasis.opendocument.database': 'ODB',
        'application/x-dbf': 'dBase',
        'application/x-msaccess': 'MDB',
        'text/plain': 'TXT',
        'text/html': 'HTML',
        'application/vnd.rar': 'RAR',
        }
        # Default score
        score = 0

        # If data_format is missing, try to get it from mimetype
        if not data_format:
            data_format = mimetype_to_format.get(mimetype, '')
        
        if data_format:
            score = openness_score.get(data_format.upper(), 0)

        return score
        # openness_score = { "JSON-LD": 5,"N3": 5, "SPARQL": 5, "RDF": 5,
        # "TTL": 5, "KML": 3, "GML": 3, "WCS": 3, "NetCDF": 3,
        # "TSV": 3, "WFS": 3, "KMZ": 3, "QGIS": 3,
        # "ODS": 3, "JSON": 3,"ODB": 3, "ODF": 3,
        # "ODG": 3, "XML": 3,"WMS": 3, "WMTS": 3,
        # "SVG": 3, "JPEG": 3,"CSV": 3, "Atom Feed": 3,
        # "XYZ": 3, "PNG": 3,"RSS": 3, "GeoJSON": 3,
        # "IATI": 3, "ICS": 3,"XLS": 2, "MDB": 2,
        # "ArcGIS Map Service": 2,"BMP": 2, "TIFF": 2,
        # "XLSX": 2, "GIF": 2,"E00": 2, "MrSID": 2,
        # "ArcGIS Map Preview": 2,"MOP": 2, "Esri REST": 2,
        # "dBase": 2, "SHP": 2,"PPTX": 1, "DOC": 1,
        # "ArcGIS Online Map": 1, "ZIP": 1, "GZ": 1,
        # "ODT": 1, "RAR": 1,"TXT": 1, "DCR": 1,
        # "DOCX": 1, "BIN": 1,"PPT": 1, "ODP": 1,
        # "PDF": 1, "ODC": 1,"MXD": 1, "TAR": 1,"EXE": 0,
        # "JS": 0,"Perl": 0,"OWL": 0, "HTML": 0,
        # "XSLT": 0, "RDFa": 0}

        # #if user add a resource as a link, data type will be null
        # if(data_format == ''):
        #     score = 0
        #     if(mimetype == 'text/csv'):
        #         data_format = 'CSV'    
        #     elif(mimetype == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'):
        #         data_format = 'XLSX'
        #     elif(mimetype == 'application/vnd.ms-excel'):
        #         data_format = 'XLS'
        #     elif(mimetype == 'application/pdf'):
        #         data_format = 'PDF'
        #     elif(mimetype == 'application/rdf+xml'):
        #         data_format = 'RDF'
        #     elif(mimetype == 'application/ld+json'):
        #         data_format = 'JSON-LD'
        #     elif(mimetype == 'application/xml'):
        #         data_format = 'XML'           
        #     elif(mimetype == 'application/vnd.google-earth.kml+xml'):
        #         data_format = 'KML'
        #     elif(mimetype == 'application/gml+xml'):
        #         data_format = 'GML'          
        #     elif(mimetype == 'application/json'):
        #         data_format = 'JSON'
        #     elif(mimetype == 'image/png'):
        #         data_format = 'PNG'
        #     elif(mimetype == 'image/jpeg'):
        #         data_format = 'JPEG'
        #     elif(mimetype == 'image/bmp'):
        #         data_format = 'BMP'
        #     elif(mimetype == 'image/gif'):
        #         data_format = 'GIF'
        #     elif(mimetype == 'image/tiff'):
        #         data_format = 'TIFF'
        #     elif(mimetype == 'application/zip'):
        #         data_format = 'ZIP'
        #     elif(mimetype == 'application/msword'):
        #         data_format = 'DOC'
        #     elif(mimetype == 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'):
        #         data_format = 'DOCX'
        #     elif(mimetype == 'application/vnd.oasis.opendocument.text'):
        #         data_format = 'ODT'
        #     elif(mimetype == 'application/vnd.oasis.opendocument.spreadsheet'):
        #         data_format = 'ODS'
        #     elif(mimetype == 'application/vnd.oasis.opendocument.presentation'):
        #         data_format = 'ODP'
        #     elif(mimetype == 'application/vnd.ms-powerpoint'):
        #         data_format = 'PPT'
        #     elif(mimetype == 'application/vnd.openxmlformats-officedocument.presentationml.presentation'):
        #         data_format = 'PPTX'
        #     elif(mimetype == 'text/html'):
        #         data_format = 'HTML'
        #     elif(mimetype == 'application/vnd.rar'):
        #         data_format = 'RAR'
        #     elif(mimetype == 'text/plain'):
        #         data_format = 'TXT'
        #     if(data_format != '' ):
        #         score =  openness_score.get(data_format)
        # else:
        #     score =  openness_score.get(data_format)
        # # data type is not in the list
        # if (score == None):
        #     score = 0
        # return score
    def calculate_metric(self, resource):
        '''Calculates the openness dimension metric for the given resource
        from the resource data.

        :param resource: `dict`, CKAN resource.
        :param data: `dict`, the resource data as a dict with the following
            values:
                * `total`, `int`, total number of rows.
                * `fields`, `list` of `dict`, column metadata - name, type.
                * `records`, `iterable`, iterable over the rows in the resource
                    where each row is a `dict` itself.

        :returns: `dict`, the report contaning the calculated values:
            * `value`, `float`, the percentage of complete values in the data.
            * `total`, `int`, total number of values expected to be populated.
            * `complete`, `int`, number of cells that have value.
        '''
        data_format = resource['format']
        mimetype    = resource['mimetype']
    
        if ((data_format == '' or data_format == None) and (mimetype == '' or mimetype == None)):
            resource_url = resource['url']   
            format_url = resource_url.split(".")[-1]
            data_format = format_url.upper()

        data_format = data_format.replace(".", "")
        data_format = data_format.upper()  
        openness_score = self.get_openness_score(data_format,mimetype)
        # log.debug('Openness score: %f%%', openness_score)
        return {
            'format': data_format,
            'value': openness_score,
            'mimetype':mimetype
        }
        
    def calculate_cumulative_metric(self, resources, metrics):
        '''Calculates the cumulative report for all resources from the
        calculated results for each resource.

        The calculation is done as `all_complete/all_total * 100`, where
            * `all_complete` is the total number of completed values in all
                resources.
            * all_total is the number of expected values (rows*columns) in all
                resources.
        The final value is the percentage of completed values in all resources
        in the dataset.

        :param resources: `list` of CKAN resources.
        :param metrics: `list` of `dict` results for each resource.

        :returns: `dict`, a report for the total percentage of complete values:
            * `value`, `float`, the percentage of complete values in the data.
            * `total`, `int`, total number of values expected to be populated.
            * `complete`, `int`, number of cells that have value.
        '''
        # log.debug('-------------Calculate openness metrics -----------')
        # log.debug(metrics)
        openness_list = []
        total = 0
        
        for item_metric in metrics:
            #check dict is not Empty
            if item_metric:                  
                openness_score = item_metric.get('value')
                if openness_score is not None:
                    total = total+openness_score
                    openness_list.append(openness_score)
        if openness_list:
            result_score = max(openness_list)
            return {
                'total': total,
                'value': result_score,
            }
        else:
            return {
                'total': 0,
                'value': 0,
            }
class Downloadable():#DimensionMetric
    '''Calculates the Downloadable Data Qualtiy dimension.

    The calculation is performed over all values in the resource data.
    In each row, ever cell is inspected if there is a value present in it.

    The calculation is: `cells_with_value/total_numbr_of_cells * 100`, where:
        * `cells_with_value` is the number of cells containing a value. A cell
            contains a value if the value in the cell is not `None` or an empty
            string or a string containing only whitespace.
        * `total_numbr_of_cells` is the total number of cells expected to be
            populted. This is calculated from the number of rows multiplied by
            the number of columns in the tabular data.

    The return value is a percentage of cells that are populated from the total
    number of cells.
    '''

    def __init__(self):
        self.name = 'downloadable'

    def is_downloadable(self,url):
        try:
            response = requests.head(url, allow_redirects=True)
            if 'Content-Type' in response.headers:
                content_type = response.headers['Content-Type'].lower()
                if 'text' not in content_type or 'csv' in content_type or 'application/octet-stream' in content_type:
                    return True
            # Check if the URL ends with common file extensions
            if os.path.splitext(url)[1].lower() in ['.zip', '.pdf', '.jpg', '.jpeg', '.png', '.gif', '.doc', '.docx', '.xls', '.xlsx', '.ppt', '.pptx','.gml','.kml','.kmz','.json','.geojson','.shp','.wms','.xml']:
                return True
            return False
        except requests.RequestException:
            return False
    def calculate_metric(self, resource):
        '''Calculates the openness dimension metric for the given resource
        from the resource data.

        :param resource: `dict`, CKAN resource.
        :param data: `dict`, the resource data as a dict with the following
            values:
                * `total`, `int`, total number of rows.
                * `fields`, `list` of `dict`, column metadata - name, type.
                * `records`, `iterable`, iterable over the rows in the resource
                    where each row is a `dict` itself.

        :returns: `dict`, the report contaning the calculated values:
            * `value`, `float`, the percentage of complete values in the data.
            * `total`, `int`, total number of values expected to be populated.
            * `complete`, `int`, number of cells that have value.
        '''
        # log.debug ('------------downloadable--')
        # log.debug(resource)
        downloadable_score = 1
        resource_data_format = resource['format'] 
        resource_url    = resource['url']
        
        if self.is_downloadable(resource_url):
            downloadable_score = 1 # 2 can download
            #ตรวจสอบ format ไม่ตรงตามที่กำหนด แต่เว้นว่างได้ เลยตรวจยาก เพราะเค้าอาจจะไม่กำหนดก็ได้
            # if(pd.notna(resource_url) and resource_data_format != ""): 
            #     format_url = resource_url.split(".")[-1]
            #     lower_format = resource_data_format.lower()
            #     if(format_url != lower_format): # set wrong format type
            #         downloadable_score = 0
        else: 
            downloadable_score = 0 # 1 cannot download such as HTML.

        return {
            'format': resource_data_format,
            'value': downloadable_score,
        }
        
    def calculate_cumulative_metric(self, resources, metrics):
        '''Calculates the cumulative report for all resources from the
        calculated results for each resource.

        The calculation is done as `all_complete/all_total * 100`, where
            * `all_complete` is the total number of completed values in all
                resources.
            * all_total is the number of expected values (rows*columns) in all
                resources.
        The final value is the percentage of completed values in all resources
        in the dataset.

        :param resources: `list` of CKAN resources.
        :param metrics: `list` of `dict` results for each resource.

        :returns: `dict`, a report for the total percentage of complete values:
            * `value`, `float`, the percentage of complete values in the data.
            * `total`, `int`, total number of values expected to be populated.
            * `complete`, `int`, number of cells that have value.
        '''
        # log.debug('-------------Calculate downloadable metrics -----------')
        # log.debug(metrics)
        # downloadable_list = []
        # total = 0
        # for item_metric in metrics:
        #     #check dict is not Empty
        #     if item_metric:
        #         downloadable_score = item_metric.get('value')
        #         total = total+downloadable_score
        #         downloadable_list.append(downloadable_score)
        
        # if downloadable_list:
        #     result_score = max(downloadable_list)
        #     return {
        #         'total': total,
        #         'value': result_score,
        #     }
        # else:
        #     return {
        #         'total': 0,
        #         'value': 0,
        #     }
        N_total = len(resources)
        N_downloadable = 0 
        for item_metric in metrics:
            #check dict is not Empty
            if item_metric and isinstance(item_metric.get('value'), (int, float)):
                val = item_metric.get('value')
                if val:
                    N_downloadable += val
                
            else:
                N_total -= 1

        downloadable_score = (N_downloadable / N_total * 100) if N_total > 0 else 0.0
        return {
            "N_total": N_total,
            "N_available": N_downloadable,
            "value": round(downloadable_score, 2)
        }
class AccessAPI():#DimensionMetric
    '''Calculates the Downloadable Data Qualtiy dimension.

    The calculation is performed over all values in the resource data.
    In each row, ever cell is inspected if there is a value present in it.

    The calculation is: `cells_with_value/total_numbr_of_cells * 100`, where:
        * `cells_with_value` is the number of cells containing a value. A cell
            contains a value if the value in the cell is not `None` or an empty
            string or a string containing only whitespace.
        * `total_numbr_of_cells` is the total number of cells expected to be
            populted. This is calculated from the number of rows multiplied by
            the number of columns in the tabular data.

    The return value is a percentage of cells that are populated from the total
    number of cells.
    '''

    def __init__(self):
        self.name = 'access_api'
    def check_api(self, url):
        log.debug("--check_api--")
        try:
            response = requests.get(url)
            log.debug(response.status_code)
            
            if response.ok:
                try:
                    data = response.json()
                    log.debug("Valid JSON Response:")
                    # log.debug(data)
                    return True
                except ValueError:
                    log.debug("Response is not valid JSON.")
                    return False
            else:
                log.debug("Error Response: ")
                log.debug(response.status_code)
                return False
        except requests.exceptions.RequestException as e:
            log.debug("Request Exception: ")
            log.debug(e)
            return False
    def calculate_metric(self, resource):
        '''Calculates the API Accesssibility dimension metric for the given resource
        from the resource data.

        :param resource: `dict`, CKAN resource.
        :param data: `dict`, the resource data as a dict with the following
            values:
                * `total`, `int`, total number of rows.
                * `fields`, `list` of `dict`, column metadata - name, type.
                * `records`, `iterable`, iterable over the rows in the resource
                    where each row is a `dict` itself.

        :returns: `dict`, the report contaning the calculated values:
            * `value`, `float`, the percentage of complete values in the data.
            * `total`, `int`, total number of values expected to be populated.
            * `complete`, `int`, number of cells that have value.
        '''
        log.debug ('-----Access API-----')
        access_api_score = 0 
        log.debug (resource['format'])
        if(resource['datastore_active'] == True):
            access_api_score = 1
        elif(resource['datastore_active'] == False): 
            access_api_score = 0
        elif(resource['format'] == 'API'):
            is_valid_api = self.check_api(resource['url'])
            # is_valid_api = self.check_api(resource['format'])
            if is_valid_api:
                access_api_score = 1 #2
                log.debug ("The URL is a valid API endpoint.")
            else:
                access_api_score = 0
                log.debug ("The URL is not a valid API endpoint.")
        return {
            'datastore': resource['datastore_active'],
            'format': resource['format'],
            'value': access_api_score
        }
   
    def calculate_cumulative_metric(self, resources, metrics):
        '''Calculates the cumulative report for all resources from the
        calculated results for each resource.

        The calculation is done as `all_complete/all_total * 100`, where
            * `all_complete` is the total number of completed values in all
                resources.
            * all_total is the number of expected values (rows*columns) in all
                resources.
        The final value is the percentage of completed values in all resources
        in the dataset.

        :param resources: `list` of CKAN resources.
        :param metrics: `list` of `dict` results for each resource.

        :returns: `dict`, a report for the total percentage of complete values:
            * `value`, `float`, the percentage of complete values in the data.
            * `total`, `int`, total number of values expected to be populated.
            * `complete`, `int`, number of cells that have value.
        '''
        # log.debug('-------------Calculate accessibility metrics -----------')
        # log.debug(metrics)
        # access_api_list = []
        # total = 0
        
        # for item_metric in metrics:
        #     #check dict is not Empty
        #     if item_metric:
        #         access_api_score = item_metric.get('value')
        #         total = total+access_api_score
        #         access_api_list.append(access_api_score)
        # if access_api_list:
        #     result_score = max(access_api_list)
        #     return {
        #         'total': total,
        #         'value': result_score,
        #     }
        # else:
        #     return {
        #         'total': 0,
        #         'value': 0,
        #     }
        N_total = len(resources)
        N_access_api = 0 #sum(1 for r in resources if r.get('value') == 1)
        for item_metric in metrics:
            #check dict is not Empty
            if item_metric and isinstance(item_metric.get('value'), (int, float)):
                val = item_metric.get('value')
                if val:
                    N_access_api += val
                
            else:
                N_total -= 1

        access_api_score = (N_access_api / N_total * 100) if N_total > 0 else 0.0
        #     if item_metric:
        #         access_api_value = item_metric.get('value')
        #         N_access_api = N_access_api+access_api_value
        # access_api_score = (N_access_api / N_total * 100) if N_total > 0 else 0.0

        return {
            "N_total": N_total,
            "N_available": N_access_api,
            "value": round(access_api_score, 2)
        }
class Availability():
    '''Calculates the Availability Data Qualtiy dimension.

    The calculation is performed over all values in the resource data.
    In each row, ever cell is inspected if there is a value present in it.

    The calculation is: `cells_with_value/total_numbr_of_cells * 100`, where:
        * `cells_with_value` is the number of cells containing a value. A cell
            contains a value if the value in the cell is not `None` or an empty
            string or a string containing only whitespace.
        * `total_numbr_of_cells` is the total number of cells expected to be
            populted. This is calculated from the number of rows multiplied by
            the number of columns in the tabular data.

    The return value is a percentage of cells that are populated from the total
    number of cells.
    '''

    def __init__(self):
        self.name = 'availability'

    def calculate_metric(self, resource):
        '''Calculates the openness dimension metric for the given resource
        from the resource data.

        :param resource: `dict`, CKAN resource.
        :param data: `dict`, the resource data as a dict with the following
            values:
                * `total`, `int`, total number of rows.
                * `fields`, `list` of `dict`, column metadata - name, type.
                * `records`, `iterable`, iterable over the rows in the resource
                    where each row is a `dict` itself.

        :returns: `dict`, the report contaning the calculated values:
            * `value`, `float`, the percentage of complete values in the data.
            * `total`, `int`, total number of values expected to be populated.
            * `complete`, `int`, number of cells that have value.
        '''
        #---------- Setting --------------------
        access   = AccessAPI()       # สร้าง object ของ AccessAPI
        download = Downloadable()    # สร้าง object ของ Downloadable
        #---------- Downloadable --------------------
        downloadable_score = 1
        resource_data_format = resource['format'] 
        resource_url    = resource['url']
        
        if download.is_downloadable(resource_url):
            downloadable_score = 1 
        else: 
            downloadable_score = 0 

        #---------- Access Api -----------------------
        
        access_api_score = 1 
        log.debug (resource['format'])
        if(resource['datastore_active'] == True):
            access_api_score = 1
        elif(resource['datastore_active'] == False): 
            access_api_score = 0
        elif(resource['format'] == 'API'):
            is_valid_api = access.check_api(resource['url'])
            if is_valid_api:
                access_api_score = 1 
                log.debug ("The URL is a valid API endpoint.")
            else:
                access_api_score = 0
                log.debug ("The URL is not a valid API endpoint.")
        else:
            access_api_score = 0

        #----------------------------------------------
        availability_score = 1
        if downloadable_score == 1 or access_api_score == 1 :
            availability_score = 1
        else:
            availability_score = 0
        return {
            'downloadable': downloadable_score,
            'access_api': access_api_score,
            'value': availability_score,
        }
        
    def calculate_cumulative_metric(self, resources, metrics):
        '''Calculates the cumulative report for all resources from the
        calculated results for each resource.

        The calculation is done as `all_complete/all_total * 100`, where
            * `all_complete` is the total number of completed values in all
                resources.
            * all_total is the number of expected values (rows*columns) in all
                resources.
        The final value is the percentage of completed values in all resources
        in the dataset.

        :param resources: `list` of CKAN resources.
        :param metrics: `list` of `dict` results for each resource.

        :returns: `dict`, a report for the total percentage of complete values:
            * `value`, `float`, the percentage of complete values in the data.
            * `total`, `int`, total number of values expected to be populated.
            * `complete`, `int`, number of cells that have value.
        '''
        N_total = len(resources)
        N_available = 0 #sum(1 for r in resources if r.get('value') == 1)
        for item_metric in metrics:
            if item_metric and isinstance(item_metric.get('value'), (int, float)):
                val = item_metric.get('value')
                if val:
                    N_available += val
                
            else:
                N_total -= 1

        availability_score = (N_available / N_total * 100) if N_total > 0 else 0.0
        # for item_metric in metrics:
        #     #check dict is not Empty
        #     if item_metric:
        #         available_score = item_metric.get('value')
        #         N_available = N_available+available_score
        # availability_score = (N_available / N_total * 100) if N_total > 0 else 0.0

        return {
            "N_total": N_total,
            "N_available": N_available,
            "value": round(availability_score, 2)
        }
class EncodingUTF8():
    '''Calculates the Encoding UTF-8 dimension.

    The calculation is performed over all values in the resource data.
    In each row, ever cell is inspected if there is a value present in it.

    The calculation is: `cells_with_value/total_numbr_of_cells * 100`, where:
        * `cells_with_value` is the number of cells containing a value. A cell
            contains a value if the value in the cell is not `None` or an empty
            string or a string containing only whitespace.
        * `total_numbr_of_cells` is the total number of cells expected to be
            populted. This is calculated from the number of rows multiplied by
            the number of columns in the tabular data.

    The return value is a percentage of cells that are populated from the total
    number of cells.
    '''

    def __init__(self):
        self.name = 'utf8'
    def calculate_metric(self,resource,data): 
        return {
            'consistency': 0,
            'validity': '',
            'encoding': '',
            'value': 0
        }
    def calculate_metric_utf8(self,resource,validity_report): 
        '''Calculates the openness dimension metric for the given resource
        from the resource data.

        :param resource: `dict`, CKAN resource.
        :param data: `dict`, the resource data as a dict with the following
            values:
                * `total`, `int`, total number of rows.
                * `fields`, `list` of `dict`, column metadata - name, type.
                * `records`, `iterable`, iterable over the rows in the resource
                    where each row is a `dict` itself.

        :returns: `dict`, the report contaning the calculated values:
            * `value`, `float`, the percentage of complete values in the data.
            * `total`, `int`, total number of values expected to be populated.
            * `complete`, `int`, number of cells that have value.
        '''
        
        #--------------Machine Readable-----------------
        machine_readable_format = ['CSV','XLSX','XLS','JSON','XML'] #'JSON','XML' 
        document_format = ['PDF','DOC','DOCX','PPTX','PPT','ODT','ODS','ODP'] #-50
        image_format = ['PNG','JPEG','GIF','TIFF']#-60
        openness_5_star_format = ['RDF','TTL','N3','GeoJSON','WMS','GML','KML','SHP','Esri REST']#100
        validity_chk = True
        encoding_utf8 = 0
        data_format = resource['format'] 
        mimetype    = resource['mimetype'] 
        file_type = ''  
      
        #if data format and mimetype is null, find format from url
        if ((data_format == '' or data_format == None) and (mimetype == '' or mimetype == None)):
            resource_url = resource['url']   
            format_url = resource_url.split(".")[-1]
            data_format = format_url.upper()
            # log.debug(data_format)
        #---- Tabular Data -------------#
        data_format = data_format.replace(".", "")
        data_format = data_format.upper()
        
        if data_format in machine_readable_format:
            encoding = validity_report.get('encoding')     
            log.debug('-----tabular_format------')
            file_type = 'tabular'
            if encoding and "utf-8" in encoding.lower():
                encoding_utf8 = 1
            else:
                encoding_utf8 = 0
            log.debug(encoding_utf8) 
        elif data_format in openness_5_star_format:
            file_type = 'spacial types'
        elif data_format in document_format:
            file_type = 'document'
        elif data_format in image_format:
            file_type = 'image'
        else:
            file_type = 'other'
        
        return {
            'format': data_format,
            'file_type': file_type,
            'value': encoding_utf8
        }
    def calculate_cumulative_metric(self, resources, metrics):
        '''Calculates the cumulative report for all resources from the
        calculated results for each resource.

        The calculation is done as `all_complete/all_total * 100`, where
            * `all_complete` is the total number of completed values in all
                resources.
            * all_total is the number of expected values (rows*columns) in all
                resources.
        The final value is the percentage of completed values in all resources
        in the dataset.

        :param resources: `list` of CKAN resources.
        :param metrics: `list` of `dict` results for each resource.

        :returns: `dict`, a report for the total percentage of complete values:
            * `value`, `float`, the percentage of complete values in the data.
            * `total`, `int`, total number of values expected to be populated.
            * `complete`, `int`, number of cells that have value.
        '''
        # log.debug('-------------Calculate machine metrics -----------')
        N_total = len(resources)
        N_encoding_utf8 = 0 

        for item_metric in metrics:
            if item_metric and isinstance(item_metric.get('value'), (int, float)):
                val = item_metric.get('value')
                if val:
                    N_encoding_utf8 += val
                
            else:
                N_total -= 1

        encoding_utf8_score = (N_encoding_utf8 / N_total * 100) if N_total > 0 else 0.0

        return {
            "N_total": N_total,
            "N_encoding_utf8": N_encoding_utf8,
            "value": round(encoding_utf8_score, 2)
        }
class Preview():
    '''Calculates the Encoding UTF-8 dimension.

    The calculation is performed over all values in the resource data.
    In each row, ever cell is inspected if there is a value present in it.

    The calculation is: `cells_with_value/total_numbr_of_cells * 100`, where:
        * `cells_with_value` is the number of cells containing a value. A cell
            contains a value if the value in the cell is not `None` or an empty
            string or a string containing only whitespace.
        * `total_numbr_of_cells` is the total number of cells expected to be
            populted. This is calculated from the number of rows multiplied by
            the number of columns in the tabular data.

    The return value is a percentage of cells that are populated from the total
    number of cells.
    '''

    def __init__(self):
        self.name = 'preview'
    def calculate_metric(self, resource):
        '''Calculates the API Accesssibility dimension metric for the given resource
        from the resource data.

        :param resource: `dict`, CKAN resource.
        :param data: `dict`, the resource data as a dict with the following
            values:
                * `total`, `int`, total number of rows.
                * `fields`, `list` of `dict`, column metadata - name, type.
                * `records`, `iterable`, iterable over the rows in the resource
                    where each row is a `dict` itself.

        :returns: `dict`, the report contaning the calculated values:
            * `value`, `float`, the percentage of complete values in the data.
            * `total`, `int`, total number of values expected to be populated.
            * `complete`, `int`, number of cells that have value.
        '''
        log.debug ('-----Preview-----')
        preview_score = 0
        log.debug (resource['format'])
        if(resource['datastore_active'] == True):
            preview_score = 1
        elif(resource['datastore_active'] == False): 
            preview_score = 0
       
        return {
            'datastore': resource['datastore_active'],
            'value': preview_score
        }
   
    def calculate_cumulative_metric(self, resources, metrics):
        '''Calculates the cumulative report for all resources from the
        calculated results for each resource.

        The calculation is done as `all_complete/all_total * 100`, where
            * `all_complete` is the total number of completed values in all
                resources.
            * all_total is the number of expected values (rows*columns) in all
                resources.
        The final value is the percentage of completed values in all resources
        in the dataset.

        :param resources: `list` of CKAN resources.
        :param metrics: `list` of `dict` results for each resource.

        :returns: `dict`, a report for the total percentage of complete values:
            * `value`, `float`, the percentage of complete values in the data.
            * `total`, `int`, total number of values expected to be populated.
            * `complete`, `int`, number of cells that have value.
        '''
        # log.debug('-------------Calculate preview metrics -----------')
        N_total = len(resources)
        N_preview = 0
        for item_metric in metrics:
            val = item_metric.get('value') if item_metric else None
            if isinstance(val, (int, float)) and val:
                N_preview += val
            else:
                N_total -= 1
        preview_score = (N_preview / N_total * 100) if N_total > 0 else 0.0
        # for item_metric in metrics:
        #     #check dict is not Empty
        #     if item_metric and isinstance(item_metric.get('value'), (int, float)):
        #         val = item_metric.get('value')
        #         if val:
        #             N_preview += val
                
        #     else:
        #         N_total -= 1

        

        return {
            "N_total": N_total,
            "N_preview": N_preview,
            "value": round(preview_score, 2)
        }
# class MachineReadable():#DimensionMetric
#     '''Calculates the MachineReadable Data Qualtiy dimension.

#     The calculation is performed over all values in the resource data.
#     In each row, ever cell is inspected if there is a value present in it.

#     The calculation is: `cells_with_value/total_numbr_of_cells * 100`, where:
#         * `cells_with_value` is the number of cells containing a value. A cell
#             contains a value if the value in the cell is not `None` or an empty
#             string or a string containing only whitespace.
#         * `total_numbr_of_cells` is the total number of cells expected to be
#             populted. This is calculated from the number of rows multiplied by
#             the number of columns in the tabular data.

#     The return value is a percentage of cells that are populated from the total
#     number of cells.
#     '''

#     def __init__(self):
#         self.name = 'machine_readable'
#     def calculate_metric(self,resource,data): 
#         return {
#             'consistency': 0,
#             'validity': '',
#             'encoding': '',
#             'value': 0
#         }
#     def calculate_metric_machine(self,resource,consistency_val,validity_report): 
#         '''Calculates the openness dimension metric for the given resource
#         from the resource data.

#         :param resource: `dict`, CKAN resource.
#         :param data: `dict`, the resource data as a dict with the following
#             values:
#                 * `total`, `int`, total number of rows.
#                 * `fields`, `list` of `dict`, column metadata - name, type.
#                 * `records`, `iterable`, iterable over the rows in the resource
#                     where each row is a `dict` itself.

#         :returns: `dict`, the report contaning the calculated values:
#             * `value`, `float`, the percentage of complete values in the data.
#             * `total`, `int`, total number of values expected to be populated.
#             * `complete`, `int`, number of cells that have value.
#         '''
#         #--------------Machine Readable-----------------
#         machine_readable_format = ['CSV','XLSX','XLS','JSON','XML'] #'JSON','XML' 
#         document_format = ['PDF','DOC','DOCX','PPTX','PPT','ODT','ODS','ODP'] #-50
#         image_format = ['PNG','JPEG','GIF','TIFF']#-60
#         openness_5_star_format = ['RDF','TTL','N3','GeoJSON','WMS','GML','KML','SHP','Esri REST']#100
#         validity_chk = True
#         encoding_utf8 = False
#         data_format = resource['format'] 
#         mimetype    = resource['mimetype']   
#         machine_readable_score = 100
#         # log.debug('-----Machine Readable ------')
#         # log.debug(data_format)
#         # log.debug(resource['mimetype'])
#         # log.debug(resource['id'])
#         #if data format and mimetype is null, find format from url
#         if ((data_format == '' or data_format == None) and (mimetype == '' or mimetype == None)):
#             resource_url = resource['url']   
#             format_url = resource_url.split(".")[-1]
#             data_format = format_url.upper()
#             # log.debug(data_format)
#         #---- Tabular Data -------------#
#         data_format = data_format.replace(".", "")
#         data_format = data_format.upper()
#         if data_format in machine_readable_format:
#             encoding = validity_report.get('encoding')
#             valid    = validity_report.get('valid')          
#             log.debug('-----machine_readable_format------')
#             # log.debug(valid)
#             # log.debug(encoding_utf8)
#             # log.debug(validity_chk)
#             # log.debug(consistency_val)

#             if "utf-8" in encoding:
#                 encoding_utf8 = True
#             # if(validity_report.get('blank-header') > 0 or validity_report.get('duplicate-header') > 0 or 
#             #    validity_report.get('blank-row') > 0 or validity_report.get('duplicate-row') > 0 or 
#             #    validity_report.get('extra-value') > 0 or validity_report.get('schema-error') > 0):
#             if(validity_report.get('blank-header') > 0 or validity_report.get('duplicate-header') > 0 or 
#                validity_report.get('extra-value') or validity_report.get('source-error') > 0 ):
#                 validity_chk = False
            
#             if(consistency_val >= 0 and consistency_val < 100):
#                 machine_readable_score = machine_readable_score-15
#             if(validity_chk == False):
#                 machine_readable_score = machine_readable_score-20
#             if(encoding_utf8 == False):
#                 machine_readable_score = machine_readable_score-10
                    
#             # log.debug('----machine_readable_val in------')
#             # log.debug('MachineReadable score: %f%%', machine_readable_score)
#             # log.debug(encoding_utf8)
#             # log.debug(consistency_val)
#             # log.debug(validity_chk)
#             return {
#                 'consistency': consistency_val,
#                 'validity': validity_chk,
#                 'encoding': encoding_utf8,
#                 'value': machine_readable_score
#             }
#         elif data_format in openness_5_star_format:
#             machine_readable_score = 100
#         elif data_format in document_format:
#             machine_readable_score = machine_readable_score-50
#         elif data_format in image_format:
#             machine_readable_score = machine_readable_score-60
#         else:
#             machine_readable_score = machine_readable_score-60
            
#         return {
#             'format': data_format,
#             'value': machine_readable_score
#         }

#     def calculate_cumulative_metric(self, resources, metrics):
#         '''Calculates the cumulative report for all resources from the
#         calculated results for each resource.

#         The calculation is done as `all_complete/all_total * 100`, where
#             * `all_complete` is the total number of completed values in all
#                 resources.
#             * all_total is the number of expected values (rows*columns) in all
#                 resources.
#         The final value is the percentage of completed values in all resources
#         in the dataset.

#         :param resources: `list` of CKAN resources.
#         :param metrics: `list` of `dict` results for each resource.

#         :returns: `dict`, a report for the total percentage of complete values:
#             * `value`, `float`, the percentage of complete values in the data.
#             * `total`, `int`, total number of values expected to be populated.
#             * `complete`, `int`, number of cells that have value.
#         '''
#         # log.debug('-------------Calculate machine metrics -----------')
#         # log.debug(metrics)
#         machine_readable_list = []
#         total = 0
        
#         for item_metric in metrics:
#             #check dict is not Empty
#             if item_metric:
#                 machine_readable_score = item_metric.get('value')
#                 if isinstance(machine_readable_score, int):
#                     total = total+machine_readable_score
#                     machine_readable_list.append(machine_readable_score)
#         if machine_readable_list:
#             result_score = max(machine_readable_list)
#             return {
#                 'total': total,
#                 'avg_score': total/len(machine_readable_list),
#                 'value': result_score,
#             }
#         else:
#             return {
#                 'total': 0,
#                 'value': 0,
#             }
class Completeness():#DimensionMetric
    '''Calculates the completeness Data Qualtiy dimension.

    The calculation is performed over all values in the resource data.
    In each row, ever cell is inspected if there is a value present in it.

    The calculation is: `cells_with_value/total_numbr_of_cells * 100`, where:
        * `cells_with_value` is the number of cells containing a value. A cell
            contains a value if the value in the cell is not `None` or an empty
            string or a string containing only whitespace.
        * `total_numbr_of_cells` is the total number of cells expected to be
            populted. This is calculated from the number of rows multiplied by
            the number of columns in the tabular data.

    The return value is a percentage of cells that are populated from the total
    number of cells.
    '''
    def __init__(self):
        # super(Completeness, self).__init__('completeness')
        self.name = 'completeness'
    # def calculate_metric(self, resource, data):
    #     log.debug('---completeness start----')
    #     log.debug("Fields: %s", data.get('fields'))
    #     log.debug("Total records: %s", data.get('total'))

    #     fields = data.get('fields', [])
    #     records = data.get('records', [])
    #     columns_count = len(fields)
    #     rows_count = len(records)
    #     total_values_count = columns_count * rows_count

    #     log.debug('---Rows: %d, Columns: %d, Total Values: %d',
    #               rows_count, columns_count, total_values_count)

    #     total_complete_values = sum(self._completeness_row(row) for row in records)

    #     result = (float(total_complete_values) / float(total_values_count) * 100.0) if total_values_count else 0.0

    #     log.debug('Complete (non-empty) values: %d', total_complete_values)
    #     log.debug('Completeness score: %.2f%%', result)

    #     return {
    #         'value': result,
    #         'total': total_values_count,
    #         'complete': total_complete_values,
    #     }
    def calculate_metric(self, resource, data):
        '''Calculates the completeness dimension metric for the given resource
        from the resource data.

        :param resource: `dict`, CKAN resource.
        :param data: `dict`, the resource data as a dict with the following
            values:
                * `total`, `int`, total number of rows.
                * `fields`, `list` of `dict`, column metadata - name, type.
                * `records`, `iterable`, iterable over the rows in the resource
                    where each row is a `dict` itself.

        :returns: `dict`, the report contaning the calculated values:
            * `value`, `float`, the percentage of complete values in the data.
            * `total`, `int`, total number of values expected to be populated.
            * `complete`, `int`, number of cells that have value.
        '''
        log.debug('---completeness start----')
        log.debug(data['fields'])
        log.debug(data['total'])
        # log.debug('---completeness list records----')
        # log.debug(list(data['records']))
        # columns_count = len(data['fields'])
        # rows_count = data['total']
        # total_values_count = columns_count * rows_count

        # log.debug('---Rows: %d, Columns: %d, Total Values: %d',
        #                   rows_count, columns_count, total_values_count)
        # total_complete_values = 0
        # for row in data['records']:
        #     log.debug('---row---')
        #     log.debug(row)
        #     total_complete_values += self._completenes_row(row)

        # result = \
        #     float(total_complete_values)/float(total_values_count) * 100.0 if \
        #     total_values_count else 0
        # log.debug ('Complete (non-empty) values: %s',
        #                   total_complete_values)
        # log.debug('Completeness score: %f%%', result)
        # return {
        #     'value': round(result,2),
        #     'total': total_values_count,
        #     'complete': total_complete_values,
        # }


        # log.debug('------raw_data------')
        # raw_data = data['raw_data']
        # headers = raw_data[0]
        # rows = raw_data[1:]
        # #  แปลงเป็น list of dict
        # records = [dict(zip(headers, row)) for row in rows]
        # # log.debug(records)
        # # สร้าง DataFrame จาก records
        # df = pd.DataFrame(records)
        #--------------------------------------------
        rows_count = data['total']
        # df = pd.DataFrame(data['records'])
        # แปลง raw_data เป็น DataFrame โดยข้ามแถวแรก (header)
        df = pd.DataFrame(data['raw_data'][1:], columns=data['raw_data'][0])
        log.debug(df)
        # เช็คว่ามีข้อมูลดิบหรือไม่ และ df แปลงเป็น DataFrame ได้หรือไม่
        if rows_count > 1 and df.empty:
            log.debug("--df.empty--")
            records = data['raw_data']  # ดึงข้อมูลทั้งหมดออกจาก LazyStreamingList
            log.debug(records)
            clean_records = []
            for r in records:
                if None in r:
                    r.pop(None)
                clean_records.append(r)

            df = pd.DataFrame(clean_records) 
            if rows_count > 1 and df.empty: 
                return {
                    'error': 'Cannot parse table to DataFrame', 
                    'value': None
                } 

        # เช็คว่าข้อมูลมีหรือไม่ (กรณีไม่มีข้อมูลดิบเลย หรือ df ว่างจริง)
        elif df.empty or df.shape[0] == 0 or df.shape[1] == 0:
            return {
                'error': 'Data is empty', 
                'value': None
            }
        else:
            # ลบช่องว่างและจัดการค่าว่างใน string
            # df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
            # # แปลงค่าที่ถือว่าเป็น "ข้อมูลว่าง" ให้กลายเป็น NaN (missing value) ของ Pandas
            # df.replace(to_replace=["", " ", "-", "ไม่มีข้อมูล", "null", "NaN","N/A","n/a","NA","na"], value=np.nan, inplace=True)
            missing_values = [
                "", " ", "-", "ไม่มีข้อมูล", "null", "n/a", "na", "none", "nan"
            ]

            # ฟังก์ชัน normalize เพื่อจัดการ missing value
            def normalize_empty_values(x):
                if isinstance(x, str):
                    x_str = x.strip().lower()
                    return np.nan if x_str in missing_values else x
                return x
            # ลบช่องว่างและแปลงค่าที่ไม่มีข้อมูลให้เป็น NaN
            df = df.applymap(normalize_empty_values)
            # คำนวณจำนวนช่องทั้งหมดในตาราง
            rows_count, columns_count = df.shape 
            total_values_count = rows_count * columns_count
            # คำนวณจำนวนค่าที่ ไม่เป็น NaN
            total_complete_values = df.notna().sum().sum()

            log.debug('---Rows: %d, Columns: %d, Total Values: %d',
                    rows_count, columns_count, total_values_count)
            log.debug('Complete (non-empty) values: %d', total_complete_values)

            result = float(total_complete_values) / float(total_values_count) * 100.0 if total_values_count else 0
            log.debug('Completeness score: %f%%', result)

            return {
                'value': round(result, 2),
                'total': total_values_count,
                'complete': int(total_complete_values),
            }
    # def _completenes_row(self, row):
    #         count = 0
    #         for _, value in row.items():
    #             if value is None:
    #                 continue
    #             if isinstance(value, str):
    #                 if not value.strip():
    #                     continue
    #             count += 1
    #         return count
    # def _completeness_row(self, row):
    #     return sum(1 for v in row.values() if v is not None and (not isinstance(v, str) or v.strip()))

    def calculate_cumulative_metric(self, resources, metrics):
        # total, complete = reduce(
        #     lambda acc, result: (
        #         acc[0] + result.get('total', 0),
        #         acc[1] + result.get('complete', 0)
        #     ),
        #     metrics,
        #     (0, 0)
        # )
        # return {
        #     'total': total,
        #     'complete': complete,
        #     'value': round(float(complete) / float(total) * 100.0,2) if total else 0.0,
        # }
        N_total = len(resources)
        N_complete = 0 
        for item_metric in metrics:
            #check dict is not Empty
            # if item_metric and isinstance(item_metric.get('value'), (int, float)):
            #     val = item_metric.get('value')
            val = item_metric.get('value') if item_metric else None
            if isinstance(val, (int, float)):
                if val:
                    N_complete += val
                
            else:
                N_total -= 1

        completeness_score = (N_complete / N_total ) if N_total > 0 else 0.0
        return {
            "N_total": N_total,
            "N_complete": N_complete,
            "value": round(completeness_score, 2)
        }
    # def calculate_metric(self, resource, data):
    #     '''Calculates the completeness dimension metric for the given resource
    #     from the resource data.

    #     :param resource: `dict`, CKAN resource.
    #     :param data: `dict`, the resource data as a dict with the following
    #         values:
    #             * `total`, `int`, total number of rows.
    #             * `fields`, `list` of `dict`, column metadata - name, type.
    #             * `records`, `iterable`, iterable over the rows in the resource
    #                 where each row is a `dict` itself.

    #     :returns: `dict`, the report contaning the calculated values:
    #         * `value`, `float`, the percentage of complete values in the data.
    #         * `total`, `int`, total number of values expected to be populated.
    #         * `complete`, `int`, number of cells that have value.
    #     '''
    #     log.debug('---completeness start----')
    #     log.debug(data['fields'])
    #     log.debug(data['total'])
    #     columns_count = len(data['fields'])
    #     rows_count = data['total']
    #     total_values_count = columns_count * rows_count

    #     log.debug('---Rows: %d, Columns: %d, Total Values: %d',
    #                       rows_count, columns_count, total_values_count)
    #     log.debug(data)
    #     total_complete_values = 0
    #     for row in data['records']:
    #         total_complete_values += self._completenes_row(row)

    #     result = \
    #         float(total_complete_values)/float(total_values_count) * 100.0 if \
    #         total_values_count else 0
    #     log.debug ('Complete (non-empty) values: %s',
    #                       total_complete_values)
    #     log.debug('Completeness score: %f%%', result)
    #     return {
    #         'value': result,
    #         'total': total_values_count,
    #         'complete': total_complete_values,
    #     }

    # def _completenes_row(self, row):
    #     count = 0
    #     for _, value in row.items():
    #         if value is None:
    #             continue
    #         if isinstance(value, str):
    #             if not value.strip():
    #                 continue
    #         count += 1
    #     return count

    # def calculate_cumulative_metric(self, resources, metrics):
    #     '''Calculates the cumulative report for all resources from the
    #     calculated results for each resource.

    #     The calculation is done as `all_complete/all_total * 100`, where
    #         * `all_complete` is the total number of completed values in all
    #             resources.
    #         * all_total is the number of expected values (rows*columns) in all
    #             resources.
    #     The final value is the percentage of completed values in all resources
    #     in the dataset.

    #     :param resources: `list` of CKAN resources.
    #     :param metrics: `list` of `dict` results for each resource.

    #     :returns: `dict`, a report for the total percentage of complete values:
    #         * `value`, `float`, the percentage of complete values in the data.
    #         * `total`, `int`, total number of values expected to be populated.
    #         * `complete`, `int`, number of cells that have value.
    #     '''
    #     total, complete = reduce(lambda total, complete, result: (
    #         total + result.get('total', 0),
    #         complete + result.get('complete', 0)
    #     ), metrics, (0, 0))
    #     return {
    #         'total': total,
    #         'complete': complete,
    #         'value': float(complete)/float(total) * 100.0 if total else 0.0,
    #     }
    #     # return {
    #     #     'total': 0,
    #     #     'complete': 0,
    #     #     'value': 0,
    #     # }
class Uniqueness(): #DimensionMetric
    '''Calculates the uniqueness of the data.

    The general calculation is: `unique_values/total_values * 100`, where:
        * `unique_values` is the number of unique values
        * `total_values` is the total number of value

    The dimension value is a percentage of unique values in the data.
    '''
    def __init__(self):
        # super(Uniqueness, self).__init__('uniqueness')
        self.name = 'uniqueness'
    def nan_to_none(self, value):
        if isinstance(value, float) and math.isnan(value):
            return None
        return value
    def _rename_columns(self, columns):
        """ตั้งชื่อใหม่ให้คอลัมน์ซ้ำหรือชื่อ NaN"""
        seen = {}
        null_count = 0
        new_cols = []
        for col in columns:
            # ถ้าชื่อคอลัมน์เป็น NaN หรือ None → แปลงเป็น Null_1, Null_2, ...
            if pd.isna(col) or str(col).strip() == "":
                null_count += 1
                new_col = f"Null_{null_count}"
            else:
                new_col = str(col)

            # ถ้ามีชื่อซ้ำ → เพิ่ม _1, _2, ...
            if new_col in seen:
                seen[new_col] += 1
                new_col = f"{new_col}_{seen[new_col]}"
            else:
                seen[new_col] = 0

            new_cols.append(new_col)
        return new_cols
    def calculate_metric(self, resource, data):
        '''Calculates the uniqueness of the values in the data for the given
        resource.

        For each column of the data, the number of unique values is calculated
        and the total number of values (basically the number of rows).

        Then, to calculate the number of unique values in the data, the sum of
        all unique values is calculated and the sum of the total number of
        values for each column. The the percentage is calculated from those two
        values.

        :param resource: `dict`, CKAN resource.
        :param data: `dict`, the resource data as a dict with the following
            values:
                * `total`, `int`, total number of rows.
                * `fields`, `list` of `dict`, column metadata - name, type.
                * `records`, `iterable`, iterable over the rows in the resource
                    where each row is a `dict` itself.

        :returns: `dict`, a report on the uniqueness metrics for the given
            resource data:
            * `value`, `float`, the percentage of unique values in the data.
            * `total`, `int`, total number of values in the data.
            * `unique`, `int`, number unique values in the data.
            * `columns`, `dict`, detailed report for each column in the data.
        '''
        try:
            # ✅ แปลง raw_data → DataFrame โดยข้าม header
            df = pd.DataFrame(data['raw_data'][1:], columns=data['raw_data'][0])

            # ✅ ตั้งชื่อคอลัมน์ใหม่ (รวม NaN และชื่อซ้ำ)
            df.columns = self._rename_columns(df.columns)

            # ✅ ไม่ลบคอลัมน์ชื่อ NaN แล้ว
            df = df.dropna(axis=1, how='all')  # ลบเฉพาะคอลัมน์ที่ว่างทั้งหมด (ไม่มีข้อมูลเลย)

            if df.empty or df.shape[0] == 0 or df.shape[1] == 0:
                return {'error': 'Data is empty', 'value': None}

            total_rows = len(df)

            # ✅ หาแถวซ้ำ
            duplicates_bool = df.duplicated(keep=False)
            duplicate_rows = df[duplicates_bool].copy()

            if not duplicate_rows.empty:
                grouped = (
                    duplicate_rows
                    .reset_index()
                    .groupby(list(df.columns), dropna=False, as_index=False)
                    .agg({'index': list})
                )

                duplicate_details = []
                for _, row in grouped.iterrows():
                    row_data = {col: self.nan_to_none(row[col]) for col in df.columns}
                    human_rows = [i + 1 for i in row["index"]]
                    duplicate_details.append({
                        "row": row_data,
                        "rows": human_rows
                    })
            else:
                duplicate_details = []

            # ✅ คำนวณ uniqueness
            unique_rows = len(df.drop_duplicates())
            uniqueness_score = (unique_rows / total_rows) * 100 if total_rows > 0 else 0

            return {
                'value': round(uniqueness_score, 2),
                'total': total_rows,
                'unique': unique_rows,
                'duplicates': duplicate_details
            }

        except Exception as e:
            import traceback
            log.error(f"[Uniqueness] Failed: {e}\n{traceback.format_exc()}")
            return {'error': str(e), 'value': None}
        #-----old1 ------
        # # df = pd.DataFrame(data['records'])
        # # แปลง raw_data เป็น DataFrame โดยข้ามแถวแรก (header)
        # df = pd.DataFrame(data['raw_data'][1:], columns=data['raw_data'][0])
        # if df.empty or df.shape[0] == 0 or df.shape[1] == 0:
        #     return {
        #         'error': 'Data is empty',
        #         'value': None
        #     }
        # else:
        #     total_rows = len(df)

        #     # หาแถวที่ซ้ำทั้งหมด
        #     duplicates_bool = df.duplicated(keep=False)
        #     duplicate_rows = df[duplicates_bool].copy()

        #     # group duplicates
        #     grouped = (
        #         duplicate_rows
        #         .reset_index()
        #         .groupby(list(df.columns), dropna=False)['index']
        #         .apply(list)
        #         .reset_index()
        #     )

        #     # แปลงให้อ่านง่าย
        #     duplicate_details = []
        #     for _, row in grouped.iterrows():
        #         # row_data = {col: row[col] for col in df.columns}
        #         row_data = {col: self.nan_to_none(row[col]) for col in df.columns}
        #         # แปลง index ให้เป็นลำดับ row (เริ่มที่ 1 แทน 0)
        #         human_rows = [i + 1 for i in row["index"]]
        #         duplicate_details.append({
        #             "row": row_data,          # ค่า row ที่ซ้ำ
        #             "rows": human_rows        # บรรทัดที่ซ้ำ (อ่านง่าย)
        #         })

        #     # นับจำนวน unique rows
        #     unique_rows = len(df.drop_duplicates())

        #     # คำนวณ score
        #     uniqueness_score = (unique_rows / total_rows) * 100 if total_rows > 0 else 0

        #     return {
        #         'value': round(uniqueness_score, 2),
        #         'total': total_rows,
        #         'unique': unique_rows,
        #         'duplicates': duplicate_details
        #     }
        #------old old --------------------------------------
        # # แปลง data['records'] เป็น DataFrame
        # df = pd.DataFrame(data['records'])
        # if df.empty or df.shape[0] == 0 or df.shape[1] == 0:
        #     return {
        #         'error': 'Data is empty', 
        #         'value': None
        #     }
        # else:
            
        #     total_rows = len(df)
        #     # duplicates = df[df.duplicated(keep=False)]  # แสดงแถวที่ซ้ำทั้งหมด ไม่ว่าจะแถวแรกหรือถัดไป
        #     # หาแถวที่ซ้ำทั้งหมด (keep=False คือ mark ทั้งแถวที่ซ้ำ)
        #     duplicates_bool = df.duplicated(keep=False)
        #     # ดึง index ของแถวที่ซ้ำ
        #     duplicate_indices = df.index[duplicates_bool].tolist()

        #     unique_rows = len(df.drop_duplicates())

        #     uniqueness_score = (unique_rows / total_rows) * 100 if total_rows > 0 else 0
        #     log.debug("----Uniqueness start----")
        #     log.debug("Total rows: %d", total_rows)
        #     log.debug("Unique rows: %d", unique_rows)
        #     log.debug("Uniqueness score: %.2f%%", uniqueness_score)
        #     # log.debug(df)
        #     # log.debug(duplicate_indices)
        #     return {
        #         'value': round(uniqueness_score, 2),
        #         'total': total_rows,
        #         'unique': unique_rows,
        #         'duplicates': duplicate_indices
        #     }
        #---------------------------------
    def calculate_cumulative_metric(self, resources, metrics):
        '''Calculates uniqueness for all resources based on the metrics
        calculated in the previous phase for each resource.

        The calculation is performed based on the total unique values as a
        percentage of the total number of values present in all data from all
        given resources.

        :param resources: `list` of CKAN resources.
        :param metrics: `list` of `dict` results for each resource.

        :returns: `dict`, a report for the total percentage of unique values:
            * `value`, `float`, the percentage of unique values in the data.
            * `total`, `int`, total number of values in the data.
            * `unique`, `int`, number of unique values.
        '''
        # result = {}
        # result['total'] = sum([r.get('total', 0) for r in metrics])
        # result['unique'] = sum([r.get('unique', 0) for r in metrics])
        # if result['total'] > 0:
        #     result['value'] = round(100.0 *
        #                        float(result['unique'])/float(result['total']),2)
        # else:
        #     result['value'] = 0.0

        # return result
        N_total = len(resources)
        N_unique = 0 
        for item_metric in metrics:
            #check dict is not Empty
            val = item_metric.get('value') if item_metric else None
            if isinstance(val, (int, float)):
                if val:
                    N_unique += val
                
            else:
                N_total -= 1

        uniqueness_score = (N_unique / N_total ) if N_total > 0 else 0.0
        return {
            "N_total": N_total,
            "N_unique": N_unique,
            "value": round(uniqueness_score, 2)
        }
class Validity():#DimensionMetric
    '''Calculates Data Quality dimension validity.

    The validity is calculated as a percentage of valid records against the
    total number of records in the data.

    The validation is performed using the validation provided by
    https://github.com/frictionlessdata/goodtables-py.
    '''
    def __init__(self):
        # super(Validity, self).__init__('validity')
        self.name = 'validity'

    def _perform_validation(self, resource, total_rows,data):
        rc = {}
        rc.update(resource)
        rc['validation_options'] = {
            'row_limit': total_rows,
        }
        # return validate_resource_data(rc)
        try:
            return validate_resource_data(rc,data)
        except ValueError as ve:
            log.error(f"Validation failed for resource {resource.get('id')}: {ve}")
            # บางกรณี อาจเจอ unmatched '{' ให้ตรวจว่าข้อความมี curly braces
            if "unmatched '{'" in str(ve):
                log.warning("Found unmatched '{' in validation message. Skipping validation.")
            return None
        except Exception as e:
            log.exception(f"Unexpected error during validation of resource {resource.get('id')}")
            return None

    def calculate_metric(self, resource, data):
        '''Calculates the percentage of valid records in the resource data.

        The validation of the resource data is done based on the resource
        schema. The schema is specified in the resource metadata itself.

        :param resource: `dict`, CKAN resource.
        :param data: `dict`, the resource data as a dict with the following
            values:
                * `total`, `int`, total number of rows.
                * `fields`, `list` of `dict`, column metadata - name, type.
                * `records`, `iterable`, iterable over the rows in the resource
                    where each row is a `dict` itself.

        :returns: `dict`, a report on the validity metrics for the given
            resource data:
            * `value`, `float`, percentage of valid records in the resource
                data.
            * `total`, `int`, total number of records.
            * `valid`, `int`, number of valid records.
        '''
        total_rows = 0
        total_errors = 0
        encoding = ''
        valid = ''
        error_message = ''
        error_message_all = ''
        headers = []
        table_count = 0
        relevant_errors = 0
        filepath = resource['url']
        dict_error = {}
        log.debug("---validity---")
        #ตรวจสอบการอ่านข้อมูล ถ้า Failed to Fetch Data จะไม่ตรวจ Validity
        if data.get('error'):  # ตรวจว่า 'error' มีค่าและไม่ใช่ค่าว่าง
            dict_error['encoding'] = None
            dict_error['error'] = 'Data is empty/Invalid file format or structure'
            return {
                'error': 'Data is empty/Invalid file format or structure',
                'value': None,
                'report': dict_error
            }
        elif (data.get('total') == 0):
            dict_error['encoding'] = None
            dict_error['error'] = 'Data is empty'
            return {
                'error': 'Data is empty',
                'value': None,
                'report': dict_error
            }
        validation = None
        try:
            validation = self._perform_validation(resource,
                                                  data.get('total', 0),data)
        except Exception as e:
            log.error('Failed to validate data for resource: %s. '
                              'Error: %s', resource['id'], str(e))
            log.exception(e)
            return {
                'failed': True,
                'error': str(e),
            }      
        if validation:
            # log.debug('---validation get report---') remove , 'duplicate-row': 0
            dict_error = {'blank-header': 0, 'duplicate-header': 0, 'blank-row': 0 ,'extra-value':0,'extra-header':0,'missing-value':0,'format-error':0, 'schema-error':0, 'encoding-error':0, 'source-error':0,'encoding':'', 'error':''}
            error_types=['blank-header', 'duplicate-header', 'extra-value','source-error']
            for table in validation.get('tables', []):
                total_rows += table.get('row-count', 0)
                total_errors += table.get('error-count', 0)
                for error in table.get('errors', []):
                    item_code = error.get('code')
                    if item_code != 'duplicate-row':   
                        error_message += error.get('message', '') + ','   
                        count_val = dict_error.get(item_code, 0) + 1      
                        dict_error[item_code] = count_val

                        if item_code in error_types:   
                            relevant_errors += 1
                    else:                              
                        continue
                encoding = table.get('encoding')
                valid = table.get('valid')
                log.debug("---encoding---")
                log.debug(encoding)
                # log.debug(table.get('errors'))
                dict_error['encoding'] = encoding
                dict_error['valid']    = valid
                dict_error['error'] = error_message
            #โหลดและตรวจสอบไฟล์ได้แล้ว แต่ไม่มีข้อมูลในไฟล์นั้นเลย เช่น ไม่มีข้อมูลในไฟล์เลย หรือมี header แต่ไม่มี row จริง
            if total_rows == 0:
                log.debug('--total_rows=0--')
                default_error_message = 'Data is empty/Invalid file format or structure'
                if dict_error.get('source-error', 0) > 0 :
                    default_error_message = "Source-error"
                   
                return {
                    'value': None,
                    'error': default_error_message,
                    'report': dict_error
                }
            blank_header = 1
            duplicate_header = 1
            extra_value =  1
            if dict_error.get('blank-header', 0) > 0 :
                blank_header = 0
            if dict_error.get('duplicate-header', 0) > 0 :
                duplicate_header = 0
            if dict_error.get('extra-value', 0) > 0 :
                extra_value = 0

            validity_score = (blank_header+duplicate_header+extra_value) / 3 * 100
            # log.debug("---validity_score---")
            # log.debug(blank_header)
            # log.debug(duplicate_header)
            # log.debug(extra_value)
            # log.debug(validity_score)
            is_valid = 1 if validity_score == 100 else 0

            return {
                'value': round(validity_score,2),
                'total': 1,        # นับเป็น 1 dataset
                'valid': is_valid, # 1 = valid, 0 = invalid
                'report': dict_error
            }
        else:
            #ไม่สามารถตรวจสอบข้อมูลได้เลย เช่น โหลดไฟล์ไม่สำเร็จ, URL เสีย / response ไม่ใช่ไฟล์จริง
            dict_error['encoding'] = None
            dict_error['error'] = 'Failed to validate the resource'
            return {
                'error': 'Failed to validate the resource',
                'value': None,
                'report': dict_error
            }

    def calculate_cumulative_metric(self, resources, metrics):
        '''Calculates the percentage of valid records in all the data for the
        given resources based on the calculation for each individual resource.

        The percentage is calculated based on the total number of valid records
        across all of the data in all resources against the total number of
        records in the data.

        :param resources: `list` of CKAN resources.
        :param metrics: `list` of `dict` results for each resource.

        :returns: `dict`, a report on the validity metrics for all of the data.
            * `value`, `float`, percentage of valid records in the data.
            * `total`, `int`, total number of records.
            * `valid`, `int`, number of valid records.
        '''
        # total = sum([r.get('total', 0) for r in metrics])
        # valid = sum([r.get('valid', 0) for r in metrics])
        # if total == 0:
        #     return {
        #         'value': 0.0,
        #         'total': 0,
        #         'valid': 0,
        #     }
        # return {
        #         'value': round(float(valid)/float(total) * 100.0,2) if total else 0.0,
        #         'total': total,
        #         'valid': valid,
        #     }
        N_total = len(resources)
        N_validity = 0 
        for item_metric in metrics:
            #check dict is not Empty
            
            # if item_metric and isinstance(item_metric.get('value'), (int, float)):
            #     val = item_metric.get('value')
            val = item_metric.get('value') if item_metric else None
            if isinstance(val, (int, float)):
                if val:
                    N_validity += val
                
            else:
                N_total -= 1

        validity_score = (N_validity / N_total ) if N_total > 0 else 0.0
        return {
            "N_total": N_total,
            "N_validity": N_validity,
            "value": round(validity_score, 2)
        }
class Relevance(): #DimensionMetric
    '''Calculates the Relevance of the data.

    The general calculation is: `unique_values/total_values * 100`, where:
        * `unique_values` is the number of unique values
        * `total_values` is the total number of value

    The dimension value is a percentage of unique values in the data.
    '''
    def __init__(self):
        # super(Uniqueness, self).__init__('uniqueness')
        self.name = 'relevance'
    # def get_organization_id(self,base_url, org_name):
    #     url = f"{base_url}/api/3/action/organization_show?id={org_name}"
    #     try:
    #         response = requests.get(url, timeout=10)
    #         response.raise_for_status()
    #         data = response.json()
    #         if data.get('success'):
    #             return data['result']['id'], None
    #         return None, f"API responded success=False: {data}"
    #     except Exception as e:
    #         return None, f"Error fetching organization ID: {str(e)}"

    # def get_first_harvest_source(self,base_url, org_id):
    #     url = f"{base_url}/api/3/action/harvest_source_list?organization_id={org_id}"
    #     try:
    #         response = requests.get(url, timeout=10)
    #         response.raise_for_status()
    #         data = response.json()
    #         if data.get('success') and data['result']:
    #             return data['result'][0], None  # ดึงตัวแรก
    #         return None, "No harvest sources found"
    #     except Exception as e:
    #         return None, f"Error fetching harvest source: {str(e)}"

    # def get_package_from_url(self,harvest_url, package_id):
    #     url = f"{harvest_url}/api/3/action/package_show?id={package_id}&include_tracking=true"
    #     try:
    #         response = requests.get(url, timeout=10)
    #         response.raise_for_status()
    #         data = response.json()
    #         if data.get('success'):
    #             return data['result'], None
    #         return None, f"API returned success=False: {data}"
    #     except Exception as e:
    #         return None, f"Error fetching package data: {str(e)}"
    def analyze_package_statistics(self, package_data):
        resources = package_data.get("resources", [])
        total_download = 0

        for res in resources:
            tracking = res.get("tracking_summary", {})
            downloads = tracking.get("total", 0)
            total_download += downloads

        total_view = package_data.get("tracking_summary", {}).get("total", 0)

        if total_view == 0:
            relevance = 0  
        else:
            relevance = min((total_download / total_view) * 100, 100)
        return {
            "total_download": total_download,
            "total_view": total_view,
            "relevance_percent": round(relevance, 2)
        }
    # def get_owner_org_by_package_id(self,package_id):
    #     package = Session.query(Package).get(package_id)
    #     if package:
    #         return package.owner_org
    #     return None
    def calculate_metric(self, resource, level_name, execute_type):
        '''Calculates the relevance of the values in the data for the given
        resource.

        For each column of the data, the number of unique values is calculated
        and the total number of values (basically the number of rows).

        Then, to calculate the number of unique values in the data, the sum of
        all unique values is calculated and the sum of the total number of
        values for each column. The the percentage is calculated from those two
        values.

        :param resource: `dict`, CKAN resource.
        :param data: `dict`, the resource data as a dict with the following
            values:
                * `total`, `int`, total number of rows.
                * `fields`, `list` of `dict`, column metadata - name, type.
                * `records`, `iterable`, iterable over the rows in the resource
                    where each row is a `dict` itself.

        :returns: `dict`, a report on the uniqueness metrics for the given
            resource data:
            * `value`, `float`, the percentage of unique values in the data.
            * `total`, `int`, total number of values in the data.
            * `unique`, `int`, number unique values in the data.
            * `columns`, `dict`, detailed report for each column in the data.
        '''

        return {
            'value': None
        }
    def calculate_cumulative_metric(self, package_id, resources, metrics):
        '''Calculates uniqueness for all resources based on the metrics
        calculated in the previous phase for each resource.

        The calculation is performed based on the total unique values as a
        percentage of the total number of values present in all data from all
        given resources.

        :param resources: `list` of CKAN resources.
        :param metrics: `list` of `dict` results for each resource.

        :returns: `dict`, a report for the total percentage of unique values:
            * `value`, `float`, the percentage of unique values in the data.
            * `total`, `int`, total number of values in the data.
            * `unique`, `int`, number of unique values.
        '''
        # # collect download in resource level
        # # org_name = 'mrta'
        # base_url = config.get("ckan.site_url")
        # org_name = self.get_owner_org_by_package_id(package_id)
        # # base_url = "https://ckan-dev.opend.cloud"

        # # Step 1: Get org ID
        # org_id, err1 = self.get_organization_id(base_url, org_name)
        # if err1:
        #     return {"error": f"Step 1 failed: {err1}"}

        # # Step 2: Get first harvest source
        # harvest, err2 = self.get_first_harvest_source(base_url, org_id)
        # if err2:
        #     return {"error": f"Step 2 failed: {err2}", "org_id": org_id}

        # harvest_url = harvest.get("url")

        # if not harvest_url or not package_id:
        #     return {"error": "Missing url or resource_id in harvest source", "org_id": org_id, "harvest": harvest}

        try:
            package_data = logic.get_action('package_show')(None, {'id': package_id, 'include_tracking': True})
        except toolkit.ObjectNotFound:
            # print(f"Dataset not found: {package_id}")
            return {"error": f"Dataset not found: {package_id}"}
            package_data = None
        except toolkit.NotAuthorized:
            return {"error": f"Unauthorized to access dataset: {package_id}"}
            package_data = None
        except Exception as e:
            return {"error": f"Unexpected error for package: {package_id}"}
            package_data = None

        stats = self.analyze_package_statistics(package_data)
        
        return {
            "error": None,
            # "org_id": org_id,
            # "harvest_url": harvest_url,
            "package_id": package_id,
            "N_download": stats['total_download'],
            "N_view": stats['total_view'],
            "value": stats['relevance_percent']
        }
# class Accuracy():#DimensionMetric
#     '''Calculates Data Qualtiy dimension accuracy.

#     Accuracy of a record is determined through background research and it is
#     not quite possible to determine if a record is accurate or inaccurate via
#     a generic algorithm.

#     This metric calculates the percentage of accurate records *only* for
#     records that already have been marked as accurate or inaccurate in a
#     particular dataset.

#     To calculate this, the record must contain a flag, a column in the dataset,
#     that marks the record as accurate, inaccurate or not determined.

#     If such column is present, then the calculation is:
#     `number_of_accurate/(number_of_accurate + number_of_inaccurate) * 100)`.

#     Accuracy is measured as percentage of accurate records, from the set of
#     records that have been checked and marked as accurate or inaccurate.
#     '''
#     def __init__(self):
#         # super(Accuracy, self).__init__('accuracy')
#         self.name = 'accuracy'

#     def calculate_metric(self, resource, data):
#         '''Calculates the percentage of accurate records for the given resource
#         data.

#         The resource must contain a data quality setting to configure which
#         column contains the flag that marks the record as accurate.

#         :param resource: `dict`, CKAN resource. The resource dict must contain
#             a property `data_quality_settings` which contain the name of the
#             column used to determine the accuracy of that record. This
#             property must have the following format:

#                 .. code-block: python

#                     resource = {
#                         ...
#                         'data_quality_settings': {
#                             'accuracy': {
#                                 'column': '<column name>',
#                             }
#                         }
#                     }
#         :param data: `dict`, the resource data as a dict with the following
#             values:
#                 * `total`, `int`, total number of rows.
#                 * `fields`, `list` of `dict`, column metadata - name, type.
#                 * `records`, `iterable`, iterable over the rows in the resource
#                     where each row is a `dict` itself.

#         :returns: `dict`, a report on the accuracy metrics for the resource
#             data:
#             * `value`, `float`, percentage of accurate records in the data.
#             * `accurate`, `int`, number of accurate records.
#             * `inaccurate`, `int`, number of inaccurate records.
#             * `total`, `int`, `accurate` + `inaccurate` - total number of
#                 checked records.
#         '''
        # settings = resource.get('data_quality_settings',
        #                         {}).get('accuracy', {})
        # column = settings.get('column')
        # if not column:
        #     log.error('Cannot calculate accuracy on this resource '
        #                       'because no accuracy column is specified.')
        #     return {
        #         'failed': True,
        #         'error': 'Missing accuracy column.',
        #     }

        # accurate = 0
        # inaccurate = 0

        # for row in data['records']:
        #     flag = row.get(column)
        #     if flag is None or flag.strip() == '':
        #         # neither accurate or inaccurate
        #         continue
        #     if flag.lower() in ['1', 'yes', 'accurate', 't', 'true']:
        #         accurate += 1
        #     else:
        #         inaccurate += 1

        # total = accurate + inaccurate
        # value = 0.0
        # if total:
        #     value = float(accurate)/float(total) * 100.0

        # log.debug('Accurate: %d', accurate)
        # log.debug('Inaccurate: %d', inaccurate)
        # log.debug('Accuracy: %f%%', value)
        # return {
        #     'value': value,
        #     'total': total,
        #     'accurate': accurate,
        #     'inaccurate': inaccurate,
        # }

    # def calculate_cumulative_metric(self, resources, metrics):
    #     '''Calculates the percentage of accurate records in all data for the
    #     given resources, based on the calculations for individual resources.

    #     The value is calculated as percentage of accurate records of the total
    #     number of checked records (all accurate + all inaccurate records).

    #     :param resources: `list` of CKAN resources.
    #     :param metrics: `list` of `dict` results for each resource.

    #     :returns: `dict`, a report on the accuracy metrics for all data in all
    #         resources:
    #         * `value`, `float`, percentage of accurate records in the data.
    #         * `accurate`, `int`, number of accurate records.
    #         * `inaccurate`, `int`, number of inaccurate records.
    #         * `total`, `int`, `accurate` + `inaccurate` - total number of
    #             checked records.
    #     '''
        # accurate = sum([r.get('accurate', 0) for r in metrics])
        # inaccurate = sum([r.get('inaccurate', 0) for r in metrics])

        # total = accurate + inaccurate
        # value = 0.0
        # if total:
        #     value = float(accurate)/float(total) * 100.0

        # return {
        #     'value': value,
        #     'total': total,
        #     'accurate': accurate,
        #     'inaccurate': inaccurate,
        # }
class Consistency():#DimensionMetric
    '''Calculates Data Quality dimension consistency.

    This dimension gives a metric of how consistent are the values of same type
    accross the dataset (or resource).

    The type of the values is determined by the type of the column in CKAN's
    data store (like int, numeric, timestamp, text etc).
    For each type, a validator for the consistency is used to try to categorize
    the value in some category - for example all dates with same formats would
    be in the same category, all numbers that use the same format, how many
    of the numbers in a column are floating point and how many are integer.

    For each column, we may get multiple categories (plus a special category
    called `unknown` which means we could not determine the category for those
    values), and we count the number of values for each category.
    The consistency, for a given column, would be the number of values in the
    category with most values, expressed as a percentage of the total number of
    values in that column.

    To calculate the consistency for the whole data, we calculate the total
    number of consistent values (sum of all columns) expressed as a percentage
    of the total number of values present in the data.
    '''
    def __init__(self):
        # super(Consistency, self).__init__('consistency')
        self.name = 'consistency'

    # def validate_date(self, field, value, _type, report):
    #     date_format = detect_date_format(value) or 'unknown'
    #     formats = report['formats']
    #     formats[date_format] = formats.get(date_format, 0) + 1

    # def validate_numeric(self, field, value, _type, report):
    #     if value is None or (isinstance(value, str) and value.strip() == ""):
    #         return
    #     formats = report['formats']
    #     num_format = detect_numeric_format(value)

    #     if num_format:
    #         formats[num_format] = formats.get(num_format, 0) + 1
    #     else:
    #         formats['non-numeric'] = formats.get('non-numeric', 0) + 1
        
    # def validate_string(self, field, value, _type, report):
    #     if value is None or (isinstance(value, str) and value.strip() == ""):
    #         return
    #     formats = report['formats']
    #     formats['text'] = formats.get('text', 0) + 1  # ใช้ 'text' แทนทุกกรณี
    def validate_date(self, field, value, _type, report):
        if value is None or (isinstance(value, str) and value.strip() == ""):
            return
        formats = report['formats']
        date_format = detect_date_format(value)
        if date_format:
            formats['timestamp'] = formats.get('timestamp', 0) + 1
        else:
            formats['other'] = formats.get('other', 0) + 1

    def validate_numeric(self, field, value, _type, report):
        if value is None or (isinstance(value, str) and value.strip() == ""):
            return
        formats = report['formats']
        num_format = detect_numeric_format(value)
        if num_format:
            formats['numeric'] = formats.get('numeric', 0) + 1
        else:
            formats['other'] = formats.get('other', 0) + 1

    def validate_string(self, field, value, _type, report):
        if value is None or (isinstance(value, str) and value.strip() == ""):
            return
        formats = report['formats']
        # text ทุกค่า ถือเป็น text
        formats['text'] = formats.get('text', 0) + 1
    def validate_default(self, field, value, _type, report):
        # ข้ามค่าว่าง ไม่ต้องนับ
        if value is None or (isinstance(value, str) and value.strip() == ""):
            return
        formats = report['formats']
        formats['unknown'] = formats.get('unknown', 0) + 1
    def get_consistency_validators(self):
        return {
            'timestamp': self.validate_date,
            'numeric': self.validate_numeric,
            'int': self.validate_numeric,
            'float': self.validate_numeric,
            'string': self.validate_string,
            'text': self.validate_string,
            'default': self.validate_default  # ใช้กรณีไม่มี type
        }

    def calculate_metric(self, resource , data):
        '''Calculates the consistency of the resource data.

        For each of the columns of the resource data, we determine the number
        of values that belong in the same category (have the same format, are
        of exactly the same type, etc). Then, we find the category with most
        values, per column, and use that expressed as percentage to calculate
        the consistency for that column.
        The final value is calculated from the results obtained for each
        column.

        :param resource: `dict`, CKAN resource.
        :param data: `dict`, the resource data as a dict with the following
            values:
                * `total`, `int`, total number of rows.
                * `fields`, `list` of `dict`, column metadata - name, type.
                * `records`, `iterable`, iterable over the rows in the resource
                    where each row is a `dict` itself.

        :returns: `dict`, a report on the consistency metrics the date in the
            resource:
            * `value`, `float`, percentage of consistent records in the data.
            * `consistent`, `int`, number of consistent records.
            * `total`, `int`, total number ofrecords.
            * `report`, `dict`, detailed, per column, report for the
                consistency of the data.
        '''
        log.debug('-----consistency start-------')
        # log.debug(data['fields'])
        validators = self.get_consistency_validators()
        fields = {f['id']: f for f in data['fields']}
        report = {f['id']: {'count': 0, 'formats': {}} for f in data['fields']}  
        #
        # log.debug(data)
        # log.debug(fields)
        count_row=0
        log.debug('-----consistency record--------') # version แรกใช้ data record
        field_names = [f['id'] for f in data['fields']]
        # ข้าม header แถวแรกของ raw_data
        for row in data['raw_data'][1:]:#data['records']:
            # count_row=count_row+1
            # for field, value in row.items():
            count_row += 1
            row_dict = dict(zip(field_names, row))  # แปลง list -> dict
            log.debug(row_dict)
            for field, value in row_dict.items():
                # ข้ามค่าว่าง
                if value is None or (isinstance(value, str) and str(value).strip() == ""):
                    continue  
                field_type = fields.get(field, {}).get('type')
                validator = validators.get(field_type)
                field_report = report[field]         
                if validator:
                    validator(field, value, field_type, field_report)
                    field_report['count'] += 1
      
        for field, field_report in report.items():
            merged_data = {}   # <-- สร้าง dict ใหม่
            datetime_formats = {
                '%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d',
                '%Y-%m-%d %H:%M:%S', '%d/%m/%Y %H:%M:%S',
                '%Y-%m-%dT%H:%M:%S', '%Y/%m/%d %H:%M',
            }
            numeric_count = 0
            timestamp_count = 0
            format_dict = field_report['formats'] 
            log.debug('--format_dict before--')
            log.debug(field) 
            log.debug(format_dict) 
            for key, value in format_dict.items():
                if key in {'int', 'float', 'unknown',
                        '^\\d+$', '^[+-]\\d+$',
                        '^(\\d{1,3},)+(\\d{3})$', '^(\\d{1,3},)+(\\d{3})\\.\\d+$',
                        '^[+-](\\d{1,3},)(\\d{3})+$', '^[+-](\\d{1,3},)+(\\d{3})\\.\\d+$',
                        '^\\d+\\.\\d+$', '^[+-]\\d+\\.\\d+$'}:
                    numeric_count  += value      # <-- นับแต่ไม่ใส่ key เดิม       
                elif key in datetime_formats:
                    timestamp_count  += value
                else:
                    # เก็บ string/text หรือ format ที่ไม่ใช่ numeric/timestamp
                    # merged_data[key] = merged_data.get(key, 0) + value
                    merged_data[key] = value


            # เพิ่ม numeric/timestamp ลงไป
            if numeric_count > 0:
                merged_data['numeric'] = numeric_count
            if timestamp_count > 0:
                merged_data['timestamp'] = timestamp_count

            # เขียนกลับ → เอา key เดิมออกแล้วรวม numeric/timestamp/text
            field_report['formats'] = merged_data
                
            # log.debug('--format_dict--')
            # log.debug(field) 
            # log.debug(format_dict) 
            # log.debug(merged_data)
            #[End Pang Edit]--------------------------------------------------------------------
            if field_report['formats']:
                #เลือก count format ที่มีค่ามากสุด เช่น int 30 float 4 แปลว่าคอลัมน์นี้คือ int
                most_consistent = max([count if fmt != 'unknown' else 0
                                    for fmt, count
                                    in field_report['formats'].items()])
                field_report['consistent'] = most_consistent
        #total = จำนวนค่าทั้งหมดที่ตรวจได้ column * row 
        total = sum([f.get('count', 0) for _, f in report.items()])
        consistent = sum([f.get('consistent', 0) for _, f in report.items()])
        value = float(consistent)/float(total) * 100.0 if total else 0.0
        #ไม่มีข้อมูลให้ตรวจเลย เช่น data['records'] ว่าง หรือแต่ละ record ไม่มี field
        if total == 0:
            return {
                'error': 'Data is empty',
                'value': None,
                'report': report
            }
        #---- check key in report {} : for excel, if key is not str type, set report null because the data structure is invalid---
        list_key = report.keys()
        for key_item in list_key:
            if not isinstance(key_item, str):
                report = {}
                break
        #------------------------------------------------
        # if (consistent == None):
        #     consistent = 0
        return {
            'total': total,
            'consistent': consistent,
            'value': round(value, 2),
            'report': report
        }
        
      # for field, field_report in report.items():
        #     #------------------------------------------------------------------
        #     # # New dictionary to store the merged result
        #     merged_data = {'numeric': 0,'timestamp': 0}
        #     datetime_formats = {
        #         '%Y-%m-%d',
        #         '%d/%m/%Y',
        #         '%Y-%m-%d',
        #         '%Y/%m/%d',
        #         '%Y-%m-%d %H:%M:%S',
        #         '%d/%m/%Y %H:%M:%S',
        #         '%Y-%m-%dT%H:%M:%S',
        #         '%Y/%m/%d %H:%M',
        #         # เพิ่ม pattern ที่อยากรองรับ
        #     }
        #     chk_timestamp = False

            
        #     #[Pang Edit]update report for numeric values ==> merge int, float, unknown to numeric
        #     format_dict = field_report['formats']   
        #     keys_to_merge = {'int', 'float', 'unknown',
        #                     '^\\d+$',
        #                     '^[+-]\\d+$',
        #                     '^(\\d{1,3},)+(\\d{3})$',
        #                     '^(\\d{1,3},)+(\\d{3})\\.\\d+$',
        #                     '^[+-](\\d{1,3},)(\\d{3})+$',
        #                     '^[+-](\\d{1,3},)+(\\d{3})\\.\\d+$',
        #                     '^\\d+\\.\\d+$',
        #                     '^[+-]\\d+\\.\\d+$'
        #                     } 
            
        #     # # Iterate through the original dictionary
        #     chk_numeric = False
        #     for key, value in format_dict.items():
        #         if key in keys_to_merge:
        #             merged_data['numeric'] += value
        #             chk_numeric = True
        #         if key in datetime_formats:
        #             merged_data['timestamp'] += value
        #             chk_timestamp = True
        #     # log.debug('--format_dict--')
        #     # log.debug(field) 
        #     # log.debug(format_dict) 
        #     # log.debug(merged_data)
        #     # if(chk_numeric):
        #     #     field_report['formats'] = merged_data
        #     # แค่มีอย่างใดอย่างหนึ่ง ก็เขียนกลับ
        #     if merged_data['timestamp'] > 0 or merged_data['numeric'] > 0:
        #         field_report['formats'] = merged_data
        #         log.debug(field_report['formats']) 
        #-------------       
    def calculate_cumulative_metric(self, resources, metrics):
        '''Calculates the total percentage of consistent values in the data for
        all the given resources.

        :param resources: `list` of CKAN resources.
        :param metrics: `list` of `dict` results for each resource.

        :returns: `dict`, a report on the consistency metrics the date in the
            resource:
            * `value`, `float`, percentage of consistent records in the data.
            * `consistent`, `int`, number of consistent records.
            * `total`, `int`, total number ofrecords.
            * `report`, `dict`, detailed, per column, report for the
                consistency of the data.
        '''
        # total = sum([r.get('total', 0) for r in metrics])
        # consistent = sum([r.get('consistent', 0) for r in metrics])

        # # FIXME: This is not the proper calculation. We need to merge all
        # # data to calculate the consistency properly.
        # value = float(consistent)/float(total) * 100.0 if total else 0.0
        # return {
        #     'total': total,
        #     'consistent': consistent,
        #     'value': value,
        # }
        N_total = len(resources)
        N_consistency = 0 
        for item_metric in metrics:
            #check dict is not Empty
            if item_metric and isinstance(item_metric.get('value'), (int, float)):
                val = item_metric.get('value')
                if val:
                    N_consistency += val
                
            else:
                N_total -= 1

        consistency_score = (N_consistency / N_total ) if N_total > 0 else 0.0
        return {
            "N_total": N_total,
            "N_consistency": N_consistency,
            "value": round(consistency_score, 2)
        }
class Timeliness():#DimensionMetric
    '''Calculates the timeliness Data Quality dimension.

    The timeliness of a records is the difference between the times when the
    measurement was made and the time that record have entered our system.

    The measurement requires extra configuration - the name of the column in
    the resource data that holds the time of when the measurement was made. If
    this time is missing or the setting is not configured, then the calculation
    cannot be performed automatically.

    The calculation is performed by taking a time delta between the time that
    the resource have been last modified (the time when the data have entered
    and was stored in CKAN) and the time in the specified column, for each
    record.

    The total difference (time delta in resolution of seconds) is then divided
    by the number of records checked, givin the average time delta in seconds.
    This average is then used as the value for this dimension.
    '''
    def __init__(self):
        # super(Timeliness, self).__init__('timeliness')
        self.name = 'timeliness'

    def calculate_metric(self, resource): #, data
        '''Calculates the timeliness of the data in the given resource.

        :param resource: `dict`, CKAN resource. The resource is expected to
            have a property `data_quality_settings` that configures the column
            that holds the timestamp of when the measurement was taken. This
            property is a `dict`, having the following structure:

                .. code-block: python

                    resoruce = {
                        ...
                        'data_quality_settings': {
                            'timeliness': {
                                'column': '<the name of the column',
                            }
                        }
                    }

        :param data: `dict`, the resource data as a dict with the following
            values:
                * `total`, `int`, total number of rows.
                * `fields`, `list` of `dict`, column metadata - name, type.
                * `records`, `iterable`, iterable over the rows in the resource
                    where each row is a `dict` itself.

        :returns: `dict`, a report on the timeliness metrics for the given
            resource data:
            * `value`, `str`, string representation of the average time delta,
                for example: `'+3:24:00'`, `'+3 days, 3:24:00'`
            * `total`, `int`, total number of seconds that the measurements
                have beed delayed.
            * `average`, `int`, the average delay in seocnds.
            * `records`, `int`, number of checked records.
        '''
        package_id = resource.get('package_id')
        packages = Package.get(package_id)
        update_frequency_unit = Session.query(PackageExtra).filter(PackageExtra.key == 'update_frequency_unit', PackageExtra.package_id == package_id).first()
        update_frequency_interval = Session.query(PackageExtra).filter(PackageExtra.key == 'update_frequency_interval', PackageExtra.package_id == package_id).first()

        if resource.get('last_modified') is not None:
            created = dateutil.parser.parse(resource.get('last_modified') or resource.get('created'))
        else:
            created = dateutil.parser.parse(resource.get('created'))


        # measured_.count = 0
        # total_delta = 0
        # diff_date = (created.date() - datetime.now().date()).days
        # tln = abs((created.date() - datetime.now().date()).days)
        # if update_frequency_unit.value == u'วัน':
        #     if update_frequency_interval.value != '':
        #         tln_val = int(update_frequency_interval.value) - tln
        #     else:
        #         tln_val = 1 - tln
        # elif update_frequency_unit.value == u'สัปดาห์':
        #     if update_frequency_interval.value != '':
        #         tln_val = (7*int(update_frequency_interval.value)) - tln
        #     else:
        #         tln_val = 7 - tln
        # elif update_frequency_unit.value == u'เดือน':
        #     if update_frequency_interval.value != '':
        #         tln_val = (30 * int(update_frequency_interval.value)) - tln
        #     else:
        #         tln_val = 30 - tln
        # elif update_frequency_unit.value == u'ไตรมาส':
        #     if update_frequency_interval.value != '':
        #         tln_val = (90 * int(update_frequency_interval.value)) - tln
        #     else:
        #         tln_val = 90 - tln
        # elif update_frequency_unit.value == u'ครึ่งปี':
        #     if update_frequency_interval.value != '':
        #         tln_val = (180 * int(update_frequency_interval.value)) - tln
        #     else:
        #         tln_val = 180 - tln
        # elif update_frequency_unit.value == u'ปี':
        #     if update_frequency_interval.value != '':
        #         tln_val =  (365 * int(update_frequency_interval.value)) - tln
        #     else:
        #         tln_val = 365 - tln
        # else:
        #     tln_val = 999

        # return {         
        #     'frequency': update_frequency_unit.value,
        #     'value': tln_val,
        #     'date_diff': diff_date
        # }
        measured_count = 0
        total_delta = 0
        elapsed_days = (datetime.now().date() - created.date()).days
        created_date_str = created.date().strftime('%Y-%m-%d')
        overdue_day =  0
        update_cycle_days = 0
        freshness = 0
        # ดึงความถี่ ปกติใส่เป็นตัวเลข
        # ดึงข้อความจากฟิลด์
        value = update_frequency_interval.value or ''  # ป้องกัน None
        match = re.search(r'\d+', value)

        # ถ้า match ได้เลข ให้ใช้เลขนั้น ไม่งั้นใช้ default = 1
        interval = int(match.group()) if match else 1

        unit = update_frequency_unit.value or ''

        if unit == u'วัน':
            update_cycle_days = interval
        elif unit == u'สัปดาห์':
            update_cycle_days = 7 * interval
        elif unit == u'เดือน':
            update_cycle_days = 30 * interval
        elif unit == u'ไตรมาส':
            update_cycle_days = 90 * interval
        elif unit == u'ครึ่งปี':
            update_cycle_days = 180 * interval
        elif unit == u'ปี':
            update_cycle_days = 365 * interval
        else:
            update_cycle_days = None  # หรือ 0 หรือ raise Exception
 
        # if update_frequency_unit.value == u'วัน':
        #     if update_frequency_interval.value != '':
        #         update_cycle_days = int(update_frequency_interval.value)
        #     else:
        #         update_cycle_days = 1
        # elif update_frequency_unit.value == u'สัปดาห์':
        #     if update_frequency_interval.value != '':
        #         update_cycle_days = (7*int(update_frequency_interval.value))
        #     else:
        #         update_cycle_days = 7
        # elif update_frequency_unit.value == u'เดือน':
        #     if update_frequency_interval.value != '':
        #         update_cycle_days = (30 * int(update_frequency_interval.value)) 
        #     else:
        #         update_cycle_days = 30
        # elif update_frequency_unit.value == u'ไตรมาส':
        #     if update_frequency_interval.value != '':
        #         update_cycle_days = (90 * int(update_frequency_interval.value))
        #     else:
        #         update_cycle_days = 90 
        # elif update_frequency_unit.value == u'ครึ่งปี':
        #     if update_frequency_interval.value != '':
        #         update_cycle_days = (180 * int(update_frequency_interval.value))
        #     else:
        #         update_cycle_days = 180
        # elif update_frequency_unit.value == u'ปี':
        #     if update_frequency_interval.value != '':
        #         update_cycle_days =  (365 * int(update_frequency_interval.value))
        #     else:
        #         update_cycle_days = 365

        if  update_cycle_days:
            overdue_day =  elapsed_days - update_cycle_days
            safe_overdue = max(0, overdue_day) #abs(overdue_day)#
            acceptable_latency = (safe_overdue / update_cycle_days) * 100
            #freshness = ข้อมูลใหม่แค่ไหน, 100% = เพิ่งอัปเดต, 0% ถึงรอบพอดี, ติดลบข้อมูลล่าช้า
            freshness = (update_cycle_days - elapsed_days) / update_cycle_days * 100

            log.debug('resource.id')
            log.debug(resource.get('id'))
            log.debug('created.date()')
            log.debug(created.date())
            log.debug(created_date_str)
            log.debug('elapsed_days')
            log.debug(elapsed_days)
            log.debug('overdue_day')
            log.debug(overdue_day)
            log.debug('safe_overdue')
            log.debug(safe_overdue)
            log.debug('update_cycle_days')
            log.debug(update_cycle_days)
            log.debug(update_frequency_unit.value)
        else:
            acceptable_latency = -1  # ไม่ได้กำหนดค่า หรือกำหนดเป็นค่าอื่นๆ
            freshness = -1
            safe_overdue = ""
        
        #----Timeliness------------
        timeliness = 0
        if acceptable_latency == 0:
            timeliness = 0 # อัพเดทตรงเวลา
        elif acceptable_latency>0 and acceptable_latency <= 25:
            timeliness = 1 # รบกวนปรับปรุง
        elif acceptable_latency>25 and acceptable_latency <= 50:
            timeliness = 2 # ควรปรับปรุง
        elif acceptable_latency>50 and acceptable_latency <= 100:
            timeliness = 3 # ต้องปรับปรุง
        elif acceptable_latency > 100:
            timeliness = 4 # ล้าสมัยเกินไป
        elif acceptable_latency == -1:
            timeliness = -1 # ไม่ได้กำหนด
        
        return {         
            'frequency': update_frequency_unit.value,
            'value': timeliness,
            'acceptable_latency': round(acceptable_latency,2),
            'freshness':round(freshness,2),
            'updated_date': created_date_str,
            'elapsed_days': elapsed_days,
            'update_cycle_days':update_cycle_days,
            'safe_overdue': safe_overdue
        }
    def calculate_cumulative_metric(self, resources, metrics):
        '''Calculates the timeliness of all data in all of the given resources
        based on the results for each individual resource.

        The average delay is calculated by getting the total number of seconds
        that all records have been delayed, then dividing that to the number of
        total records in all the data.

        :param resources: `list` of CKAN resources.
        :param metrics: `list` of `dict` results for each resource.

        :returns: `dict`, a report for the timeliness of the data:
            * `value`, `str`, string representation of the average time delta,
                for example: `'+3:24:00'`, `'+3 days, 3:24:00'`
            * `total`, `int`, total number of seconds that the measurements
                have beed delayed.
            * `average`, `int`, the average delay in seocnds.
            * `records`, `int`, number of checked records.
        '''
        timeliness_list = []
        total = 0
        
        for item_metric in metrics:
            #check dict is not Empty
            # if ((item_metric) or (item_metric is not None)):
            if item_metric and isinstance(item_metric.get('value'), (int, float)):
                timeliness_score = item_metric.get('value')
                total = total+timeliness_score
                timeliness_list.append(timeliness_score)
        if timeliness_list:
            result_score = min(timeliness_list)
            return {
                'total': total,
                'value': result_score,
            }
        else:
            return {
                'total': 0,
                'value': 0,
            }
class AcceptableLatency():
    '''Calculates the timeliness Data Quality dimension.

    The timeliness of a records is the difference between the times when the
    measurement was made and the time that record have entered our system.

    The measurement requires extra configuration - the name of the column in
    the resource data that holds the time of when the measurement was made. If
    this time is missing or the setting is not configured, then the calculation
    cannot be performed automatically.

    The calculation is performed by taking a time delta between the time that
    the resource have been last modified (the time when the data have entered
    and was stored in CKAN) and the time in the specified column, for each
    record.

    The total difference (time delta in resolution of seconds) is then divided
    by the number of records checked, givin the average time delta in seconds.
    This average is then used as the value for this dimension.
    '''
    def __init__(self):
        # super(Timeliness, self).__init__('timeliness')
        self.name = 'acc_latency'

    def calculate_metric(self, resource,timeliness_val): 
   
        log.debug('resource.id')
        log.debug(resource.get('id'))
       
        acc_latency = timeliness_val.get('acceptable_latency')
        return {                   
            'value': acc_latency,
        }
    
    def calculate_cumulative_metric(self, resources, metrics):
        '''Calculates the timeliness of all data in all of the given resources
        based on the results for each individual resource.

        The average delay is calculated by getting the total number of seconds
        that all records have been delayed, then dividing that to the number of
        total records in all the data.

        :param resources: `list` of CKAN resources.
        :param metrics: `list` of `dict` results for each resource.

        :returns: `dict`, a report for the timeliness of the data:
            * `value`, `str`, string representation of the average time delta,
                for example: `'+3:24:00'`, `'+3 days, 3:24:00'`
            * `total`, `int`, total number of seconds that the measurements
                have beed delayed.
            * `average`, `int`, the average delay in seocnds.
            * `records`, `int`, number of checked records.
        '''
        acc_latency_list = []
        total = 0
        
        for item_metric in metrics:
            #check dict is not Empty
            if item_metric and isinstance(item_metric.get('value'), (int, float)):
                acc_latency_score = item_metric.get('value')
                total = total+acc_latency_score
                acc_latency_list.append(acc_latency_score)
        if acc_latency_list:
            result_score = min(acc_latency_list)
            # if(result_score < 0):
            #     result_score = 0
            return {
                'total': total,
                'value': result_score,
            }
        else:
            return {
                'total': 0,
                'value': 0,
            }
        # N_total = len(resources)
        # N_acc_latency = 0 
        # for item_metric in metrics:
        #     #check dict is not Empty
        #     if item_metric and isinstance(item_metric.get('value'), (int, float)):
        #         val = item_metric.get('value')
        #         if val:
        #             N_acc_latency += val
                
        #     else:
        #         N_total -= 1
        # #avg N_acc_latency is percent, so we will not *100
        # acc_latency_score = (N_acc_latency / N_total) if N_total > 0 else 0.0
        # log.debug('acc_latency')
        # log.debug(acc_latency_score)
        # return {
        #     "N_total": N_total,
        #     "N_acc_latency": N_acc_latency,
        #     "value": round(acc_latency_score, 2)
        # }
class Freshness():
    '''Calculates the timeliness Data Quality dimension.

    The timeliness of a records is the difference between the times when the
    measurement was made and the time that record have entered our system.

    The measurement requires extra configuration - the name of the column in
    the resource data that holds the time of when the measurement was made. If
    this time is missing or the setting is not configured, then the calculation
    cannot be performed automatically.

    The calculation is performed by taking a time delta between the time that
    the resource have been last modified (the time when the data have entered
    and was stored in CKAN) and the time in the specified column, for each
    record.

    The total difference (time delta in resolution of seconds) is then divided
    by the number of records checked, givin the average time delta in seconds.
    This average is then used as the value for this dimension.
    '''
    def __init__(self):
        # super(Timeliness, self).__init__('timeliness')
        self.name = 'freshness'

    def calculate_metric(self, resource,timeliness_val): 
   
        log.debug('resource.id')
        log.debug(resource.get('id'))
       
        freshness = timeliness_val.get('freshness')
        return {                   
            'value': freshness,
            'timeliness': timeliness_val
        }
    
    def calculate_cumulative_metric(self, resources, metrics):
        '''Calculates the timeliness of all data in all of the given resources
        based on the results for each individual resource.

        The average delay is calculated by getting the total number of seconds
        that all records have been delayed, then dividing that to the number of
        total records in all the data.

        :param resources: `list` of CKAN resources.
        :param metrics: `list` of `dict` results for each resource.

        :returns: `dict`, a report for the timeliness of the data:
            * `value`, `str`, string representation of the average time delta,
                for example: `'+3:24:00'`, `'+3 days, 3:24:00'`
            * `total`, `int`, total number of seconds that the measurements
                have beed delayed.
            * `average`, `int`, the average delay in seocnds.
            * `records`, `int`, number of checked records.
        '''
        freshness_list = []
        total = 0
        
        for item_metric in metrics:
            #check dict is not Empty
            if item_metric and isinstance(item_metric.get('value'), (int, float)):
                freshness_score = item_metric.get('value')
                total = total+freshness_score
                freshness_list.append(freshness_score)
        if freshness_list:
            result_score = max(freshness_list)
            if result_score < 0: #for showing in radar chart
               result_score = 0 
            return {
                'total': total,
                'value': result_score,
            }
        else:
            return {
                'total': 0,
                'value': 0,
            }
        # N_total = len(resources)
        # positive = 0
        # for item_metric in metrics:
        #     #check dict is not Empty
        #     if item_metric and isinstance(item_metric.get('value'), (int, float)):
        #         val = item_metric.get('value')
        #         if(val >= 0):
        #             positive += 1
                
        #     else:
        #         N_total -= 1
        # freshness_score = (positive / N_total) * 100 if N_total else 0.0
        # return {
        #         'value': round(freshness_score,2),
        #         'total': N_total
        #     }
      
_all_date_formats = [
    '%Y-%m-%d',
    '%y-%m-%d',
    '%Y/%m/%d',
    '%y/%m/%d',
    '%Y.%m.%d',
    '%y.%m.%d',
    '%d-%m-%Y',
    '%d-%m-%y',
    '%d/%m/%Y',
    '%d/%m/%y',
    '%d.%m.%Y',
    '%d.%m.%y',
    '%m-%d-%Y',
    '%m-%d-%y',
    '%m/%d/%Y',
    '%m/%d/%y',
    '%m.%d.%Y',
    '%m.%d.%y',
]


def _generate_time_formats():
    additional = []
    for time_sep in ['T', ' ', ', ', '']:
        if time_sep:
            for date_fmt in _all_date_formats:
                for time_fmt in ['%H:%M:%S',
                                 '%H:%M:%S.%f',
                                 '%H:%M:%S.%fZ',
                                 '%H:%M:%S.%f%Z',
                                 '%H:%M:%S.%f%z']:
                    additional.append(date_fmt + time_sep + time_fmt)
    return additional


_all_date_formats += _generate_time_formats()


_all_numeric_formats = [
    r'^\d+$',
    r'^[+-]\d+$',
    r'^(\d{1,3},)+(\d{3})$',
    r'^(\d{1,3},)+(\d{3})\.\d+$',
    r'^[+-](\d{1,3},)(\d{3})+$',
    r'^[+-](\d{1,3},)+(\d{3})\.\d+$',
    r'^\d+\.\d+$',
    r'^[+-]\d+\.\d+$',
]


def detect_date_format(datestr):
    '''Tries to detect the date-time format from the given date or timestamp
    string.

    :param datestr: `str`, the date string

    :returns: `str`, the guessed format of the date, otherwise `None`.
    '''
    try:
        datestr = str(datestr)
        log.debug(datestr)
        if re.match(r'$\d^', datestr):
            return 'unix-timestamp'
        if re.match(r'$\d+\.\d+', datestr):
            return 'timestamp'
        if re.match(r'$\d+[\+\-]\d+^', datestr):
            return 'timestamp-tz'
        for dt_format in _all_date_formats:
            try:
                datetime.strptime(datestr, dt_format)
                return dt_format
            except:
                pass
    except Exception as e:
        log.debug("An error occurred in detect_date_format", e)
        return None 
          
def detect_numeric_format(numstr):
    '''Tries to detect the format of a number (int, float, number format with
    specific separator such as comma "," etc).

    :param numstr: `str`, `int`, `float`, the number string (or numeric value)
        to try to guess the format for.

    :returns: `str`, the guessed number format, otherwise `None`.
    '''
    try:
        for parser in [int, float]:
            try:
                parser(numstr)
                return parser.__name__
            except:
                pass

        for num_format in _all_numeric_formats:
            m = re.match(num_format, str(numstr))
            if m:
                log.debug(num_format)
                return num_format
        
    except Exception as e:
        log.debug("An error occurred in detect_numeric_format", e)
        return None            

# resource validation
# def validate_resource_data(resource,data):
#     '''Performs a validation of a resource data, given the resource metadata.

#     :param resource: CKAN resource data_dict.

#     :returns: `dict`, a validation report for the resource data.
#     '''
#     log.debug(u'Validating resource {}'.format(resource['id']))

#     options = toolkit.config.get(
#         u'ckanext.validation.default_validation_options')
#     if options:
#         options = json.loads(options)
#     else:
#         options = {}
#     resource_options = resource.get(u'validation_options')
#     if resource_options and isinstance(resource_options, string_types):#basestring):
#         resource_options = json.loads(resource_options)
#     if resource_options:
#         options.update(resource_options)

#     dataset = toolkit.get_action('package_show')(
#         {'ignore_auth': True}, {'id': resource['package_id']})

#     source = None
#     if resource.get(u'url_type') == u'upload':
#         upload = uploader.get_resource_uploader(resource)
#         if isinstance(upload, uploader.ResourceUpload):
#             source = upload.get_path(resource[u'id'])
#         else:
#             # Upload is not the default implementation (ie it's a cloud storage
#             # implementation)
#             pass_auth_header = toolkit.asbool(
#                 toolkit.config.get(u'ckanext.validation.pass_auth_header',
#                                    True))
#             if dataset[u'private'] and pass_auth_header:
#                 s = requests.Session()
#                 s.headers.update({
#                     u'Authorization': t.config.get(
#                         u'ckanext.validation.pass_auth_header_value',
#                         _get_site_user_api_key())
#                 })

#                 options[u'http_session'] = s
#     if not source:
#         source = resource[u'url']

#     schema = resource.get(u'schema')
#     if schema and isinstance(schema, string_types):#basestring):
#         if schema.startswith('http'):
#             r = requests.get(schema)
#             schema = r.json()
#         else:
#             schema = json.loads(schema)

#     _format = resource[u'format'].lower()
#     #----- check mimetype -----------------------------
#     log.debug('--validate: check mimetype--')
#     log.debug('---data records--')
#     # df = pd.DataFrame(data['records'])
#     records = list(data['records']) 
#     # log.debug(records)
#     # log.debug('---dataframe--')
#     # log.debug(df)
#     #--------------------------
#     report = _validate_table_by_list(records)
#     # report = _validate_table_by_datastore(df)
#     # log.debug(report)
#     # mimetype =  ResourceFetchData2.detect_mimetype(source)
#     # log.debug('---mimetype--')
#     # log.debug(mimetype)
#     # if (mimetype == 'text/csv'):
#     #     if ResourceFetchData2.has_valid_filename(source,'.csv'):
#     #         log.debug('---validate:readfile csv--')
#     #         report = _validate_table(source, _format=_format, schema=schema, **options)
#     #         report['validate_source'] = 'file'
#     #     else:
#     #         log.debug('---validate:datastore csv--')
#     #         report = _validate_table_by_datastore(df)
#     #         report['validate_source'] = 'datastore'

#     # elif(mimetype == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ):
#     #     if ResourceFetchData2.has_valid_filename(source,'.xlsx'):
#     #         log.debug('has_valid_filename')
#     #         report = _validate_table(source, _format=_format, schema=schema, **options)
#     #         report['validate_source'] = 'file'
#     #     else:
#     #         log.debug('not has_valid_filename')
#     #         report = _validate_table_by_datastore(df)
#     #         report['validate_source'] = 'datastore'

#     # elif(mimetype == 'application/vnd.ms-excel'):
#     #     log.debug(source)
#     #     if ResourceFetchData2.has_valid_filename(source,'.xls'):
#     #         log.debug('has_valid_filename')
#     #         report = _validate_table(source, _format=_format, schema=schema, **options)
#     #         report['validate_source'] = 'file'
#     #     else:
#     #         log.debug('not has_valid_filename')
#     #         report = _validate_table_by_datastore(df)
#     #         report['validate_source'] = 'datastore'
#     # else:
#     #     report = _validate_table(source, _format=_format, schema=schema, **options)

#     # log.debug(report)
#     # if (report.get('error') != ''):
#     #     report = _validate_table_by_datastore(df)
#     #     report['validate_source'] = 'datastore'
#     #     log.debug(report)
#     # report = _validate_table(source, _format=_format, schema=schema, **options)
#     # report = _validate_table_by_ByteIO(source, _format=_format, schema=schema, **options)

#     # Hide uploaded files
#     for table in report.get('tables', []):
#         if table['source'].startswith('/'):
#             table['source'] = resource['url']
#     for index, warning in enumerate(report.get('warnings', [])):
#         report['warnings'][index] = re.sub(r'Table ".*"', 'Table', warning)

#     return report
#--- old extra value-----
# def detect_extra_columns_excel(records, sample_size=20):
#     # แปลง dict เป็น list ของ row values
#     rows = [list(record.values()) for record in records]

#     # ตรวจว่ามีข้อมูลหรือไม่
#     if not rows:
#         return {
#             "expected_columns": 0,
#             "extra_rows": []
#         }

#     # 1) ตรวจหา "จำนวนคอลัมน์ที่มีข้อมูลจริง" จาก sample rows
#     sample_rows = rows[1:sample_size+1]  # ข้าม header
#     def count_nonempty_except_last(row):
#         if not row:
#             return 0
#         trimmed = row[:-1] if str(row[-1]).strip() == '' else row
#         return sum(1 for cell in trimmed if str(cell).strip() != '')

#     col_counts = [count_nonempty_except_last(r) for r in sample_rows if r]
#     if not col_counts:
#         return {
#             "expected_columns": 0,
#             "extra_rows": [],
#             "error": "No valid sample rows"
#         }

#     expected_columns = Counter(col_counts).most_common(1)[0][0]

#     # 2) ตรวจหาว่าแถวไหนมี extra values เกิน expected columns
#     extra_rows = []
#     for i, row in enumerate(rows, start=1):
#         # ตรวจว่าเกิน expected_columns หรือไม่
#         if len(row) > expected_columns:
#             extra_values = row[expected_columns:]
#             if any(str(cell).strip() != '' for cell in extra_values):  # ต้องไม่ใช่ช่องว่างล้วน
#             # if any(cell.strip() != '' for cell in extra_values):  # ต้องไม่ใช่ช่องว่างล้วน
#                 extra_rows.append({
#                     "row_number": i,
#                     "columns": len(row),
#                     "extra_values": extra_values
#                 })

#     return {
#         "expected_columns": expected_columns,
#         "extra_rows": extra_rows
#     }
#---- new extra value----------------------------------

def detect_extra_columns_from_rows(rows, expected_columns):
    extra_rows = []
    for i, row in enumerate(rows, start=1):
        # รองรับ dict เช่นจาก Excel
        if isinstance(row, dict):
            ordered_values = [row[k] for k in sorted(row.keys(), key=lambda x: str(x) if x is not None else 'zzz')]
        else:
            ordered_values = list(row)

        # ตัดช่องว่างท้ายแถวออก (None, '', '  ')
        trimmed_values = list(ordered_values)
        while trimmed_values and (trimmed_values[-1] is None or str(trimmed_values[-1]).strip() == ''):
            trimmed_values.pop()

        # ถ้าเหลือมากกว่า expected_columns ถือว่าเกิน
        if len(trimmed_values) > expected_columns:
            extra_values = trimmed_values[expected_columns:]
            extra_rows.append({
                "row_number": i,
                "columns": len(trimmed_values),
                "extra_values": extra_values
            })

    return extra_rows
# def detect_columns_and_validate(records):
#     # แปลง dict เป็น list ของ row values
#     # rows = [list(record.values()) for record in records]
#     # ตรวจว่ามีข้อมูลอย่างน้อย 2 แถว (1 header + 1 data)
#     if not records or len(records) < 2:
#         return {
#             "expected_columns": 0,
#             "extra_rows": []
#         }

#     # ใช้แถวที่ 1 เป็น header
#     header = records[1]
#     data_rows = records[2:]  # ข้ามแถวคำอธิบาย (row 0)

#     # แปลงเป็น list of dict (ใช้ header)
#     rows = [dict(zip(header, row)) for row in data_rows]
#     # 1) ตรวจหา "จำนวนคอลัมน์ที่มีข้อมูลจริง" จาก sample rows
#     sample_size=20
#     sample_rows = rows[:sample_size] 
#     log.debug('---sample_rows---')
#     log.debug(sample_rows)
# #     # sample_rows = rows[1:sample_size+1]  # ข้าม header
#     def count_nonempty_except_last(row):
#         """
#         รับได้ทั้ง row ที่เป็น list หรือ dict
#         แปลงค่าทุก cell เป็น string และนับ cell ที่ไม่ว่าง (ยกเว้นช่องสุดท้ายถ้าว่าง)
#         """
#         if not row:
#             return 0
#         # รองรับกรณี row เป็น dict (เช่นจาก XLSX) → แปลงเป็น list
#         if isinstance(row, dict):
#             # sort ตาม key เพื่อความสม่ำเสมอ ถ้า key เป็น None จะไว้ท้าย
#             ordered_values = [row[k] for k in sorted(row.keys(), key=lambda x: str(x) if x is not None else 'zzz')]
#         else:
#             ordered_values = row
#         # แปลงทุก cell เป็น string และ trim
#         str_row = [str(cell).strip() if cell is not None else '' for cell in ordered_values]
#         # ถ้าช่องสุดท้ายว่าง → ไม่เอา
#         trimmed = str_row[:-1] if str_row and str_row[-1] == '' else str_row
#         # นับช่องที่มีข้อมูลจริง
#         return sum(1 for cell in trimmed if cell != '')
    
#     #กรองแถวที่มีข้อมูลน้อยกว่า 50% ทิ้งไป
#     cleaned_sample_rows = [
#         r for r in sample_rows if r and count_nonempty_except_last(r) >= len(r) * 0.5
#     ]
#     col_counts = [count_nonempty_except_last(r) for r in cleaned_sample_rows]

#     # col_counts = [count_nonempty_except_last(r) for r in sample_rows if r]
#     log.debug('---cleaned_sample_rows---')
#     log.debug(cleaned_sample_rows)
#     # col_counts = [count_nonempty_except_last(r) for r in sample_rows if r]
#     if not col_counts:
#         expected_columns = 0
#     else:
#         expected_columns, freq = Counter(col_counts).most_common(1)[0]

#     # new logic: sampling_confidence based on missing value ratio
#     total_cells = 0
#     total_missing = 0
#     for row in sample_rows:
#         if isinstance(row, dict):
#             # แปลง dict เป็น list ของ values (sorted keys รวม None)
#             ordered_values = [row[k] for k in sorted(row.keys(), key=lambda x: str(x) if x is not None else 'zzz')]
#         else:
#             ordered_values = row

#         for cell in ordered_values:
#             total_cells += 1
#             # กรณี cell เป็น None, '' หรือ '   ' → ถือว่า missing
#             if cell is None or str(cell).strip() == '':
#                 total_missing += 1
#     # log.debug('missing/total')            
#     # log.debug(total_missing)
#     # log.debug(total_cells)
#     if total_cells == 0:
#         sampling_confidence = 0
#     else:
#         sampling_confidence = 1 - (total_missing / total_cells)

#     sampling_confidence_percent = round(sampling_confidence * 100, 2)

#     result = {
#         "method": "sampling",
#         "expected_columns": expected_columns,
#         "header_row": None,
#         "sampling_confidence": sampling_confidence_percent
#     }
    
#     if sampling_confidence_percent >= 80:
#         result["extra_rows"] = detect_extra_columns_from_rows(rows, expected_columns)
#     else:
#         result["extra_rows"] = []
#         result["note"] = "Low confidence from sampling; skipped extra value check"
#     return result
#---------[Start] detect extra value ---
# ---------- normalize rows -----------
def normalize_row(row: List[Any], trim_trailing_empty: bool = True) -> List[str]:
    """Convert each cell to string, strip, convert None -> ''. Optionally trim trailing empty cells."""
    str_cells = []
    for c in row:
        if c is None:
            s = ''
        else:
            s = str(c).strip()
        str_cells.append(s)
    if trim_trailing_empty:
        # pop trailing empty cells
        while str_cells and str_cells[-1] == '':
            str_cells.pop()
    return str_cells

# ---------- header detection ----------
def analyze_header_candidates(rows: List[List[str]], top_n: int = 10) -> Dict[str, Any]:
    """
    Examine top_n rows, produce best header candidate with confidence.
    returns dict: { 'line_no', 'row', 'num_columns', 'confidence' }
    """
    def is_label_like(cell: str) -> bool:
        # label-like if not empty and not numeric-only and shortish
        if not cell: 
            return False
        s = cell.strip()
        # consider numeric if digits and optionally comma/dot
        digits = s.replace(',', '').replace('.', '').isdigit()
        if digits:
            return False
        # also treat if contains many letters
        return len(s) <= 200

    candidates = []
    for i in range(min(top_n, len(rows))):
        r = normalize_row(rows[i], trim_trailing_empty=False)  # keep length to judge
        if not r:
            continue
        num_cols = len(r)
        non_empty = sum(1 for c in r if c != '')
        nonempty_ratio = non_empty / num_cols if num_cols else 0
        label_score = sum(1 for c in r if is_label_like(c)) / num_cols if num_cols else 0
        # line weight: higher weight for earlier lines
        line_weight = 1.0 - (i * 0.05)  # adjust factor if wanted
        # confidence = (nonempty_ratio * 0.6) + (label_score * 0.3) + (line_weight * 0.1)
        confidence = (nonempty_ratio * 0.6) +  (line_weight * 0.3) + (label_score * 0.1)
        candidates.append({
            'line_no': i,
            'row': r,
            'num_columns': num_cols,
            'nonempty_ratio': nonempty_ratio,
            'label_score': label_score,
            'confidence': confidence
        })
    if not candidates:
        return {}
    # choose best by confidence
    best = max(candidates, key=lambda x: x['confidence'])
    return best

# ---------- sampling-based expected column ----------
def sample_expected_columns(rows: List[List[Any]], header_index: int = 0, sample_size: int = 20) -> Tuple[int, float]:
    """
    Count columns across a sample of rows and return (expected_columns, sampling_confidence)
    sampling_confidence = fraction of sample rows that match the mode column count AND have low missing ratio
    """
    sample_rows = []
    start = header_index + 1 if header_index is not None else 0
    for r in rows[start:start + sample_size]:
        sr = normalize_row(r, trim_trailing_empty=True)
        sample_rows.append(sr)

    if not sample_rows:
        return 0, 0.0

    col_counts = [len(r) for r in sample_rows]
    counter = Counter(col_counts)
    mode_count, freq = counter.most_common(1)[0]
    # confidence compute: proportion of sample rows that equal mode AND have < 50% missing
    good = 0
    for r in sample_rows:
        if len(r) == mode_count:
            non_empty = sum(1 for c in r if c != '')
            if mode_count == 0:
                continue
            if (non_empty / mode_count) >= 0.5:
                good += 1
    sampling_confidence = good / len(sample_rows)
    return mode_count, sampling_confidence

# ---------- detect_empty_columns ----------
def detect_empty_columns(rows: List[List[Any]], header_index: int = 0) -> List[int]:
    """
    คืน index ของ column ที่เป็นค่าว่างทั้งหมด (ไม่รวม header)
    """
    # normalize ทุก row
    norm_rows = [normalize_row(r, trim_trailing_empty=False) for r in rows[header_index+1:]]
    
    if not norm_rows:
        return []

    # transpose
    cols = list(zip(*norm_rows))
    
    empty_col_idx = []
    for i, col in enumerate(cols):
        if all(c == '' for c in col):
            empty_col_idx.append(i)
    return empty_col_idx
# ---------- detect extra rows ----------
def detect_extra_rows(rows: List[List[Any]], expected_cols: int, trim_trailing_empty: bool = True):
    extra = []
    for i, raw in enumerate(rows, start=1):
        r = normalize_row(raw, trim_trailing_empty=trim_trailing_empty)
        if len(r) > expected_cols:
            extra_vals = r[expected_cols:]
            # ignore all-empty extras
            if any(v != '' for v in extra_vals):
                extra.append({
                    'row_number': i,
                    'cols': len(r),
                    'extra_values': extra_vals
                })
    return extra
# ---------- [End] detect extra rows ----------
def validate_resource_data(resource,data):
    '''Performs a validation of a resource data, given the resource metadata.

    :param resource: CKAN resource data_dict.

    :returns: `dict`, a validation report for the resource data.
    '''
    log.debug(u'Validating resource {}'.format(resource['id']))

    options = toolkit.config.get(
        u'ckanext.validation.default_validation_options')
    if options:
        options = json.loads(options)
    else:
        options = {}
    resource_options = resource.get(u'validation_options')
    if resource_options and isinstance(resource_options, string_types):#basestring):
        resource_options = json.loads(resource_options)
    if resource_options:
        options.update(resource_options)

    dataset = toolkit.get_action('package_show')(
        {'ignore_auth': True}, {'id': resource['package_id']})

    source = None
    if resource.get(u'url_type') == u'upload':
        upload = uploader.get_resource_uploader(resource)
        if isinstance(upload, uploader.ResourceUpload):
            source = upload.get_path(resource[u'id'])
        else:
            # Upload is not the default implementation (ie it's a cloud storage
            # implementation)
            pass_auth_header = toolkit.asbool(
                toolkit.config.get(u'ckanext.validation.pass_auth_header',
                                   True))
            if dataset[u'private'] and pass_auth_header:
                s = requests.Session()
                s.headers.update({
                    u'Authorization': t.config.get(
                        u'ckanext.validation.pass_auth_header_value',
                        _get_site_user_api_key())
                })

                options[u'http_session'] = s
    if not source:
        source = resource[u'url']

    schema = resource.get(u'schema')
    log.debug('---data schema1--')
    log.debug(schema)
    if schema and isinstance(schema, string_types):#basestring):
        if schema.startswith('http'):
            r = requests.get(schema)
            schema = r.json()
        else:
            schema = json.loads(schema)
    _format = resource[u'format'].lower()
    #----- check mimetype -----------------------------
    # log.debug('--validate: check mimetype--')
    # log.debug('---data schema2--')
    # log.debug(schema)
    # log.debug(data['fields'])
    #------------------------
    mimetype = data['mimetype']
    if (mimetype == 'application/json'):
        report = _validate_table(source, _format=_format, schema=schema, **options)
    else:
        log.debug('------raw_data------')
        raw_data = data['raw_data']
        headers = raw_data[0]
        rows = raw_data[1:]
        #  check duplicate header
        duplicates = [item for item, count in Counter(headers).items() if count > 1]

        #  แปลงเป็น list of dict
        records = [dict(zip(headers, row)) for row in rows]
        # log.debug(raw_data)
        #[old]check extra values ------------------
        # extra_result = detect_columns_and_validate(records)
        # extra_rows = extra_result.get("extra_rows", [])
        # log.debug('----extra_rows----')
        # log.debug(extra_result)
        #[new]check extra values ---------------

        # 1) header candidate
        top_n_header=10
        sample_size=20
        trim_trailing=True

        header_info = analyze_header_candidates(raw_data, top_n=top_n_header)
        header_idx = header_info.get('line_no') if header_info else 0
        header_row = header_info.get('row') if header_info else raw_data[0]
        header_conf = header_info.get('confidence', 0)

        # 4) คำนวณ sampling expected cols
        mode_cols, sampling_conf = sample_expected_columns(raw_data, header_index=header_idx, sample_size=sample_size)

        # 5) ตัดสินใจเลือก expected cols
        while header_row and (header_row[-1] is None or str(header_row[-1]).strip() == ''):
                header_row.pop()
                
        extra_by_column = []
        if sampling_conf < 0.5:
            empty_col_idx = detect_empty_columns(raw_data, header_index=header_idx)
            log.debug('empty_col_idx:', empty_col_idx)

            if empty_col_idx:
                # กำหนด expected_cols = column ก่อน column ว่างแรก
                expected_cols = min(empty_col_idx)
            else:
                # ถ้าไม่มี column ว่าง ให้ใช้ header length
                expected_cols = len(header_row)

            chosen_from = 'header (sampling low)'
            
        
        elif header_conf >= sampling_conf + 0.15:
            expected_cols = len(header_row)
            chosen_from = 'header (conf higher)'
        else:
            expected_cols = mode_cols
            chosen_from = 'sampling'

        # 6) ตรวจ extra rows
        extra_rows = detect_extra_rows(raw_data, expected_cols, trim_trailing_empty=trim_trailing)
        # header_info = analyze_header_candidates(raw_data, top_n=top_n_header)
        # header_idx = header_info.get('line_no') if header_info else None
        # header_row = header_info.get('row') if header_info else None
        # header_conf = header_info.get('confidence', 0)

        # # 2) เลือกเฉพาะ rows หลัง header
        # rows_for_extra = raw_data[header_idx + 1:]

        # # 3) clean rows
        # clean_rows = [r for r in rows_for_extra if any((c is not None and str(c).strip() != '') for c in r)]
        # if not clean_rows:
        #     return {'error': 'no data', 'rows': 0}

        # # 4) คำนวณ sampling expected cols
        # mode_cols, sampling_conf = sample_expected_columns(rows_for_extra, header_index=header_idx, sample_size=sample_size)

        # # 5) ตัดสินใจเลือก expected cols
        # if sampling_conf < 0.5:
        #     expected_cols = len(header_row)
        #     chosen_from = 'header (sampling low)'
        # elif header_conf >= sampling_conf + 0.15:
        #     expected_cols = len(header_row)
        #     chosen_from = 'header (conf higher)'
        # else:
        #     expected_cols = mode_cols
        #     chosen_from = 'sampling'

        # # 6) ตรวจ extra rows
        # extra_rows = detect_extra_rows(rows_for_extra, expected_cols, trim_trailing_empty=trim_trailing)
        log.debug(header_info)
        log.debug(chosen_from)
        log.debug(extra_rows)
        #---------------------------------------
        report = validate_from_records(records)

        #--- เพิ่ม error เข้า report ถ้าพบคอลัมน์ซ้ำ
        if duplicates:
            log.debug("duplicate header: %s", duplicates)
            table = report['tables'][0]  # สมมุติว่าเรามีแค่ table เดียว
            table['valid'] = False
            table.setdefault('errors', []).append({
                'code': 'duplicate-header',
                'message': f'Duplicate headers found: {duplicates}',
                'message-data': {'duplicates': duplicates},
            })
        else:
            log.debug("no duplicate header")
        # --- ตรวจสอบ extra values
        if extra_rows:
            log.debug("--พบ extra value ที่ rows--: %s", [r['row_number'] for r in extra_rows])
            table = report['tables'][0]  # สมมุติว่า table เดียว
            table['valid'] = False
            table.setdefault('errors', []).append({
                'code': 'extra-value',
                'message': f'Extra values found in rows: {[r["row_number"] for r in extra_rows]}',
                'message-data': {
                    'header_confidence': round(header_conf, 3),
                    'sampling_confidence': round(sampling_conf, 3),
                    'expected_columns': expected_cols,
                    'extra_rows': extra_rows
                }
            })
        else:
            log.debug("--no extra value--")
    #----------------------------
    # report = _validate_table(source, _format=_format, schema=schema, **options)
    log.debug(report)
    # Hide uploaded files
    for table in report.get('tables', []):
        if table['source'].startswith('/'):
            table['source'] = resource['url']
    for index, warning in enumerate(report.get('warnings', [])):
        report['warnings'][index] = re.sub(r'Table ".*"', 'Table', warning)

    return report


def _validate_table(source, _format=u'csv', schema=None, **options):
    report = validate(source, format=_format, schema=schema, **options)
    return report
# add by Pang
def validate_from_records(records, _format='csv', schema=None, **options):
    raw_headers = list(records[0].keys())
    headers = []
    for h in raw_headers:
        # ดักกรณี header เพี้ยน
        if h is None or str(h).lower().startswith("none"):
            headers.append("")  # ให้เป็น header ว่างจริง
        else:
            headers.append(h)

    rows = [list(r.values()) for r in records]

    tmp = tempfile.NamedTemporaryFile(mode='w+', newline='', suffix='.csv', delete=False)
    try:
        writer = csv.writer(tmp)
        writer.writerow(headers)
        for row in rows:
            # ensure row length == len(headers)
            if len(row) < len(headers):
                row += [''] * (len(headers) - len(row))
            elif len(row) > len(headers):
                row = row[:len(headers)]
            writer.writerow(row)   
        tmp.flush()
        tmp_path = tmp.name
        report = validate(tmp_path, format=_format, schema=schema, **options)
    finally:
        tmp.close()              # ปิด handle ก่อน
        os.remove(tmp.name)      # ลบไฟล์เอง
    return report

# def validate_from_records(records, _format='csv', schema=None, **options):
#     # ---- 1) สร้าง header ----
#     # headers = list(records[0].keys())  # ต้องเป็น key ที่ตรงกับไฟล์จริง
#     raw_headers = list(records[0].keys())
#     headers = []
#     for h in raw_headers:
#         # ถ้า key เป็น None หรือเป็นชื่อที่ auto-gen เช่น None_1
#         if h is None or str(h).lower().startswith("none"):
#             headers.append("")  # header ว่างจริง
#         else:
#             headers.append(h)
#     rows = [list(r.values()) for r in records]

#     # ---- 2) เขียนลง temp CSV ----
#     with tempfile.NamedTemporaryFile(mode='w+', newline='', suffix='.csv', delete=False) as tmp:
#         writer = csv.writer(tmp)
#         writer.writerow(headers)
#         for row in rows:
#             # ensure row length == len(headers)
#             if len(row) < len(headers):
#                 row += [''] * (len(headers) - len(row))
#             elif len(row) > len(headers):
#                 row = row[:len(headers)]
#             writer.writerow(row)
#         tmp_path = tmp.name

#     # ---- 3) ใช้ goodtables.validate ----
#     report = validate(tmp_path, format=_format, schema=schema, **options)
#     return report

def _validate_table_by_list(records):
    report = validate(records, format='inline')
    return report
def _get_site_user_api_key():
    site_user_name = toolkit.get_action('get_site_user')({
        'ignore_auth': True
    }, {})
    site_user = toolkit.get_action('get_site_user')(
        {'ignore_auth': True}, {'id': site_user_name})
    return site_user['apikey']
